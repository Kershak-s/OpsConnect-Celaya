"""
Sistema de descarga Excel para PAE PAPA
Incluye todos los campos A-P con validación de colores
"""

from flask import request, jsonify, send_file
from datetime import datetime, date
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import json

# Colores para Excel
COLORS = {
    'success': 'C6EFCE',  # Verde claro
    'warning': 'FFEB9C',  # Amarillo claro
    'danger': 'FFC7CE',   # Rojo claro
    'empty': 'F2F2F2',    # Gris claro
    'header': '4F81BD'    # Azul header
}

# Rangos PAPA según tabla de especificaciones (sincronizado con JS frontend)
RANGOS_PAPA = {
    'A': {'verde': (0, 4), 'amarillo': (4.1, 10), 'descripcion': 'Defectos de color'},
    'B': {'verde': (0, 4), 'amarillo': (4.1, 10), 'descripcion': 'Daño seco'},
    'C': {'verde': (0, 4), 'amarillo': (4.1, 10), 'descripcion': 'Color indeseable'},
    'D': {'verde': (0, 10), 'amarillo': (10.1, 20), 'descripcion': 'Defectos internos papa'},
    'E': {'verde': (0, 10), 'amarillo': (10.1, 20), 'descripcion': 'Defectos externos papa'},
    'F': {'verde': (0, 10), 'amarillo': (10.1, 20), 'descripcion': 'Defectos totales de papa'},
    'G': {'verde': (0, 1), 'amarillo': (1.1, 2), 'descripcion': 'Centros suaves + clusters'},
    'H': {'verde': (0, 6), 'amarillo': (6.1, 20), 'descripcion': 'Exceso de cáscara'},
    'I': {'verde': (0, 6), 'amarillo': (6.1, 20), 'descripcion': 'Hojuelas aceitosas'},
    'J': {'verde': (0, 6), 'amarillo': (6.1, 20), 'descripcion': 'Ampulas'},
    'K': {'verde': (0, 6), 'amarillo': (6.1, 20), 'descripcion': 'Puntos obscuros'},
    'L': {'verde': (0, 20), 'amarillo': (20.1, 40), 'descripcion': 'Defectos totales de proceso'},
    'M': {'verde': (0, 30), 'amarillo': (30.1, 35), 'descripcion': 'Hojuelas dobladas'},
    'N': {'verde': (75, 100), 'amarillo': None, 'descripcion': 'Hojuela Entera'},  # Verde: 75-100%, Rojo: <75%, Sin amarillo
    'O': {'verde': (73, 100), 'amarillo': None, 'descripcion': 'Hojuela Entera (FIESTA)'},  # Verde: 73-100%, Rojo: <73%
    'P': {'verde': (0, 12), 'amarillo': (12.1, 15), 'descripcion': 'Pedacera (scrap)'},
    'Q': {'verde': (3, 100), 'amarillo': None, 'descripcion': 'Color de la Base L'},  # Verde: ≥3, Rojo: <3
    'R': {'verde': (-3, 2.5), 'amarillo': (2.51, 10), 'descripcion': 'Color de la base a'},
    # Campos 4H - Registros cada 4 horas
    '4H-PELADO': {'verde': (95, 100), 'amarillo': None, 'descripcion': 'Pelado - Remoción de cáscara', 'unidad': '%'},
    '4H-GROSOR': {'verde': (1.3, 1.4), 'amarillo': None, 'descripcion': 'Rebanado - Grosor', 'unidad': 'mm'},
    '4H-DESVIACION': {'verde': (0, 0.071), 'amarillo': None, 'descripcion': 'Rebanado - Desviación', 'unidad': 'mm'},
    '4H-ALMIDON': {'verde': (0, 3), 'amarillo': None, 'descripcion': 'Lavador - Almidón superficial', 'unidad': ''},
    '4H-HUMEDAD': {'verde': (0, 9.99), 'amarillo': None, 'descripcion': 'Lavador - Humedad superficial', 'unidad': '%'},
    '4H-TIEMPO': {'verde': (2.83, 3.16), 'amarillo': None, 'descripcion': 'Freidor - Tiempo de residencia', 'unidad': 'min'},
    '4H-TEMP': {'verde': (171, 9999), 'amarillo': None, 'descripcion': 'Freidor - Temperatura entrada', 'unidad': '°C'},
    '4H-OV': {'verde': (0, 15), 'amarillo': (15.1, 25), 'descripcion': 'Freidor - OV', 'unidad': 'ppm'},
    '4H-AGL': {'verde': (0, 0.25), 'amarillo': (0.26, 0.35), 'descripcion': 'Freidor - AGL', 'unidad': '%'}
}

def setup_papa_excel_routes(app):
    """Configura las rutas de descarga Excel para PAPA"""
    
    @app.route('/excel-papa')
    def excel_papa():
        """Descarga Excel con datos PAE PAPA"""
        
        from models import PAERegistro  # Importar aquí para evitar import circular

        # Parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        turno = request.args.get('turno', 'all')
        producto = request.args.get('producto', 'all')
        incluir_rangos = request.args.get('incluir_rangos', 'true').lower() == 'true'

        # Validar fechas
        if not fecha_inicio or not fecha_fin:
            return jsonify({'error': 'Fechas de inicio y fin requeridas'}), 400

        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de fecha inválido (YYYY-MM-DD)'}), 400

        # Construir query para PAPA
        query = PAERegistro.query.filter(
            PAERegistro.categoria == 'PAPA',
            PAERegistro.fecha >= fecha_inicio,
            PAERegistro.fecha <= fecha_fin
        )

        if turno != 'all':
            query = query.filter(PAERegistro.turno == turno)

        if producto != 'all':
            query = query.filter(PAERegistro.producto == producto)

        # Obtener registros
        registros = query.order_by(PAERegistro.fecha.desc(), PAERegistro.hora.desc()).all()
        
        if not registros:
            return jsonify({'error': 'No hay datos PAE PAPA para exportar en el rango seleccionado'}), 404
        
        # Crear Excel
        buffer = BytesIO()
        
        try:
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Hoja 1: Datos PAE PAPA
                df_datos = crear_dataframe_papa(registros)
                df_datos.to_excel(writer, sheet_name='PAE PAPA', index=False)
                aplicar_formato_papa(writer, 'PAE PAPA', df_datos, registros)
                
                # Hoja 2: Estadísticas por campo
                df_stats = crear_estadisticas_papa(registros)
                df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
                
                # Hoja 3: Rangos de referencia
                if incluir_rangos:
                    df_rangos = crear_rangos_papa()
                    df_rangos.to_excel(writer, sheet_name='Rangos PAPA', index=False)
                
                # Hoja 4: Resumen de conformidad
                df_conformidad = crear_resumen_conformidad(registros)
                df_conformidad.to_excel(writer, sheet_name='Conformidad', index=False)
            
            buffer.seek(0)
            filename = f'PAE_PAPA_{fecha_inicio}_{fecha_fin}.xlsx'
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            print(f"Error generando Excel PAPA: {str(e)}")
            return jsonify({'error': f'Error generando archivo Excel: {str(e)}'}), 500

def crear_dataframe_papa(registros):
    """Crea DataFrame con datos PAE PAPA completos"""

    data = []
    for registro in registros:
        # Parsear datos JSON si existen
        atributos = {}
        if hasattr(registro, 'data') and registro.data:
            try:
                atributos = json.loads(registro.data)

                # Calcular porcentajes retroactivos si no existen
                # Esto asegura compatibilidad con registros antiguos
                # NOTA: N, O, P ya son porcentajes, no se calculan
                campos_porcentaje_directo = ['N', 'O', 'P']  # Estos campos ya contienen porcentajes

                for campo in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
                    campo_porcentaje = f'{campo}_porcentaje'
                    if campo_porcentaje not in atributos and campo in atributos:
                        valor_campo = atributos.get(campo, '')
                        if valor_campo and str(valor_campo).strip():
                            try:
                                valor = float(valor_campo)

                                if campo in campos_porcentaje_directo:
                                    # N, O, P: el valor YA ES el porcentaje
                                    porcentaje = valor
                                else:
                                    # A-M: calcular porcentaje (gramos/200*100)
                                    porcentaje = ((valor / 200) * 100)

                                atributos[campo_porcentaje] = f"{porcentaje:.2f}"
                            except (ValueError, ZeroDivisionError):
                                atributos[campo_porcentaje] = ''

            except:
                pass
        
        row = {
            # Información básica
            'ID': registro.id,
            'Folio': getattr(registro, 'folio', ''),
            'Fecha': registro.fecha.strftime('%d/%m/%Y'),
            'Hora Registro': registro.hora.strftime('%H:%M:%S') if hasattr(registro, 'hora') and registro.hora else '',
            'Hora Bloque': getattr(registro, 'hora_bloque', ''),
            'Hora Muestreo': registro.hora_muestreo.strftime('%H:%M:%S') if hasattr(registro, 'hora_muestreo') and registro.hora_muestreo else '',
            'Turno': registro.turno,
            'Producto': registro.producto,
            'Categoría': registro.categoria,

            # DEFECTOS MATERIA PRIMA - Gramos y Porcentajes
            'A - Defectos de color (g)': atributos.get('A', ''),
            'A - Defectos de color (%)': atributos.get('A_porcentaje', ''),
            'B - Daño seco (g)': atributos.get('B', ''),
            'B - Daño seco (%)': atributos.get('B_porcentaje', ''),
            'C - Color indeseable (g)': atributos.get('C', ''),
            'C - Color indeseable (%)': atributos.get('C_porcentaje', ''),
            'D - Defectos internos papa (g)': atributos.get('D', ''),
            'D - Defectos internos papa (%)': atributos.get('D_porcentaje', ''),
            'E - Defectos externos papa (g)': atributos.get('E', ''),
            'E - Defectos externos papa (%)': atributos.get('E_porcentaje', ''),
            'F - Defectos totales de papa (g)': atributos.get('F', ''),
            'F - Defectos totales de papa (%)': atributos.get('F_porcentaje', ''),

            # DEFECTOS DE PROCESO - Gramos y Porcentajes
            'G - Centros suaves + clusters (g)': atributos.get('G', ''),
            'G - Centros suaves + clusters (%)': atributos.get('G_porcentaje', ''),
            'H - Exceso de cáscara (g)': atributos.get('H', ''),
            'H - Exceso de cáscara (%)': atributos.get('H_porcentaje', ''),
            'I - Hojuelas aceitosas (g)': atributos.get('I', ''),
            'I - Hojuelas aceitosas (%)': atributos.get('I_porcentaje', ''),
            'J - Ampulas (g)': atributos.get('J', ''),
            'J - Ampulas (%)': atributos.get('J_porcentaje', ''),
            'K - Puntos obscuros (g)': atributos.get('K', ''),
            'K - Puntos obscuros (%)': atributos.get('K_porcentaje', ''),
            'L - Defectos totales proceso (g)': atributos.get('L', ''),
            'L - Defectos totales proceso (%)': atributos.get('L_porcentaje', ''),
            'M - Hojuelas dobladas (g)': atributos.get('M', ''),
            'M - Hojuelas dobladas (%)': atributos.get('M_porcentaje', ''),

            # CALIDAD DEL PRODUCTO - Gramos y Porcentajes
            'N - Hojuela Entera (g)': atributos.get('N', ''),
            'N - Hojuela Entera (%)': atributos.get('N_porcentaje', ''),
            'O - Hojuela Entera FIESTA (g)': atributos.get('O', ''),
            'O - Hojuela Entera FIESTA (%)': atributos.get('O_porcentaje', ''),
            'P - Peladeras/Scrap (g)': atributos.get('P', ''),
            'P - Peladeras/Scrap (%)': atributos.get('P_porcentaje', ''),

            # COLOR DE LA BASE
            'Q - Color de la Base L': atributos.get('Q', ''),
            'R - Color de la base a': atributos.get('R', ''),

            # REGISTROS CADA 4 HORAS
            '4H - Pelado Remoción (%)': atributos.get('4H-PELADO', ''),
            '4H - Rebanado Grosor (mm)': atributos.get('4H-GROSOR', ''),
            '4H - Rebanado Desviación (mm)': atributos.get('4H-DESVIACION', ''),
            '4H - Lavador Almidón': atributos.get('4H-ALMIDON', ''),
            '4H - Lavador Humedad (%)': atributos.get('4H-HUMEDAD', ''),
            '4H - Freidor Tiempo (min)': atributos.get('4H-TIEMPO', ''),
            '4H - Freidor Temp Entrada (°C)': atributos.get('4H-TEMP', ''),
            '4H - Freidor OV (ppm)': atributos.get('4H-OV', ''),
            '4H - Freidor AGL (%)': atributos.get('4H-AGL', ''),

            # SENSORIAL
            'Sensorial - Apariencia': getattr(registro, 'sensorial_apariencia', ''),
            'Sensorial - Apariencia Comentario': getattr(registro, 'sensorial_apariencia_comentario', ''),
            'Sensorial - Textura': getattr(registro, 'sensorial_textura', ''),
            'Sensorial - Textura Comentario': getattr(registro, 'sensorial_textura_comentario', ''),
            'Sensorial - Sabor': getattr(registro, 'sensorial_sabor', ''),
            'Sensorial - Sabor Comentario': getattr(registro, 'sensorial_sabor_comentario', ''),

            # SECCIÓN ROTURA (solo para PAPA)
            'Rotura - Aplica': 'Sí' if getattr(registro, 'rotura_aplica', False) else 'No',
            'Rotura - Hojuela Entera (%)': getattr(registro, 'hojuela_entera', ''),
            'Rotura - Hojuela Entera FIESTA (%)': getattr(registro, 'hojuela_entera_fiesta', ''),
            'Rotura - Peladeras/Scrap (%)': getattr(registro, 'peladeras_scrap', ''),
            'Rotura - Observaciones': getattr(registro, 'rotura_observaciones', ''),

            # OBSERVACIONES GENERALES
            'Observaciones': getattr(registro, 'observaciones', ''),

            # AUDITORÍA
            'Creado por (ID)': getattr(registro, 'created_by', ''),
            'Fecha Creación': registro.created_at.strftime('%d/%m/%Y %H:%M:%S') if hasattr(registro, 'created_at') and registro.created_at else ''
        }
        data.append(row)
    
    return pd.DataFrame(data)

def crear_estadisticas_papa(registros):
    """Crea estadísticas por campo PAE PAPA"""

    campos_numericos = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                        '4H-PELADO', '4H-GROSOR', '4H-DESVIACION', '4H-ALMIDON', '4H-HUMEDAD',
                        '4H-TIEMPO', '4H-TEMP', '4H-OV', '4H-AGL']
    stats_data = []
    
    for campo in campos_numericos:
        valores = []
        total_registros = 0
        conformes = 0
        no_conformes = 0
        
        for registro in registros:
            total_registros += 1
            if hasattr(registro, 'data') and registro.data:
                try:
                    atributos = json.loads(registro.data)
                    valor_str = atributos.get(campo, '')
                    if valor_str and valor_str.strip():
                        valor = float(valor_str)
                        valores.append(valor)
                        
                        # Determinar conformidad
                        color = determinar_color_papa(valor, campo)
                        if color == 'success':
                            conformes += 1
                        else:
                            no_conformes += 1
                except:
                    pass
        
        # Calcular estadísticas
        if valores:
            promedio = round(sum(valores) / len(valores), 3)
            minimo = round(min(valores), 3)
            maximo = round(max(valores), 3)
            
            # Calcular varianza y desviación estándar
            if len(valores) > 1:
                varianza = sum((x - promedio) ** 2 for x in valores) / (len(valores) - 1)
                desv_std = round(varianza ** 0.5, 3)
            else:
                desv_std = 0
        else:
            promedio = minimo = maximo = desv_std = 'N/A'
        
        stats_data.append({
            'Campo': f'{campo} - {RANGOS_PAPA[campo]["descripcion"]}',
            'Total Muestras': len(valores),
            'Promedio': promedio,
            'Mínimo': minimo,
            'Máximo': maximo,
            'Desv. Estándar': desv_std,
            'Conformes': conformes,
            'No Conformes': no_conformes,
            '% Conformidad': round((conformes / len(valores) * 100), 1) if valores else 'N/A',
            'Rango Verde': f"{RANGOS_PAPA[campo]['verde'][0]} - {RANGOS_PAPA[campo]['verde'][1]}",
            'Rango Amarillo': (f"{RANGOS_PAPA[campo]['amarillo'][0]} - {RANGOS_PAPA[campo]['amarillo'][1]}"
                              if RANGOS_PAPA[campo]['amarillo'] else 'N/A')
        })
    
    return pd.DataFrame(stats_data)

def crear_rangos_papa():
    """Crea DataFrame con rangos de referencia PAPA"""

    rangos_data = []
    for campo, info in RANGOS_PAPA.items():
        amarillo_info = info.get('amarillo')

        # Determinar unidad
        if campo in ['A','B','C','D','E','F','G','H','I','J','K','L','M']:
            unidad = 'g'
        elif campo.startswith('4H-'):
            unidad = info.get('unidad', '')
        else:
            unidad = '%'

        rangos_data.append({
            'Campo': campo,
            'Descripción': info['descripcion'],
            'Rango Verde (Aceptable)': f"{info['verde'][0]} - {info['verde'][1]}",
            'Rango Amarillo (Advertencia)': (f"{amarillo_info[0]} - {amarillo_info[1]}" if amarillo_info else 'N/A'),
            'Rojo (No Conforme)': (f"> {amarillo_info[1]}" if amarillo_info else f"< {info['verde'][0]}"),
            'Unidad': unidad
        })

    return pd.DataFrame(rangos_data)

def crear_resumen_conformidad(registros):
    """Crea resumen de conformidad por producto y turno"""
    
    resumen = {}
    
    for registro in registros:
        key = f"{registro.producto}_{registro.turno}"
        
        if key not in resumen:
            resumen[key] = {
                'Producto': registro.producto,
                'Turno': registro.turno,
                'Total Registros': 0,
                'Registros Conformes': 0,
                'Registros No Conformes': 0
            }
        
        resumen[key]['Total Registros'] += 1
        
        # Evaluar conformidad del registro
        es_conforme = True
        if hasattr(registro, 'data') and registro.data:
            try:
                atributos = json.loads(registro.data)
                campos_a_evaluar = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R',
                                   '4H-PELADO', '4H-GROSOR', '4H-DESVIACION', '4H-ALMIDON', '4H-HUMEDAD',
                                   '4H-TIEMPO', '4H-TEMP', '4H-OV', '4H-AGL']
                for campo in campos_a_evaluar:
                    valor_str = atributos.get(campo, '')
                    if valor_str and valor_str.strip():
                        valor = float(valor_str)
                        color = determinar_color_papa(valor, campo)
                        if color == 'danger':  # Rojo = No conforme
                            es_conforme = False
                            break
            except:
                pass
        
        if es_conforme:
            resumen[key]['Registros Conformes'] += 1
        else:
            resumen[key]['Registros No Conformes'] += 1
    
    # Calcular porcentajes
    conformidad_data = []
    for data in resumen.values():
        total = data['Total Registros']
        if total > 0:
            conformidad = round((data['Registros Conformes'] / total) * 100, 1)
        else:
            conformidad = 0
        
        conformidad_data.append({
            'Producto': data['Producto'],
            'Turno': data['Turno'],
            'Total Registros': total,
            'Registros Conformes': data['Registros Conformes'],
            'Registros No Conformes': data['Registros No Conformes'],
            '% Conformidad': conformidad
        })
    
    return pd.DataFrame(conformidad_data)

def aplicar_formato_papa(writer, sheet_name, df, registros):
    """Aplica formato con colores a la hoja PAPA"""
    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Formato headers
    header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Aplicar colores a datos
    for row_idx, registro in enumerate(registros, 2):
        aplicar_colores_fila_papa(worksheet, row_idx, registro)
    
    # Ajustar anchos de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        # Columnas de porcentaje más angostas (10), gramos normales (15), headers amplios (30)
        if '(%)' in str(column[0].value):
            adjusted_width = 12  # Columnas de porcentaje más compactas
        elif '(g)' in str(column[0].value):
            adjusted_width = 18  # Columnas de gramos
        else:
            adjusted_width = min(max_length + 2, 30)  # Otras columnas

        worksheet.column_dimensions[column_letter].width = adjusted_width

def aplicar_colores_fila_papa(worksheet, row, registro):
    """Aplica colores a una fila específica según rangos PAPA"""
    
    if not hasattr(registro, 'data') or not registro.data:
        return
    
    try:
        atributos = json.loads(registro.data)
    except:
        return
    
    # Mapeo columnas - Ahora con gramos (g) y porcentajes (%)
    # Formato: 'Campo': (columna_gramos, columna_porcentaje) o 'Campo': columna_unica
    # Columnas: ID(1), Folio(2), Fecha(3), Hora Registro(4), Hora Bloque(5), Hora Muestreo(6), Turno(7), Producto(8), Categoría(9)
    # Luego campos A-P desde columna 10
    campos_columnas_dobles = {
        'A': (10, 11),  # A(g), A(%)
        'B': (12, 13),  # B(g), B(%)
        'C': (14, 15),  # C(g), C(%)
        'D': (16, 17),  # D(g), D(%)
        'E': (18, 19),  # E(g), E(%)
        'F': (20, 21),  # F(g), F(%)
        'G': (22, 23),  # G(g), G(%)
        'H': (24, 25),  # H(g), H(%)
        'I': (26, 27),  # I(g), I(%)
        'J': (28, 29),  # J(g), J(%)
        'K': (30, 31),  # K(g), K(%)
        'L': (32, 33),  # L(g), L(%)
        'M': (34, 35),  # M(g), M(%)
        'N': (36, 37),  # N(g), N(%)
        'O': (38, 39),  # O(g), O(%)
        'P': (40, 41),  # P(g), P(%)
    }

    # Campos con una sola columna (Q, R, y 4H)
    campos_columnas_simples = {
        'Q': 42,  # Q - Color de la Base L
        'R': 43,  # R - Color de la base a
        '4H-PELADO': 44,
        '4H-GROSOR': 45,
        '4H-DESVIACION': 46,
        '4H-ALMIDON': 47,
        '4H-HUMEDAD': 48,
        '4H-TIEMPO': 49,
        '4H-TEMP': 50,
        '4H-OV': 51,
        '4H-AGL': 52,
    }

    # Aplicar colores a campos con columnas dobles (A-P)
    for campo, (col_gramos, col_porcentaje) in campos_columnas_dobles.items():
        valor_str = atributos.get(campo, '')
        if valor_str and valor_str.strip():
            try:
                valor = float(valor_str)
                color_class = determinar_color_papa(valor, campo)

                # Aplicar color a la columna de gramos
                cell_gramos = worksheet.cell(row=row, column=col_gramos)
                cell_gramos.fill = PatternFill(start_color=COLORS[color_class], end_color=COLORS[color_class], fill_type='solid')

                # Aplicar color a la columna de porcentaje
                cell_porcentaje = worksheet.cell(row=row, column=col_porcentaje)
                cell_porcentaje.fill = PatternFill(start_color=COLORS[color_class], end_color=COLORS[color_class], fill_type='solid')

                # Texto en negrita para valores fuera de rango (en ambas columnas)
                if color_class == 'danger':
                    cell_gramos.font = Font(bold=True)
                    cell_porcentaje.font = Font(bold=True)

            except ValueError:
                pass

    # Aplicar colores a campos simples (Q, R, 4H)
    for campo, col in campos_columnas_simples.items():
        valor_str = atributos.get(campo, '')
        if valor_str and valor_str.strip():
            try:
                valor = float(valor_str)
                color_class = determinar_color_papa(valor, campo)

                # Aplicar color
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=COLORS[color_class], end_color=COLORS[color_class], fill_type='solid')

                # Texto en negrita para valores fuera de rango
                if color_class == 'danger':
                    cell.font = Font(bold=True)

            except ValueError:
                pass

def determinar_color_papa(valor, campo):
    """Determina color según rangos PAPA"""

    if campo not in RANGOS_PAPA:
        return 'empty'

    rango = RANGOS_PAPA[campo]
    verde_min, verde_max = rango['verde']
    amarillo_info = rango.get('amarillo')  # Puede ser None

    # Campo N: Hojuela Entera
    # Verde: 75-100%, Rojo: <75%
    if campo == 'N':
        if 75 <= valor <= 100:
            return 'success'  # Verde: 75-100%
        return 'danger'       # Rojo: <75%

    # Campo O: Hojuela Entera FIESTA
    # Verde: 73-100%, Rojo: <73%
    if campo == 'O':
        if 73 <= valor <= 100:
            return 'success'  # Verde: 73-100%
        return 'danger'       # Rojo: <73%

    # Campo Q: Color de la Base L
    # Verde: ≥3, Rojo: <3
    if campo == 'Q':
        if valor >= 3:
            return 'success'  # Verde: ≥3
        return 'danger'       # Rojo: <3

    # Caso especial para campo R (puede ser negativo)
    if campo == 'R':
        if verde_min <= valor <= verde_max:
            return 'success'
        if amarillo_info and ((amarillo_info[0] <= valor <= amarillo_info[1]) or valor < -3):
            return 'warning'
        return 'danger'

    # Lógica estándar para campos A-M y P
    # Verde: si está en rango verde
    if verde_min <= valor <= verde_max:
        return 'success'

    # Amarillo: si está en rango amarillo (si existe)
    if amarillo_info:
        amarillo_min, amarillo_max = amarillo_info
        if amarillo_min <= valor <= amarillo_max:
            return 'warning'

    # Rojo: si está fuera de ambos rangos
    return 'danger'
