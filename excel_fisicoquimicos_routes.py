"""
Rutas y lógica para descarga Excel de análisis fisicoquímicos
Incluye campos de Producto Terminado (PT)
"""

from flask import request, jsonify, send_file
from datetime import datetime, date
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from models import db, AnalisisCalidad

# Colores para Excel
COLORS = {
    'success': 'C6EFCE',  # Verde claro
    'warning': 'FFEB9C',  # Amarillo claro
    'danger': 'FFC7CE',   # Rojo claro
    'empty': 'F2F2F2',    # Gris claro
    'header': '4F81BD'    # Azul header
}

def setup_excel_fisicoquimicos_routes(app):
    """Configura las rutas de descarga Excel para análisis fisicoquímicos"""
    
    @app.route('/excel-fisicoquimicos/<categoria>')
    def excel_fisicoquimicos_categoria(categoria):
        """Descarga Excel con análisis fisicoquímicos incluyendo campos PT (nueva versión)"""
        
        # Obtener parámetros
        # categoria viene del URL, el resto de query parameters
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        turno = request.args.get('turno', 'all')
        producto = request.args.get('producto', 'all')
        incluir_rangos = request.args.get('incluir_rangos', 'true').lower() == 'true'
        
        # Validar fechas
        if not fecha_inicio or not fecha_fin:
            return jsonify({'error': 'Fechas requeridas'}), 400
        
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'error': 'Formato de fecha inválido'}), 400
        
        # Construir query
        query = AnalisisCalidad.query.filter(
            AnalisisCalidad.categoria == categoria,
            AnalisisCalidad.fecha >= fecha_inicio,
            AnalisisCalidad.fecha <= fecha_fin
        )
        
        # Filtros adicionales
        if turno != 'all':
            query = query.filter(AnalisisCalidad.turno == turno)
        
        if producto != 'all':
            query = query.filter(AnalisisCalidad.producto == producto)
        
        # Obtener datos
        analisis_records = query.order_by(AnalisisCalidad.fecha.desc(), AnalisisCalidad.horario.desc()).all()
        
        if not analisis_records:
            return jsonify({'error': 'No hay datos para exportar'}), 404
        
        # Crear Excel
        buffer = BytesIO()
        
        try:
            # Generar Excel con múltiples hojas
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Hoja 1: Datos principales con colores
                df_datos = crear_dataframe_analisis(analisis_records, categoria)
                df_datos.to_excel(writer, sheet_name='Análisis Fisicoquímicos', index=False)
                
                # Aplicar formato con colores
                aplicar_formato_excel(writer, 'Análisis Fisicoquímicos', df_datos, analisis_records, categoria)
                
                # Hoja 2: Estadísticas
                df_stats = crear_dataframe_estadisticas(analisis_records)
                df_stats.to_excel(writer, sheet_name='Estadísticas', index=False)
                
                # Hoja 3: Rangos de referencia (si se solicita)
                if incluir_rangos:
                    df_rangos = crear_dataframe_rangos(categoria)
                    df_rangos.to_excel(writer, sheet_name='Rangos de Referencia', index=False)
            
            buffer.seek(0)
            
            # Nombre del archivo
            filename = f'analisis_fisicoquimicos_{categoria.lower()}_{fecha_inicio}_{fecha_fin}.xlsx'
            
            return send_file(
                buffer,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            print(f"Error generando Excel: {str(e)}")
            return jsonify({'error': 'Error generando archivo Excel'}), 500

def crear_dataframe_analisis(records, categoria):
    """Crea DataFrame con todos los datos incluyendo campos PT"""

    data = []
    for record in records:
        row = {
            'Folio': record.folio,
            'Fecha': record.fecha.strftime('%d/%m/%Y'),
            'Turno': record.turno,
            'Producto': record.producto,
            'Horario': record.horario,

            # Base Frita
            'Humedad Base Frita': record.humedad_base_frita or '',
            'Aceite Base Frita': record.aceite_base_frita or '',

            # PT Producto Terminado General
            'Aceite PT General': record.aceite_pt_producto_terminado or '',
            'Humedad PT General': record.humedad_pt_producto_terminado or '',
            'Sal PT General': record.sal_pt_producto_terminado or '',

            # Tambor 1
            'Aceite PT T1': record.tanque1_aceite_pt or '',
            'Humedad PT T1': record.tanque1_humedad_pt or '',
            'Sal Titulador T1': record.tanque1_sal_titulador or '',
            'Sal PT T1': record.tanque1_sal_pt or '',

            # Tambor 2
            'Aceite PT T2': record.tanque2_aceite_pt or '',
            'Humedad PT T2': record.tanque2_humedad_pt or '',
            'Sal Titulador T2': record.tanque2_sal_titulador or '',
            'Sal PT T2': record.tanque2_sal_pt or '',

            # Tambor 3
            'Aceite PT T3': record.tanque3_aceite_pt or '',
            'Humedad PT T3': record.tanque3_humedad_pt or '',
            'Sal Titulador T3': record.tanque3_sal_titulador or '',
            'Sal PT T3': record.tanque3_sal_pt or '',
        }

        # Cloruros Base solo para PAPA
        if categoria == 'PAPA':
            row['Cloruros Base'] = record.cloruros_base or ''

        row['Observaciones'] = record.observaciones or ''
        row['Fecha Creación'] = record.created_at.strftime('%d/%m/%Y %H:%M')

        data.append(row)

    return pd.DataFrame(data)

def crear_dataframe_estadisticas(records):
    """Crea DataFrame con estadísticas resumidas"""
    
    productos = {}
    for record in records:
        producto = record.producto
        if producto not in productos:
            productos[producto] = {
                'producto': producto,
                'total_registros': 0,
                'humedad_base': [],
                'aceite_base': [],
                'aceite_pt_general': [],
                'humedad_pt_general': [],
                'sal_pt_general': []
            }
        
        productos[producto]['total_registros'] += 1
        
        # Recopilar valores para estadísticas
        if record.humedad_base_frita:
            try:
                productos[producto]['humedad_base'].append(float(record.humedad_base_frita))
            except:
                pass
                
        if record.aceite_base_frita:
            try:
                productos[producto]['aceite_base'].append(float(record.aceite_base_frita))
            except:
                pass
        
        # Campos PT Producto Terminado
        if record.aceite_pt_producto_terminado:
            try:
                productos[producto]['aceite_pt_general'].append(float(record.aceite_pt_producto_terminado))
            except:
                pass
                
        if record.humedad_pt_producto_terminado:
            try:
                productos[producto]['humedad_pt_general'].append(float(record.humedad_pt_producto_terminado))
            except:
                pass
                
        if record.sal_pt_producto_terminado:
            try:
                productos[producto]['sal_pt_general'].append(float(record.sal_pt_producto_terminado))
            except:
                pass
    
    # Calcular estadísticas
    stats_data = []
    for producto_data in productos.values():
        stats = {'Producto': producto_data['producto'], 'Total Registros': producto_data['total_registros']}
        
        for campo in ['humedad_base', 'aceite_base', 'aceite_pt_general', 'humedad_pt_general', 'sal_pt_general']:
            valores = producto_data[campo]
            if valores:
                stats[f'{campo.title()} - Promedio'] = round(sum(valores) / len(valores), 2)
                stats[f'{campo.title()} - Mín'] = round(min(valores), 2)
                stats[f'{campo.title()} - Máx'] = round(max(valores), 2)
            else:
                stats[f'{campo.title()} - Promedio'] = 'N/A'
                stats[f'{campo.title()} - Mín'] = 'N/A'
                stats[f'{campo.title()} - Máx'] = 'N/A'
        
        stats_data.append(stats)
    
    return pd.DataFrame(stats_data)

def crear_dataframe_rangos(categoria):
    """Crea DataFrame con rangos de referencia por producto"""
    
    # Rangos según categoría (copiados del JS unificado)
    rangos_por_categoria = {
        'EXTRUIDOS': {
            'CHEETOS TORCIDITOS': {
                'humedad_base': '0.7 - 1.7', 'aceite_base': '21.7 - 27.7',
                'humedad_pt': '0.5 - 1.9', 'aceite_pt': '32.46 - 38.46', 'sal_pt': '0.95 - 1.55'
            },
            'CHEETOS XTRA FLAMIN HOT': {
                'humedad_base': '0.7 - 1.7', 'aceite_base': '21.7 - 27.7',
                'humedad_pt': '0.47 - 1.67', 'aceite_pt': '29.52 - 35.52', 'sal_pt': '1.40 - 1.80'
            },
            'CHEETOS JALAQUEÑO': {
                'humedad_base': '0.7 - 1.7', 'aceite_base': '21.7 - 27.7',
                'humedad_pt': '0.5 - 1.9', 'aceite_pt': '31.64 - 37.64 (Verde)', 'sal_pt': '1.06 - 1.66 (Verde)'
            },
            'CHEETOS XTRA FH NUEVO': {
                'humedad_base': '0.7 - 1.7', 'aceite_base': '21.7 - 27.7',
                'humedad_pt': '0.5 - 1.9', 'aceite_pt': '29.35 - 35.35', 'sal_pt': '1.16 - 1.76'
            }   
        },
        'TORTILLA': {
            'DORITOS': {
                'humedad_base': '1.00 - 1.20', 'aceite_base': '20.00 - 23.00',
                'humedad_pt': '0.78 - 1.58', 'aceite_pt': '23.45 - 26.45', 'sal_pt': '0.90 - 1.50'
            },
            'DORITOS INCÓGNITA': {
                'humedad_base': '1.02 - 1.62', 'aceite_base': '20.00 - 23.00',
                'humedad_pt': '1.07 - 1.57', 'aceite_pt': '22.35 - 25.35', 'sal_pt': '0.72 - 1.32'
            },
            'TOSTITOS SALSA VERDE': {
                'humedad_base': '0.90 - 1.30', 'aceite_base': '22.00 - 24.00',
                'humedad_pt': '1.03 - 1.63', 'aceite_pt': '23.14 - 26.14', 'sal_pt': '0.97 - 1.57'
            }
        },
        'PAPA': {
            'PAPA SAL': {
                'humedad_base': '1.35 - 1.65 (Verde)', 'aceite_base': '31 - 35 (Verde)',
                'cloruros_base': '0 - 1 (Verde)', 'humedad_pt': '1.35 - 1.8 (Verde)',
                'aceite_pt': 'N/A', 'sal_pt': '0.55 - 0.85 (Verde)'
            },
            'SABRITAS LIMON': {
                'humedad_base': '1.35 - 1.65 (Verde)', 'aceite_base': '31 - 35 (Verde)',
                'cloruros_base': 'N/A', 'humedad_pt': 'N/A',
                'aceite_pt': 'N/A', 'sal_pt': '1.23 - 1.50 (Verde)'
            },
            'RUFFLES SAL': {
                'humedad_base': '1.35 - 1.65 (Verde)', 'aceite_base': '31 - 35 (Verde)',
                'cloruros_base': '0 - 1 (Verde)', 'humedad_pt': '1.35 - 1.8 (Verde)',
                'aceite_pt': 'N/A', 'sal_pt': '0.55 - 0.85 (Verde)'
            },
            'RUFFLES QUESO': {
                'humedad_base': '1.20 - 1.5 (Verde)', 'aceite_base': '31 - 35 (Verde)',
                'cloruros_base': '0 - 1 (Verde)', 'humedad_pt': '1.35 - 1.8 (Verde)',
                'aceite_pt': 'N/A', 'sal_pt': '1.24 - 1.54 (Verde)'
            }
            ,
            'SABRITAS XTRA FH': {
                'humedad_base': '1.35 - 1.65 (Verde)', 'aceite_base': '31 - 35 (Verde)',
                'cloruros_base': '0 - 1 (Verde)', 'humedad_pt': '1.41 - 1.71 (Verde)',
                'aceite_pt': 'N/A', 'sal_pt': '2.03 - 2.63 (Verde)'
            }
        }
    }
    
    rangos_data = []
    productos_categoria = rangos_por_categoria.get(categoria, {})

    for producto, rangos in productos_categoria.items():
        row_data = {
            'Producto': producto,
            'Humedad Base Frita': rangos['humedad_base'],
            'Aceite Base Frita': rangos['aceite_base'],
            'Humedad PT': rangos['humedad_pt'],
            'Aceite PT': rangos['aceite_pt'],
            'Sal PT': rangos['sal_pt']
        }

        # Agregar Cloruros Base solo para PAPA
        if categoria == 'PAPA' and 'cloruros_base' in rangos:
            row_data['Cloruros Base'] = rangos['cloruros_base']

        rangos_data.append(row_data)

    return pd.DataFrame(rangos_data)

def aplicar_formato_excel(writer, sheet_name, df, records, categoria):
    """Aplica colores y formato al Excel basado en los rangos"""
    
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # Formato para headers
    header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Aplicar formato a headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Aplicar colores a celdas de datos
    for row_idx, record in enumerate(records, 2):  # Start from row 2 (after header)
        aplicar_colores_fila(worksheet, row_idx, record, categoria)
    
    # Ajustar ancho de columnas
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 30)
        worksheet.column_dimensions[column_letter].width = adjusted_width

def aplicar_colores_fila(worksheet, row, record, categoria):
    """Aplica colores a una fila específica basado en rangos"""

    # Mapeo de columnas a valores y tipos
    # Para PAPA (con Cloruros Base):
    # Orden: Folio(1), Fecha(2), Turno(3), Producto(4), Horario(5),
    #        Humedad Base(6), Aceite Base(7),
    #        Aceite PT Gral(8), Humedad PT Gral(9), Sal PT Gral(10),
    #        Aceite T1(11), Humedad T1(12), Sal Tit T1(13), Sal PT T1(14),
    #        Aceite T2(15), Humedad T2(16), Sal Tit T2(17), Sal PT T2(18),
    #        Aceite T3(19), Humedad T3(20), Sal Tit T3(21), Sal PT T3(22),
    #        Cloruros Base(23), Observaciones(24), Fecha Creación(25)
    #
    # Para EXTRUIDOS y TORTILLA (sin Cloruros Base):
    # Orden: Folio(1), Fecha(2), Turno(3), Producto(4), Horario(5),
    #        Humedad Base(6), Aceite Base(7),
    #        Aceite PT Gral(8), Humedad PT Gral(9), Sal PT Gral(10),
    #        Aceite T1(11), Humedad T1(12), Sal Tit T1(13), Sal PT T1(14),
    #        Aceite T2(15), Humedad T2(16), Sal Tit T2(17), Sal PT T2(18),
    #        Aceite T3(19), Humedad T3(20), Sal Tit T3(21), Sal PT T3(22),
    #        Observaciones(23), Fecha Creación(24)

    # Campos base (sin Cloruros Base)
    campos_colores = {
        6: (record.humedad_base_frita, 'humedad_base'),         # Humedad Base Frita
        7: (record.aceite_base_frita, 'aceite_base'),           # Aceite Base Frita
        8: (record.aceite_pt_producto_terminado, 'aceite_pt'),  # Aceite PT General
        9: (record.humedad_pt_producto_terminado, 'humedad_pt'), # Humedad PT General
        10: (record.sal_pt_producto_terminado, 'sal_pt'),       # Sal PT General
        11: (record.tanque1_aceite_pt, 'aceite_pt'),            # Aceite PT T1
        12: (record.tanque1_humedad_pt, 'humedad_pt'),          # Humedad PT T1
        # 13: Sal Titulador T1 (sin color, es entrada)
        14: (record.tanque1_sal_pt, 'sal_pt'),                  # Sal PT T1
        15: (record.tanque2_aceite_pt, 'aceite_pt'),            # Aceite PT T2
        16: (record.tanque2_humedad_pt, 'humedad_pt'),          # Humedad PT T2
        # 17: Sal Titulador T2 (sin color, es entrada)
        18: (record.tanque2_sal_pt, 'sal_pt'),                  # Sal PT T2
        19: (record.tanque3_aceite_pt, 'aceite_pt'),            # Aceite PT T3
        20: (record.tanque3_humedad_pt, 'humedad_pt'),          # Humedad PT T3
        # 21: Sal Titulador T3 (sin color, es entrada)
        22: (record.tanque3_sal_pt, 'sal_pt'),                  # Sal PT T3
    }

    # Agregar Cloruros Base solo para PAPA (columna 23)
    if categoria == 'PAPA':
        campos_colores[23] = (record.cloruros_base, 'cloruros_base')

    for col_idx, (valor, tipo_campo) in campos_colores.items():
        if valor is not None and str(valor).strip():
            color_class = determinar_color_excel(str(valor), tipo_campo, categoria, record.producto)

            if color_class != 'empty':
                cell = worksheet.cell(row=row, column=col_idx)
                cell.fill = PatternFill(start_color=COLORS[color_class], end_color=COLORS[color_class], fill_type='solid')

def determinar_color_excel(valor, tipo_campo, categoria, producto):
    """Determina el color basado en rangos (versión simplificada para Excel)"""

    try:
        valor_num = float(valor)
    except:
        return 'empty'

    # Rangos simplificados para Excel
    if categoria == 'EXTRUIDOS':
        if 'JALAQUEÑO' in producto:
            if tipo_campo == 'aceite_pt':
                if 31.64 <= valor_num <= 37.64:
                    return 'success'
                elif (29.64 <= valor_num <= 31.63) or (37.65 <= valor_num <= 39.64):
                    return 'warning'
                else:
                    return 'danger'
            elif tipo_campo == 'sal_pt':
                if 1.06 <= valor_num <= 1.66:
                    return 'success'
                elif (0.95 <= valor_num <= 1.05) or (1.67 <= valor_num <= 1.77):
                    return 'warning'
                else:
                    return 'danger'

        # Rangos estándar EXTRUIDOS
        rangos = {
            'humedad_base': {'verde': (0.7, 1.7), 'amarillo': [(0.60, 0.69), (1.71, 1.80)]},
            'aceite_base': {'verde': (21.7, 27.7), 'amarillo': [(20.7, 21.69), (27.71, 28.7)]},
            'humedad_pt': {'verde': (0.5, 1.9), 'amarillo': [(1.91, 2.10)]},
            'aceite_pt': {'verde': (32.46, 38.46), 'amarillo': [(31.46, 32.45), (38.47, 39.46)]},
            'sal_pt': {'verde': (0.95, 1.55), 'amarillo': [(0.85, 0.94), (1.56, 1.65)]},
            'cloruros_base': {'verde': (0, 100), 'amarillo': []}  # No aplica para EXTRUIDOS
        }
    elif categoria == 'PAPA':
        # Rangos específicos para productos PAPA
        if producto == 'RUFFLES QUESO':
            rangos = {
                'humedad_base': {'verde': (1.20, 1.5), 'amarillo': [(1.05, 1.19), (1.51, 1.65)]},
                'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                'humedad_pt': {'verde': (1.35, 1.8), 'amarillo': [(1.20, 2)]},
                'sal_pt': {'verde': (1.24, 1.54), 'amarillo': [(1.19, 1.23), (1.55, 1.59)]},
                'cloruros_base': {'verde': (0, 1), 'amarillo': []},  # Solo verde o rojo
                'aceite_pt': {'verde': (0, 0), 'amarillo': []}
            }
        elif producto == 'SABRITAS XTRA FH':
            rangos = {
                'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                'humedad_pt': {'verde': (1.41, 1.71), 'amarillo': [(1.21, 1.40), (1.70, 1.91)]},
                'sal_pt': {'verde': (1.58, 1.88), 'amarillo': [(1.38, 1.57), (1.89, 2.08)]},
                'cloruros_base': {'verde': (0, 1), 'amarillo': []},  # Solo verde o rojo
                'aceite_pt': {'verde': (0, 0), 'amarillo': []}
            }
        elif producto == 'RUFFLES SAL':
            rangos = {
                'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                'humedad_pt': {'verde': (1.35, 1.8), 'amarillo': [(1.20, 2)]},
                'sal_pt': {'verde': (0.55, 0.85), 'amarillo': [(0.45, 0.54), (0.86, 0.95)]},
                'cloruros_base': {'verde': (0, 1), 'amarillo': []},  # Solo verde o rojo
                'aceite_pt': {'verde': (0, 0), 'amarillo': []}
            }
        elif producto == 'SABRITAS LIMON':
            rangos = {
                'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                'humedad_pt': {'verde': (0, 0), 'amarillo': []},  # N/A
                'sal_pt': {'verde': (1.23, 1.50), 'amarillo': [(1.10, 1.22), (1.51, 1.63)]},
                'cloruros_base': {'verde': (0, 100), 'amarillo': []},  # N/A
                'aceite_pt': {'verde': (0, 0), 'amarillo': []}  # N/A
            }
        else:  # PAPA SAL (default)
            rangos = {
                'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                'humedad_pt': {'verde': (1.35, 1.8), 'amarillo': [(1.20, 2)]},
                'sal_pt': {'verde': (0.55, 0.85), 'amarillo': [(0.45, 0.54), (0.86, 0.95)]},
                'cloruros_base': {'verde': (0, 1), 'amarillo': []},  # Solo verde o rojo
                'aceite_pt': {'verde': (0, 0), 'amarillo': []}
            }
        
    else:  # TORTILLA
        rangos = {
            'humedad_base': {'verde': (1.0, 1.2), 'amarillo': [(0.8, 0.99), (1.21, 1.3)]},
            'aceite_base': {'verde': (20.0, 23.0), 'amarillo': [(21.0, 21.99), (23.01, 24.0)]},
            'humedad_pt': {'verde': (0.78, 1.58), 'amarillo': [(0.68, 0.77), (1.59, 1.68)]},
            'aceite_pt': {'verde': (23.45, 26.45), 'amarillo': [(22.45, 23.44), (26.46, 27.45)]},
            'sal_pt': {'verde': (0.9, 1.5), 'amarillo': [(0.8, 0.89), (1.51, 1.6)]},
            'cloruros_base': {'verde': (0, 100), 'amarillo': []}  # No aplica para TORTILLA
        }

    if tipo_campo in rangos:
        rango = rangos[tipo_campo]

        # Verificar verde
        if rango['verde'][0] <= valor_num <= rango['verde'][1]:
            return 'success'

        # Verificar amarillo
        for amarillo in rango['amarillo']:
            if amarillo[0] <= valor_num <= amarillo[1]:
                return 'warning'

        return 'danger'

    return 'empty'
