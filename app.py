import os
import re
from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_wtf.csrf import CSRFProtect
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from urllib.parse import urlparse as url_parse
from werkzeug.utils import secure_filename
from datetime import datetime,timedelta
from io import BytesIO
import pandas as pd
import json

from config import Config
from forms import LoginForm, RegistrationForm, ChangePasswordForm, MenuItemForm, CreateFormForm, FormQuestionForm, PNCForm, PNCSimpleForm, CalidadTortillaForm, AnalisisCalidadForm, PAEForm, WeakLinkForm, PesoForm
from models import db, User, MenuItem, Form, FormQuestion, FormResponse, FormAnswer, PNC, PNCSimple, CalidadTortilla, AnalisisCalidad, PAERegistro, WeakLink, Peso
from utils import admin_required, save_image, allowed_file
from aceite_routes import setup_aceite_routes
from register_pesos_routes import register_pesos_routes
from excel_fisicoquimicos_routes import setup_excel_fisicoquimicos_routes
from pae_visualizacion_routes import setup_pae_visualizacion_routes
from papa_excel_routes import setup_papa_excel_routes

def create_app(config_class=Config):
    app = Flask(__name__)
    app.config.from_object(config_class)
    
    # Inicializar extensiones
    db.init_app(app)
    
    # Inicializar protección CSRF
    csrf = CSRFProtect(app)
    
    # Configurar Login Manager
    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'
    login_manager.login_message = 'Por favor inicie sesión para acceder a esta página.'
    
    @login_manager.user_loader
    def load_user(user_id):
        # Usar Session.get() en lugar de Query.get() para evitar el warning de deprecación
        return db.session.get(User, int(user_id))
    
    # Crear directorios si no existen
    # Configurar rutas para el módulo de aceite
    setup_aceite_routes(app)
    
    # Configurar rutas para el módulo de pesos
    register_pesos_routes(app)
    
    # Configurar rutas para descarga Excel fisicoquímicos
    setup_excel_fisicoquimicos_routes(app)

    # Configurar rutas para visualización de registros PAE
    setup_pae_visualizacion_routes(app)

    # Configurar rutas para descarga Excel PAE PAPA (con todas las columnas)
    setup_papa_excel_routes(app)
    
    with app.app_context():
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        db.create_all()  # Crear tablas si no existen
        
        # Crear usuario admin por defecto si no existe
        if not User.query.filter_by(username='admin').first():
            admin = User(
                username='admin',
                user_id='GPADMIN001',
                email='admin@example.com',
                password='admin123',
                is_admin=True
            )
            db.session.add(admin)
            db.session.commit()
    
    # Contexto global para las plantillas
    @app.context_processor
    def inject_now():
        return {'now': datetime.now()}
    
    # Rutas de autenticación
    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('index'))
        
        form = LoginForm()
        if form.validate_on_submit():
            user = User.query.filter_by(user_id=form.user_id.data).first()
            
            if user is None or not user.check_password(form.password.data):
                flash('GPID o contraseña incorrectos', 'danger')
                return redirect(url_for('login'))
            
            # Actualizar último login
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            login_user(user, remember=form.remember_me.data)
            next_page = request.args.get('next')
            if not next_page or url_parse(next_page).netloc != '':
                next_page = url_for('index')
            
            return redirect(next_page)
        
        return render_template('auth/login.html', title='Iniciar Sesión', form=form)
    
    @app.route('/logout')
    def logout():
        logout_user()
        return redirect(url_for('login'))
    
    # Gestión de usuarios (solo admin)
    @app.route('/admin/users')
    @login_required
    @admin_required
    def admin_users():
        users = User.query.all()
        return render_template('admin/users.html', title='Gestión de Usuarios', users=users)
    
    @app.route('/admin/users/add', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def add_user():
        form = RegistrationForm()
        if form.validate_on_submit():
            user = User(
                username=form.username.data,
                user_id=form.user_id.data,
                password=form.password.data,
                is_admin=form.is_admin.data
            )
            db.session.add(user)
            db.session.commit()
            flash(f'Usuario {form.username.data} creado exitosamente!', 'success')
            return redirect(url_for('admin_users'))
        
        return render_template('auth/register.html', title='Agregar Usuario', form=form)
    
    @app.route('/admin/users/delete/<int:id>', methods=['POST'])
    @login_required
    @admin_required
    def delete_user(id):
        user = db.session.get(User, id)  # Usando session.get en lugar de query.get_or_404
        
        if not user:
            flash('Usuario no encontrado', 'danger')
            return redirect(url_for('admin_users'))
        
        # No permitir eliminar al propio usuario
        if user.id == current_user.id:
            flash('No puedes eliminar tu propia cuenta', 'danger')
            return redirect(url_for('admin_users'))
        
        db.session.delete(user)
        db.session.commit()
        flash(f'Usuario {user.username} eliminado correctamente', 'success')
        return redirect(url_for('admin_users'))
    
    @app.route('/change-password', methods=['GET', 'POST'])
    @login_required
    def change_password():
        form = ChangePasswordForm()
        if form.validate_on_submit():
            if not current_user.check_password(form.current_password.data):
                flash('Contraseña actual incorrecta', 'danger')
                return redirect(url_for('change_password'))
            
            current_user.set_password(form.new_password.data)
            db.session.commit()
            flash('Contraseña actualizada correctamente', 'success')
            return redirect(url_for('index'))
        
        return render_template('auth/change_password.html', title='Cambiar Contraseña', form=form)
    
    # Página principal y menú
    @app.route('/')
    @login_required
    def index():
        # Obtener elementos de menú agrupados por categoría
        extruidos = MenuItem.query.filter_by(category='EXTRUIDOS').first()
        tortilla = MenuItem.query.filter_by(category='TORTILLA').first()
        papa = MenuItem.query.filter_by(category='PAPA').first()
        
        # Si no hay elementos, crear objetos temporales con imágenes por defecto
        if not extruidos:
            extruidos = {'category': 'EXTRUIDOS', 'image_path': 'img/default/extruidos.jpg', 'description': 'Administrar productos extruidos'}
        if not tortilla:
            tortilla = {'category': 'TORTILLA', 'image_path': 'img/default/tortilla.jpg', 'description': 'Administrar productos de tortilla'}
        if not papa:
            papa = {'category': 'PAPA', 'image_path': 'img/default/papa.jpg', 'description': 'Administrar productos de papa'}
        
        return render_template('dashboard/index.html', 
                               title='Panel Principal',
                               extruidos=extruidos,
                               tortilla=tortilla,
                               papa=papa)
    
    @app.route('/line/calidad/<category>')
    @login_required
    def line_menu(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        menu_item = MenuItem.query.filter_by(category=category).first()
        if not menu_item:
            menu_item = {'category': category, 'image_path': f'img/default/{category.lower()}.jpg', 'description': f'Línea de producción {category}'}

        return render_template('dashboard/line_menu.html',
                               title=f'Secciones - {category}',
                               category=category,
                               menu_item=menu_item)
    
    # Ruta para mostrar el menú de secciones según la línea seleccionada
    @app.route('/line/<category>')
    @login_required
    def line_sections(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
            
        # Obtener el elemento del menú para esta categoría
        menu_item = MenuItem.query.filter_by(category=category).first()
        if not menu_item:
            menu_item = {'category': category, 'image_path': f'img/default/{category.lower()}.jpg', 'description': f'Línea de producción {category}'}
        
        # Definir las secciones disponibles para cada línea
        sections = [
            {'id': 'calidad', 'name': 'Calidad', 'icon': 'fa-award', 'color': 'primary'},
            {'id': 'produccion', 'name': 'Producción', 'icon': 'fa-industry', 'color': 'success'},
            {'id': 'mantenimiento', 'name': 'Mantenimiento', 'icon': 'fa-tools', 'color': 'warning'},
            {'id': 'inocuidad', 'name': 'Inocuidad', 'icon': 'fa-shield-alt', 'color': 'info'},
            {'id': 'formularios', 'name': 'Nuevo formulario y recetas', 'icon': 'fa-clipboard-list', 'color': 'secondary'}
        ]
        
        return render_template('dashboard/line_sections.html',
                               title=f'Secciones - {category}',
                               category=category,
                               menu_item=menu_item,
                               sections=sections)
    
    # Ruta para mostrar ejercicios de una sección específica
    @app.route('/line/<category>/<section>/<subsection>')
    @login_required
    def calidad_exercises(category, section, subsection):
        section_name = 'Calidad'
        section_icon = 'fa-award'
        menu_item = MenuItem.query.filter_by(category=category).first()
        if not menu_item:
            menu_item = {'category': category, 'image_path': f'img/default/{category.lower()}.jpg', 'description': f'Línea de producción {category}'}

        if subsection == 'kpi':
            exercises = [
                {'id': 'pnc', 'name': 'PNC', 'icon': 'fa-times-circle', 'color': 'danger', 'description': 'Registro y análisis de Producto No Conforme'},
                {'id': 'sensorial', 'name': 'Sensorial', 'icon': 'fa-eye', 'color': 'info', 'description': 'Evaluación sensorial de producto'},
                {'id': 'irc', 'name': 'IRC', 'icon': 'fa-chart-bar', 'color': 'secondary', 'description': 'Indicadores de Reclamaciones y Calidad'},
                {'id': 'quejas_consumidores', 'name': 'Quejas consumidores', 'icon': 'fa-user-times', 'color': 'warning', 'description': 'Registro de quejas de consumidores'}
            ]
        else:
            exercises = [
                {'id': 'pae', 'name': 'PAE', 'icon': 'fa-clipboard-check', 'color': 'primary', 'description': 'Protocolos de Análisis y Evaluación'},
                {'id': 'wl_proceso', 'name': 'WEAK LINK PROCESO', 'icon': 'fa-link', 'color': 'secondary', 'description': 'Análisis de puntos débiles en el proceso'},
                {'id': 'weaklink', 'name': 'WEAK LINK EMPAQUE', 'icon': 'fa-link-slash', 'color': 'warning', 'description': 'Análisis de puntos débiles en la cadena de empaque'},
                {'id': 'laboratorio', 'name': 'LABORATORIO', 'icon': 'fa-flask', 'color': 'info', 'description': 'Registro y seguimiento de análisis fisicoquímicos'}
            ]
            if category == 'TORTILLA':
                exercises.append({'id': 'pesos', 'name': 'PESOS PROCESOS', 'icon': 'fa-balance-scale', 'color': 'success', 'description': 'Control de pesos en los procesos productivos'})

        return render_template('calidad_tortilla/calidad.html',
                               title=f'{section_name} - {category}',
                               category=category,
                               section=section,
                                 subsection=subsection,
                               section_name=section_name,
                               section_icon=section_icon,
                               menu_item=menu_item,
                               exercises=exercises)

    @app.route('/under_construction')
    @login_required
    def under_construction():
        return render_template('under_construction.html')
                               
    # Ruta para mostrar página en construcción
    @app.route('/line/<category>/<section>')
    @login_required
    def section_exercises(category, section):
        # Para la sección de formularios, redirigir directamente a la página de formularios personalizados
        if section == 'formularios':
            return redirect(url_for('custom_forms', category=category))
            
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))

        # Para ejercicios específicos, redirigir a sus herramientas correspondientes
        if request.args.get('exercise') == 'nuevo_formulario':
            return redirect(url_for('custom_forms', category=category))
        elif request.args.get('exercise') == 'pnc':
            return redirect(url_for('list_pnc_simple', category=category))
        elif request.args.get('exercise') == 'pae':
            return redirect(url_for('pae_dashboard', category=category))
        elif request.args.get('exercise') == 'weaklink':
            return redirect(url_for('weaklink_dashboard', category=category))
        
        # Para secciones en construcción: produccion, mantenimiento, inocuidad
        if section in ['produccion', 'mantenimiento', 'inocuidad']:
            return render_template('under_construction.html',
                                  title=f'Módulo en Construcción - {category}',
                                  category=category)
            
        # Verificar sección válida para otras secciones
        valid_sections = ['calidad', 'produccion', 'mantenimiento', 'inocuidad', 'formularios']
        if section not in valid_sections:
            flash('Sección no válida', 'danger')
            return redirect(url_for('line_sections', category=category))
        
        # Obtener el elemento del menú para esta categoría
        menu_item = MenuItem.query.filter_by(category=category).first()
        if not menu_item:
            menu_item = {'category': category, 'image_path': f'img/default/{category.lower()}.jpg', 'description': f'Línea de producción {category}'}
        
        # Definir ejercicios disponibles según la sección
        exercises = []
        section_name = ''
        section_icon = ''
        
        if section == 'calidad':
            section_name = 'Calidad'
            section_icon = 'fa-award'
            exercises = [
                {'id': 'pnc', 'name': 'PNC', 'icon': 'fa-times-circle', 'color': 'danger', 'description': 'Registro y análisis de Producto No Conforme'},
                {'id': 'pae', 'name': 'PAE', 'icon': 'fa-clipboard-check', 'color': 'primary', 'description': 'Protocolos de Análisis y Evaluación'},
                {'id': 'weaklink', 'name': 'WEAK LINK', 'icon': 'fa-link-slash', 'color': 'warning', 'description': 'Análisis de puntos débiles en la cadena de producción'},
                {'id': 'laboratorio', 'name': 'LABORATORIO', 'icon': 'fa-flask', 'color': 'info', 'description': 'Registro y seguimiento de análisis fisicoquímicos'}
            ]
            # Solo agregar opción de Pesos para TORTILLA
            if category == 'TORTILLA':
                exercises.append({'id': 'pesos', 'name': 'PESOS PROCESOS', 'icon': 'fa-balance-scale', 'color': 'success', 'description': 'Control de pesos en los procesos productivos'})
        elif section == 'formularios':
            section_name = 'Nuevo formulario y recetas'
            section_icon = 'fa-clipboard-list'
            exercises = [
                {'id': 'nuevo_formulario', 'name': 'CREAR NUEVO FORMULARIO Y RECETAS', 'icon': 'fa-clipboard-list', 'color': 'secondary', 'description': 'Crear y gestionar formularios personalizados'}
            ]
        
        return render_template('dashboard/section_exercises.html',
                               title=f'{section_name} - {category}',
                               category=category,
                               section=section,
                               section_name=section_name,
                               section_icon=section_icon,
                               menu_item=menu_item,
                               exercises=exercises)
    
    # Rutas para gestionar formularios
    @app.route('/forms/<category>/<section>')
    @login_required
    def list_forms(category, section):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
            
        # Verificar sección válida
        valid_sections = ['calidad', 'produccion', 'mantenimiento', 'inocuidad', 'formularios']
        if section not in valid_sections:
            flash('Sección no válida', 'danger')
            return redirect(url_for('line_sections', category=category))
        
        # Obtener el elemento del menú para esta categoría
        menu_item = MenuItem.query.filter_by(category=category).first()
        if not menu_item:
            menu_item = {'category': category, 'image_path': f'img/default/{category.lower()}.jpg', 'description': f'Línea de producción {category}'}
        
        # Obtener formularios para esta categoría y sección
        forms = Form.query.filter_by(category=category, section=section).order_by(Form.created_at.desc()).all()
        
        # Obtener nombre de la sección y ícono
        section_info = {
            'calidad': {'name': 'Calidad', 'icon': 'fa-award'},
            'produccion': {'name': 'Producción', 'icon': 'fa-industry'},
            'mantenimiento': {'name': 'Mantenimiento', 'icon': 'fa-tools'},
            'inocuidad': {'name': 'Inocuidad', 'icon': 'fa-shield-alt'},
            'formularios': {'name': 'Nuevo formulario y recetas', 'icon': 'fa-clipboard-list'}
        }
        
        return render_template('forms/list_forms.html',
                              title=f'Formularios - {category}',
                              category=category,
                              section=section,
                              section_name=section_info[section]['name'],
                              section_icon=section_info[section]['icon'],
                              menu_item=menu_item,
                              forms=forms)
    
    @app.route('/forms/create/<category>/<section>', methods=['GET', 'POST'])
    @login_required
    def create_form(category, section):
        # Verificaciones de categoría y sección válidas
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
            
        if section not in ['calidad', 'produccion', 'mantenimiento', 'inocuidad', 'formularios']:
            flash('Sección no válida', 'danger')
            return redirect(url_for('line_sections', category=category))
        
        form = CreateFormForm()
        # Preseleccionar categoría y sección
        form.category.data = category
        form.section.data = section
        
        if form.validate_on_submit():
            # Crear nuevo formulario
            new_form = Form(
                title=form.title.data,
                description=form.description.data,
                category=form.category.data,
                section=form.section.data,
                is_active=form.is_active.data,
                created_by=current_user.id
            )
            
            db.session.add(new_form)
            db.session.commit()
            
            flash(f'Formulario "{form.title.data}" creado exitosamente. Ahora puedes agregar preguntas.', 'success')
            return redirect(url_for('edit_form', form_id=new_form.id))
        
        return render_template('forms/create_form.html',
                              title='Crear Nuevo Formulario',
                              form=form,
                              category=category,
                              section=section)
    
    @app.route('/forms/edit/<int:form_id>', methods=['GET', 'POST'])
    @login_required
    def edit_form(form_id):
        # Obtener el formulario
        form_db = Form.query.get_or_404(form_id)
        
        # Verificar si es el creador o un administrador
        if form_db.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para editar este formulario', 'danger')
            return redirect(url_for('list_forms', category=form_db.category, section=form_db.section))
        
        # Inicializar formulario para editar detalles generales
        form = CreateFormForm(obj=form_db)
        
        # Si se envió el formulario principal
        if form.validate_on_submit() and request.form.get('action') == 'update_details':
            form.populate_obj(form_db)
            db.session.commit()
            flash('Detalles del formulario actualizados correctamente', 'success')
            return redirect(url_for('edit_form', form_id=form_id))
        
        # Obtener las preguntas ordenadas
        questions = FormQuestion.query.filter_by(form_id=form_id).order_by(FormQuestion.order).all()
        
        # Formulario para agregar/editar pregunta
        question_form = FormQuestionForm()
        
        # Lógica para agregar una nueva pregunta
        if question_form.validate_on_submit() and request.form.get('action') == 'add_question':
            # Obtener el orden máximo actual
            max_order = db.session.query(db.func.max(FormQuestion.order)).filter_by(form_id=form_id).scalar() or 0
            
            # Crear la nueva pregunta
            new_question = FormQuestion(
                form_id=form_id,
                question_text=question_form.question_text.data,
                question_type=question_form.question_type.data,
                is_required=question_form.is_required.data,
                order=max_order + 1
            )
            
            # Procesar opciones si es necesario
            if question_form.question_type.data in ['radio', 'checkbox', 'select'] and question_form.options.data:
                options_list = [option.strip() for option in question_form.options.data.split('\n') if option.strip()]
                new_question.set_options(options_list)
            
            db.session.add(new_question)
            db.session.commit()
            
            flash('Pregunta agregada correctamente', 'success')
            return redirect(url_for('edit_form', form_id=form_id))
        
        # Lógica para manejar operaciones AJAX de las preguntas
        if request.method == 'POST' and request.form.get('action') == 'delete_question':
            question_id = request.form.get('question_id')
            if question_id:
                question = FormQuestion.query.get(question_id)
                if question and question.form_id == form_id:
                    db.session.delete(question)
                    db.session.commit()
                    flash('Pregunta eliminada correctamente', 'success')
                    return redirect(url_for('edit_form', form_id=form_id))
        
        return render_template('forms/edit_form.html',
                              title=f'Editar Formulario: {form_db.title}',
                              form=form,
                              form_db=form_db,
                              questions=questions,
                              question_form=question_form)
    
    @app.route('/forms/view/<int:form_id>')
    @login_required
    def view_form(form_id):
        # Obtener el formulario
        form = Form.query.get_or_404(form_id)
        
        # Obtener las preguntas ordenadas
        questions = FormQuestion.query.filter_by(form_id=form_id).order_by(FormQuestion.order).all()
        
        return render_template('forms/view_form.html',
                              title=form.title,
                              form=form,
                              questions=questions)
    
    @app.route('/forms/submit/<int:form_id>', methods=['POST'])
    @login_required
    def submit_form(form_id):
        # Obtener el formulario
        form = Form.query.get_or_404(form_id)
        
        # Crear una nueva respuesta
        response = FormResponse(
            form_id=form_id,
            user_id=current_user.id
        )
        
        db.session.add(response)
        db.session.commit()
        
        # Procesar respuestas para cada pregunta
        questions = FormQuestion.query.filter_by(form_id=form_id).all()
        for question in questions:
            # Obtener el valor de la respuesta del formulario
            answer_key = f'question_{question.id}'
            answer_value = ''
            
            if question.question_type == 'checkbox':
                # Para checkboxes, pueden venir múltiples valores
                values = request.form.getlist(answer_key)
                answer_value = ','.join(values) if values else ''
            else:
                # Para otros tipos de preguntas
                answer_value = request.form.get(answer_key, '')
            
            # Guardar la respuesta
            answer = FormAnswer(
                response_id=response.id,
                question_id=question.id,
                answer_text=answer_value
            )
            
            db.session.add(answer)
        
        db.session.commit()
        
        flash('Formulario enviado correctamente', 'success')
        return redirect(url_for('list_forms', category=form.category, section=form.section))
    
    @app.route('/forms/responses/<int:form_id>')
    @login_required
    def view_responses(form_id):
        # Obtener el formulario
        form = Form.query.get_or_404(form_id)
        
        # Verificar si es el creador o un administrador
        if form.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para ver las respuestas de este formulario', 'danger')
            return redirect(url_for('list_forms', category=form.category, section=form.section))
        
        # Obtener las respuestas
        responses = FormResponse.query.filter_by(form_id=form_id).order_by(FormResponse.submitted_at.desc()).all()
        
        # Obtener preguntas para mostrar cabeceras
        questions = FormQuestion.query.filter_by(form_id=form_id).order_by(FormQuestion.order).all()
        
        return render_template('forms/view_responses.html',
                              title=f'Respuestas: {form.title}',
                              form=form,
                              questions=questions,
                              responses=responses)
    
    # Gestión de elementos del menú
    @app.route('/menu/edit/<category>', methods=['GET', 'POST'])
    @login_required
    @admin_required  # Solo administradores pueden editar
    def edit_menu_item(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Categoría no válida', 'danger')
            return redirect(url_for('index'))
        
        # Buscar elemento existente o crear uno nuevo
        menu_item = MenuItem.query.filter_by(category=category).first()
        if not menu_item:
            menu_item = MenuItem(name=category, category=category)
            # Asignar ruta de imagen por defecto para nuevos elementos
            menu_item.image_path = f'img/default/{category.lower()}.jpg'
        
        form = MenuItemForm(obj=menu_item)
        
        if form.validate_on_submit():
            form.populate_obj(menu_item)
            
            # Procesar imagen si se proporciona
            if form.image.data:
                image_path = save_image(form.image.data, category)
                if image_path:
                    menu_item.image_path = image_path
            
            # Guardar en la base de datos
            if menu_item.id is None:
                db.session.add(menu_item)
            db.session.commit()
            
            flash(f'Elemento de menú {category} actualizado correctamente', 'success')
            return redirect(url_for('index'))
        
        return render_template('dashboard/edit_menu_item.html', 
                              title=f'Editar {category}',
                              form=form,
                              category=category,
                              menu_item=menu_item)
    
    # Rutas para la nueva herramienta de formularios personalizados
    @app.route('/custom_forms/<category>')
    @login_required
    def custom_forms(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener formularios personalizados para esta categoría
        custom_forms = Form.query.filter_by(category=category, section='nuevo_formulario').order_by(Form.created_at.desc()).all()
        
        return render_template('forms/custom_forms.html',
                             title=f'Formularios Personalizados - {category}',
                             category=category,
                             forms=custom_forms)
    
    @app.route('/custom_forms/create/<category>', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def create_custom_form(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        form = CreateFormForm()
        # Preseleccionar categoría y sección
        form.category.data = category
        form.section.data = 'nuevo_formulario'
        
        if form.validate_on_submit():
            # Crear nuevo formulario personalizado
            new_form = Form(
                title=form.title.data,
                description=form.description.data,
                category=form.category.data,
                section='nuevo_formulario',  # Sección fija para estos formularios
                is_active=form.is_active.data,
                created_by=current_user.id
            )
            
            db.session.add(new_form)
            db.session.commit()
            
            flash(f'Formulario personalizado "{form.title.data}" creado exitosamente. Ahora puedes agregar preguntas.', 'success')
            return redirect(url_for('edit_custom_form', form_id=new_form.id))
        
        return render_template('forms/create_custom_form.html',
                             title='Crear Formulario Personalizado',
                             form=form,
                             category=category)
    
    @app.route('/custom_forms/edit/<int:form_id>', methods=['GET', 'POST'])
    @login_required
    @admin_required
    def edit_custom_form(form_id):
        # Obtener el formulario
        form_db = Form.query.get_or_404(form_id)
        
        # Verificar si es un formulario personalizado
        if form_db.section != 'nuevo_formulario':
            flash('Este no es un formulario personalizado', 'danger')
            return redirect(url_for('custom_forms', category=form_db.category))
        
        # Verificar si es el creador o un administrador
        if form_db.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para editar este formulario', 'danger')
            return redirect(url_for('custom_forms', category=form_db.category))
        
        # Inicializar formulario para editar detalles generales
        form = CreateFormForm(obj=form_db)
        
        # Si se envió el formulario principal
        if form.validate_on_submit() and request.form.get('action') == 'update_details':
            form_db.title = form.title.data
            form_db.description = form.description.data
            form_db.category = form.category.data
            form_db.is_active = form.is_active.data
            db.session.commit()
            flash('Detalles del formulario actualizados correctamente', 'success')
            return redirect(url_for('edit_custom_form', form_id=form_id))
        
        # Obtener las preguntas ordenadas
        questions = FormQuestion.query.filter_by(form_id=form_id).order_by(FormQuestion.order).all()
        
        # Formulario para agregar/editar pregunta
        question_form = FormQuestionForm()
        
        # Lógica para agregar una nueva pregunta
        if question_form.validate_on_submit() and request.form.get('action') == 'add_question':
            # Obtener el orden máximo actual
            max_order = db.session.query(db.func.max(FormQuestion.order)).filter_by(form_id=form_id).scalar() or 0
            
            # Crear la nueva pregunta
            new_question = FormQuestion(
                form_id=form_id,
                question_text=question_form.question_text.data,
                question_type=question_form.question_type.data,
                is_required=question_form.is_required.data,
                order=max_order + 1
            )
            
            # Procesar opciones si es necesario
            if question_form.question_type.data in ['radio', 'checkbox', 'select'] and question_form.options.data:
                options_list = [option.strip() for option in question_form.options.data.split('\n') if option.strip()]
                new_question.set_options(options_list)
            
            db.session.add(new_question)
            db.session.commit()
            
            flash('Pregunta agregada correctamente', 'success')
            return redirect(url_for('edit_custom_form', form_id=form_id))
        
        # Lógica para manejar operaciones AJAX de las preguntas
        if request.method == 'POST' and request.form.get('action') == 'delete_question':
            question_id = request.form.get('question_id')
            if question_id:
                question = FormQuestion.query.get(question_id)
                if question and question.form_id == form_id:
                    db.session.delete(question)
                    db.session.commit()
                    flash('Pregunta eliminada correctamente', 'success')
                    return redirect(url_for('edit_custom_form', form_id=form_id, tab='questions'))
        
        return render_template('forms/edit_custom_form.html',
                             title=f'Editar Formulario Personalizado: {form_db.title}',
                             form=form,
                             form_db=form_db,
                             questions=questions,
                             question_form=question_form)
    
    @app.route('/custom_forms/view/<int:form_id>')
    @login_required
    def view_custom_form(form_id):
        # Obtener el formulario
        form = Form.query.get_or_404(form_id)
        
        # Verificar si es un formulario personalizado
        if form.section != 'nuevo_formulario':
            flash('Este no es un formulario personalizado', 'danger')
            return redirect(url_for('custom_forms', category=form.category))
        
        # Obtener las preguntas ordenadas
        questions = FormQuestion.query.filter_by(form_id=form_id).order_by(FormQuestion.order).all()
        
        return render_template('forms/view_custom_form.html',
                             title=form.title,
                             form=form,
                             questions=questions,
                             category=form.category)
    
    @app.route('/custom_forms/submit/<int:form_id>', methods=['POST'])
    @login_required
    def submit_custom_form(form_id):
        # Obtener el formulario
        form = Form.query.get_or_404(form_id)
        
        # Verificar si es un formulario personalizado
        if form.section != 'nuevo_formulario':
            flash('Este no es un formulario personalizado', 'danger')
            return redirect(url_for('custom_forms', category=form.category))
        
        # Crear una nueva respuesta
        response = FormResponse(
            form_id=form_id,
            user_id=current_user.id
        )
        
        db.session.add(response)
        db.session.commit()
        
        # Procesar respuestas para cada pregunta
        questions = FormQuestion.query.filter_by(form_id=form_id).all()
        for question in questions:
            # Obtener el valor de la respuesta del formulario
            answer_key = f'question_{question.id}'
            answer_value = ''
            
            if question.question_type == 'checkbox':
                # Para checkboxes, pueden venir múltiples valores
                values = request.form.getlist(answer_key)
                answer_value = ','.join(values) if values else ''
            else:
                # Para otros tipos de preguntas
                answer_value = request.form.get(answer_key, '')
            
            # Guardar la respuesta
            answer = FormAnswer(
                response_id=response.id,
                question_id=question.id,
                answer_text=answer_value
            )
            
            db.session.add(answer)
        
        db.session.commit()
        
        flash('Formulario enviado correctamente', 'success')
        return redirect(url_for('custom_forms', category=form.category))
    
    @app.route('/custom_forms/delete/<int:form_id>', methods=['POST'])
    @login_required
    @admin_required
    def delete_custom_form(form_id):
        # Obtener el formulario
        form = Form.query.get_or_404(form_id)
        
        # Verificar si es un formulario personalizado
        if form.section != 'nuevo_formulario':
            flash('Este no es un formulario personalizado', 'danger')
            return redirect(url_for('custom_forms', category=form.category))
        
        # Verificar si es el creador o un administrador
        if form.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para eliminar este formulario', 'danger')
            return redirect(url_for('custom_forms', category=form.category))
        
        category = form.category  # Guardar la categoría antes de eliminar
        
        # Eliminar el formulario y todas sus preguntas y respuestas (cascada)
        db.session.delete(form)
        db.session.commit()
        
        flash('Formulario eliminado correctamente', 'success')
        return redirect(url_for('custom_forms', category=category))
    
    @app.route('/custom_forms/responses/<int:form_id>')
    @login_required
    @admin_required
    def view_custom_responses(form_id):
        # Obtener el formulario
        form = Form.query.get_or_404(form_id)
        
        # Verificar si es un formulario personalizado
        if form.section != 'nuevo_formulario':
            flash('Este no es un formulario personalizado', 'danger')
            return redirect(url_for('custom_forms', category=form.category))
        
        # Verificar si es el creador o un administrador
        if form.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para ver las respuestas de este formulario', 'danger')
            return redirect(url_for('custom_forms', category=form.category))
        
        # Obtener las respuestas
        responses = FormResponse.query.filter_by(form_id=form_id).order_by(FormResponse.submitted_at.desc()).all()
        
        # Obtener preguntas para mostrar cabeceras
        questions = FormQuestion.query.filter_by(form_id=form_id).order_by(FormQuestion.order).all()
        
        return render_template('forms/view_custom_responses.html',
                             title=f'Respuestas: {form.title}',
                             form=form,
                             questions=questions,
                             responses=responses,
                             category=form.category)
    
    @app.route('/custom_forms/all_responses/<category>')
    @login_required
    @admin_required
    def view_all_custom_responses(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener todos los formularios personalizados para esta categoría
        forms = Form.query.filter_by(category=category, section='nuevo_formulario').all()
        
        # Obtener todas las respuestas de todos los formularios de esta categoría
        # Primero, obtenemos los IDs de los formularios
        form_ids = [form.id for form in forms]
        
        # Si no hay formularios, mostrar mensaje y redirigir
        if not form_ids:
            flash('No hay formularios personalizados para esta categoría', 'info')
            return redirect(url_for('custom_forms', category=category))
        
        # Obtener todas las respuestas para estos formularios
        responses_by_form = {}
        questions_by_form = {}
        
        for form_id in form_ids:
            # Obtener las respuestas para este formulario
            responses = FormResponse.query.filter_by(form_id=form_id).order_by(FormResponse.submitted_at.desc()).all()
            responses_by_form[form_id] = responses
            
            # Obtener las preguntas para este formulario
            questions = FormQuestion.query.filter_by(form_id=form_id).order_by(FormQuestion.order).all()
            questions_by_form[form_id] = questions
        
        return render_template('forms/all_custom_responses.html',
                             title=f'Todas las Respuestas - {category}',
                             category=category,
                             forms=forms,
                             responses_by_form=responses_by_form,
                             questions_by_form=questions_by_form)

    # Rutas para PNC
    @app.route('/pnc/<category>', methods=['GET'])
    @login_required
    def list_pnc(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener todos los registros PNC para esta categoría
        pnc_records = PNC.query.filter_by(categoria=category).order_by(PNC.created_at.desc()).all()
        
        # Si estamos en la categoría TORTILLA, también obtener los registros de calidad de tortillas
        calidad_tortilla_records = None
        if category == 'TORTILLA':
            calidad_tortilla_records = CalidadTortilla.query.order_by(CalidadTortilla.created_at.desc()).all()
        
        return render_template('pnc/list_pnc.html',
                               title=f'Registros PNC - {category}',
                               category=category,
                               pnc_records=pnc_records,
                               calidad_tortilla_records=calidad_tortilla_records)
    
    @app.route('/pnc/<category>/create', methods=['GET', 'POST'])
    @login_required
    def create_pnc(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        pnc_form = PNCForm()
        analisis_form = AnalisisCalidadForm()
        
        # Generar un prefijo de folio para ambos formularios basado en la fecha y categoría
        # Usamos el módulo datetime importado al inicio del archivo
        fecha_actual = datetime.now()
        fecha_str = fecha_actual.strftime('%d%m')
        
        # Definir el sufijo según la categoría
        categoria_sufijo = ''
        if category == 'PAPA':
            categoria_sufijo = 'PA'
        elif category == 'EXTRUIDOS':
            categoria_sufijo = 'EX'
        elif category == 'TORTILLA':
            categoria_sufijo = 'TO'
        else:
            categoria_sufijo = 'XX'
            
        # Crear prefijos para cada tipo de folio
        pnc_prefix = f"PNC_{fecha_str}_{categoria_sufijo}_"
        analisis_prefix = f"ANL_{fecha_str}_{categoria_sufijo}_"
        
        # Procesar el formulario PNC
        if 'submit_pnc' in request.form and pnc_form.validate_on_submit():
            # Crear nuevo registro PNC
            new_pnc = PNC(
                folio=pnc_form.folio.data,
                fecha=pnc_form.fecha.data,
                turno=pnc_form.turno.data,
                producto=pnc_form.producto.data,
                horario=pnc_form.horario.data,
                cantidad=pnc_form.cantidad.data,
                origen=pnc_form.origen.data,
                no_conformidad=pnc_form.no_conformidad.data,
                status=pnc_form.status.data,
                detector=pnc_form.detector.data,
                rechazo=pnc_form.rechazo.data,
                categoria=category,
                created_by=current_user.id
            )
            
            db.session.add(new_pnc)
            db.session.commit()
            
            flash(f'Registro PNC con folio {pnc_form.folio.data} creado exitosamente.', 'success')
            return redirect(url_for('list_pnc', category=category))
        
        # Procesar el formulario de Análisis de Calidad
        if 'submit_analisis' in request.form and analisis_form.validate_on_submit():
            # Crear nuevo registro de análisis
            new_analisis = AnalisisCalidad(
                folio=analisis_form.folio.data,
                fecha=analisis_form.fecha.data,
                turno=analisis_form.turno.data,
                producto=analisis_form.producto.data,
                horario=analisis_form.horario.data,
                lote=analisis_form.lote.data,
                analista=analisis_form.analista.data,
                peso=analisis_form.peso.data,
                observaciones=analisis_form.observaciones.data,
                categoria=category,
                created_by=current_user.id
            )
            
            db.session.add(new_analisis)
            db.session.commit()
            
            flash(f'Registro de Análisis de Calidad con folio {analisis_form.folio.data} creado exitosamente.', 'success')
            return redirect(url_for('list_analisis_calidad', category=category))
        
        # Prepopular los campos de fecha y asignar folios sugeridos
        if request.method == 'GET':
            # Asignar fecha actual a los formularios
            pnc_form.fecha.data = fecha_actual.date()
            analisis_form.fecha.data = fecha_actual.date()
            
            # Sugerir folios
            # Para PNC - buscar el último número de folio para la categoría actual
            pnc_ultimo_numero = 0
            existing_pnc_folios = PNC.query.filter(
                PNC.folio.like(f"{pnc_prefix}%")
            ).all()
            
            patron_pnc = re.compile(f"{pnc_prefix}(\\d+)")
            for pnc in existing_pnc_folios:
                match = patron_pnc.match(pnc.folio)
                if match:
                    try:
                        num = int(match.group(1))
                        if num > pnc_ultimo_numero:
                            pnc_ultimo_numero = num
                    except ValueError:
                        pass
            
            # Para Análisis - buscar el último número de folio
            analisis_ultimo_numero = 0
            existing_analisis_folios = AnalisisCalidad.query.filter(
                AnalisisCalidad.folio.like(f"{analisis_prefix}%")
            ).all()
            
            patron_analisis = re.compile(f"{analisis_prefix}(\\d+)")
            for analisis in existing_analisis_folios:
                match = patron_analisis.match(analisis.folio)
                if match:
                    try:
                        num = int(match.group(1))
                        if num > analisis_ultimo_numero:
                            analisis_ultimo_numero = num
                    except ValueError:
                        pass
            
            # Asignar folios sugeridos
            pnc_form.folio.data = f"{pnc_prefix}{pnc_ultimo_numero + 1:03d}"
            analisis_form.folio.data = f"{analisis_prefix}{analisis_ultimo_numero + 1:03d}"
        
        return render_template('pnc/create_pnc.html',
                               title=f'Nuevo Registro PNC - {category}',
                               pnc_form=pnc_form,
                               analisis_form=analisis_form,
                               category=category)
    
    @app.route('/pnc/<category>/edit/<int:pnc_id>', methods=['GET', 'POST'])
    @login_required
    def edit_pnc(category, pnc_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro PNC
        pnc = PNC.query.get_or_404(pnc_id)
        
        # Verificar que el PNC pertenece a la categoría correcta
        if pnc.categoria != category:
            flash('Registro PNC no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_pnc', category=category))
        
        # Verificar si es el creador o un administrador
        if pnc.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para editar este registro', 'danger')
            return redirect(url_for('list_pnc', category=category))
        
        form = PNCForm(obj=pnc)
        
        if form.validate_on_submit():
            # Actualizar el registro PNC
            form.populate_obj(pnc)
            db.session.commit()
            
            flash('Registro PNC actualizado correctamente', 'success')
            return redirect(url_for('list_pnc', category=category))
        
        return render_template('pnc/edit_pnc.html',
                               title=f'Editar Registro PNC - {pnc.folio}',
                               form=form,
                               pnc=pnc,
                               category=category)
    
    @app.route('/pnc/<category>/view/<int:pnc_id>')
    @login_required
    def view_pnc(category, pnc_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro PNC
        pnc = PNC.query.get_or_404(pnc_id)
        
        # Verificar que el PNC pertenece a la categoría correcta
        if pnc.categoria != category:
            flash('Registro PNC no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_pnc', category=category))
        
        return render_template('pnc/view_pnc.html',
                               title=f'Ver Registro PNC - {pnc.folio}',
                               pnc=pnc,
                               category=category)
    
    @app.route('/pnc/<category>/delete/<int:pnc_id>', methods=['POST'])
    @login_required
    def delete_pnc(category, pnc_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro PNC
        pnc = PNC.query.get_or_404(pnc_id)
        
        # Verificar que el PNC pertenece a la categoría correcta
        if pnc.categoria != category:
            flash('Registro PNC no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_pnc', category=category))
        
        # Verificar si es el creador o un administrador
        if pnc.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para eliminar este registro', 'danger')
            return redirect(url_for('list_pnc', category=category))
        
        db.session.delete(pnc)
        db.session.commit()
        
        flash('Registro PNC eliminado correctamente', 'success')
        return redirect(url_for('list_pnc', category=category))
        
    # Rutas para PNC Simple
    @app.route('/pnc_simple/<category>', methods=['GET'])
    @login_required
    def list_pnc_simple(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        try:
            # Intentar obtener todos los registros PNC Simple para esta categoría
            pnc_records = PNCSimple.query.filter_by(categoria=category).order_by(PNCSimple.created_at.desc()).all()
        except Exception as e:
            # Si hay un error (como columnas faltantes), mostrar un mensaje de advertencia
            flash('Se requiere actualizar la base de datos. Por favor ejecute el script update_db.py', 'warning')
            pnc_records = []
        
        # Crear un formulario para CSRF
        form = PNCSimpleForm()
        
        # Preparar datos para la pestaña de resultados
        productos_unicos = set()
        unidades_por_producto = {}

        # Recopilar datos de los registros existentes
        for record in pnc_records:
            if record.producto:
                productos_unicos.add(record.producto)
                
                # Agrupar por producto y unidad
                if record.producto not in unidades_por_producto:
                    unidades_por_producto[record.producto] = {}
                
                # Asegurar que existan las unidades estándar en cada producto
                for unidad_estandar in ['KILOGRAMOS', 'TONELADAS', 'TARIMAS']:
                    if unidad_estandar not in unidades_por_producto[record.producto]:
                        unidades_por_producto[record.producto][unidad_estandar] = 0
                
                # Obtener la unidad del registro (o usar KILOGRAMOS como predeterminada)
                unidad = record.unidad_cantidad
                if not unidad:
                    unidad = 'KILOGRAMOS'  # Valor predeterminado
                
                # Normalizar el nombre de la unidad (para evitar duplicados con mayúsculas/minúsculas)
                unidad = unidad.upper()
                
                # Asegurarse de que la unidad exista en el diccionario
                if unidad not in unidades_por_producto[record.producto]:
                    unidades_por_producto[record.producto][unidad] = 0
                
                # Sumar la cantidad si existe
                if record.cantidad is not None:
                    unidades_por_producto[record.producto][unidad] += record.cantidad
        # Para las demás categorías, usar la plantilla estándar
        return render_template('pnc/list_pnc_simple.html',
                                title=f'Registros PNC Simple - {category}',
                                category=category,
                                pnc_records=pnc_records,
                                productos_unicos=list(productos_unicos),
                                unidades_por_producto=unidades_por_producto,
                                form=form)
    
    @app.route('/pnc_simple/<category>/create', methods=['GET', 'POST'])
    @login_required
    def create_pnc_simple(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        form = PNCSimpleForm()
        
        # Opciones dinámicas según la categoría - ANTES DE VALIDACIÓN
        if category == 'TORTILLA':
            form.producto.choices = [
                ('DORITOS', 'DORITOS'),
                ('TOSTITOS SALSA VERDE', 'TOSTITOS SALSA VERDE'),
                ('TOSTITOS FH', 'TOSTITOS FH'),
                ('DORITOS PIZZEROLA', 'DORITOS PIZZEROLA'),
                ('DORITOS FH', 'DORITOS FH'),
                ('RANCHERITOS', 'RANCHERITOS'),
                ('DORITOS INCOGNITA', 'DORITOS INCOGNITA')
                
            ]
        elif category == 'EXTRUIDOS':
            form.producto.choices = [
                ('CHEETOS TORCIDITOS', 'CHEETOS TORCIDITOS'),
                ('CHEETOS XTRA FLAMIN HOT', 'CHEETOS XTRA FLAMIN HOT'),
                ('CHEETOS JALAPENO', 'CHEETOS JALAQUEÑO'),
                ('CHEETOS XTRA FH NUEVO', 'CHEETOS XTRA FH NUEVO'),
                ('OTROS', 'OTROS')
            ]
        elif category == 'PAPA':
            form.producto.choices = [
                ('PAPA SAL', 'PAPA SAL'),
                ('SABRITAS LIMON', 'SABRITAS LIMON'),
                ('RUFFLES QUESO', 'RUFFLES QUESO'),
                ('RUFFLES SAL', 'RUFFLES SAL'),
                ('SABRITAS XTRA FH', 'SABRITAS XTRA FH')
            ]
        else:
            form.producto.choices = [('', 'Seleccionar...')]

        if form.validate_on_submit():
            # Generar el folio automáticamente con el nuevo formato: PNC_DDMM_XX_001
            # Usar la fecha actual del servidor para asegurar consistencia
            fecha_actual = datetime.now()
            fecha_str = fecha_actual.strftime('%d%m')  # Solo día y mes
            categoria_sufijo = ''
            if category == 'PAPA':
                categoria_sufijo = 'PA'
            elif category == 'EXTRUIDOS':
                categoria_sufijo = 'EX'
            elif category == 'TORTILLA':
                categoria_sufijo = 'TO'
            else:
                categoria_sufijo = 'XX'
            
            # Base del folio sin el número secuencial
            folio_base = f"PNC_{fecha_str}_{categoria_sufijo}"
            
            # Buscar el último número de folio para este día y categoría
            # Obtener todos los folios que coincidan con el patrón base
            existing_folios = PNCSimple.query.filter(
                PNCSimple.folio.like(f"{folio_base}_%")
            ).all()
            
            # Encontrar el número más alto
            ultimo_numero = 0
            patron = re.compile(f"{folio_base}_(\\d+)")
            
            for pnc in existing_folios:
                match = patron.match(pnc.folio)
                if match:
                    try:
                        num = int(match.group(1))
                        if num > ultimo_numero:
                            ultimo_numero = num
                    except ValueError:
                        pass
            
            # Incrementar para el nuevo folio y formatear con ceros a la izquierda
            numero_folio = f"{ultimo_numero + 1:03d}"  # Formato 001, 002, etc.
            folio = f"{folio_base}_{numero_folio}"
            
            # Crear nuevo registro PNC Simple
            new_pnc = PNCSimple(
                folio=folio,
                fecha=fecha,
                turno=turno,
                producto=producto,
                horario=horario,
                cantidad=cantidad_val,
                unidad_cantidad=request.form.get('unidad_cantidad', ''),
                origen=request.form.get('origen', ''),
                no_conformidad=request.form.get('no_conformidad', ''),
                status=request.form.get('status', ''),
                detector=request.form.get('detector', ''),
                rechazo=True if request.form.get('rechazo') == 'true' else False,
                categoria=category,
                created_by=current_user.id
            )
            
            try:
                db.session.add(new_pnc)
                db.session.commit()
                flash(f'Registro PNC Simple con folio {folio} creado exitosamente.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error al guardar el registro: {str(e)}', 'danger')
            
            return redirect(url_for('list_pnc_simple', category=category))
        
        # Para peticiones POST sin validación (desde el modal)
        if request.method == 'POST':
            # AGREGAR OPCIONES DINÁMICAS TAMBIÉN AQUÍ
            if category == 'TORTILLA':
                form.producto.choices = [
                    ('DORITOS', 'DORITOS'),
                    ('TOSTITOS SALSA VERDE', 'TOSTITOS SALSA VERDE'),
                    ('TOSTITOS FH', 'TOSTITOS FH'),
                    ('DORITOS PIZZEROLA', 'DORITOS PIZZEROLA'),
                    ('DORITOS FH', 'DORITOS FH'),
                    ('RANCHERITOS', 'RANCHERITOS'),
                    ('DORITOS INCOGNITA', 'DORITOS INCOGNITA')
                ]
            elif category == 'EXTRUIDOS':
                form.producto.choices = [
                    ('TORCIDITOS', 'CHEETOS TORCIDITOS'),
                    ('CHEETOS XTRA FLAMIN HOT', 'CHEETOS XTRA FLAMIN HOT'),
                    ('CHEETOS JALAPENO', 'CHEETOS JALAQUEÑO'),
                    ('CHEETOS XTRA FH NUEVO', 'CHEETOS XTRA FH NUEVO'),
                    ('OTROS', 'OTROS')
                ]
            elif category == 'PAPA':
                form.producto.choices = [
                    ('PAPA SAL', 'PAPA SAL'),
                    ('SABRITAS LIMON', 'SABRITAS LIMON'),
                    ('RUFFLES SAL', 'RUFFLES SAL'),
                    ('RUFFLES QUESO', 'RUFFLES QUESO'),
                    ('SABRITAS XTRA FH', 'SABRITAS XTRA FH')

                ]
            else:
                form.producto.choices = [('', 'Seleccionar...')]
                
            # Aunque recibimos la fecha del formulario, usamos la fecha actual del servidor
            # para asegurar consistencia en el folio
            fecha_actual = datetime.now()
            fecha_str = fecha_actual.strftime('%d%m')  # Solo día y mes
            # Mantenemos la fecha del formulario para el registro
            fecha = datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date()
            
            # Generar el folio automáticamente con el nuevo formato: PNC_DDMM_XX_001
            categoria_sufijo = ''
            if category == 'PAPA':
                categoria_sufijo = 'PA'
            elif category == 'EXTRUIDOS':
                categoria_sufijo = 'EX'
            elif category == 'TORTILLA':
                categoria_sufijo = 'TO'
            else:
                categoria_sufijo = 'XX'
            
            # Base del folio sin el número secuencial
            folio_base = f"PNC_{fecha_str}_{categoria_sufijo}"
            
            # Buscar el último número de folio para este día y categoría
            # Obtener todos los folios que coincidan con el patrón base
            existing_folios = PNCSimple.query.filter(
                PNCSimple.folio.like(f"{folio_base}_%")
            ).all()
            
            # Encontrar el número más alto
            ultimo_numero = 0
            patron = re.compile(f"{folio_base}_(\\d+)")
            
            for pnc in existing_folios:
                match = patron.match(pnc.folio)
                if match:
                    try:
                        num = int(match.group(1))
                        if num > ultimo_numero:
                            ultimo_numero = num
                    except ValueError:
                        pass
            
            # Incrementar para el nuevo folio y formatear con ceros a la izquierda
            numero_folio = f"{ultimo_numero + 1:03d}"  # Formato 001, 002, etc.
            folio = f"{folio_base}_{numero_folio}"
            
            turno = request.form.get('turno')
            producto = request.form.get('producto')
            # Si se enviaron los campos de hora_inicio y hora_fin, usar esos para formar el horario
            if 'hora_inicio' in request.form and 'hora_fin' in request.form:
                hora_inicio = request.form.get('hora_inicio')
                hora_fin = request.form.get('hora_fin')
                horario = f"{hora_inicio}:00 - {hora_fin}:00"
            else:
                horario = request.form.get('horario', '')
            
            # Convertir cantidad a numérico si existe
            try:
                cantidad_val = float(request.form.get('cantidad', '0')) if request.form.get('cantidad') else None
            except ValueError:
                cantidad_val = None
                
            new_pnc = PNCSimple(
                folio=folio,
                fecha=fecha,
                turno=turno,
                producto=producto,
                horario=horario,
                cantidad=cantidad_val,
                unidad_cantidad=request.form.get('unidad_cantidad', ''),
                origen=request.form.get('origen', ''),
                no_conformidad=request.form.get('no_conformidad', ''),
                status=request.form.get('status', ''),
                detector=request.form.get('detector', ''),
                rechazo=True if request.form.get('rechazo') == 'true' else False,
                categoria=category,
                created_by=current_user.id
            )
            
            try:
                db.session.add(new_pnc)
                db.session.commit()
                flash(f'Registro PNC Simple con folio {folio} creado exitosamente.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error al guardar el registro: {str(e)}', 'danger')
            
            return redirect(url_for('list_pnc_simple', category=category))
        
        return render_template('pnc/create_pnc_simple.html',
                               title=f'Nuevo Registro PNC Simple - {category}',
                               form=form,
                               category=category)
    
    @app.route('/pnc_simple/<category>/edit/<int:pnc_id>', methods=['GET', 'POST'])
    @login_required
    def edit_pnc_simple(category, pnc_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro PNC Simple
        pnc = PNCSimple.query.get_or_404(pnc_id)
        
        # Verificar que el PNC pertenece a la categoría correcta
        if pnc.categoria != category:
            flash('Registro PNC Simple no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_pnc_simple', category=category))
        
        # Verificar si es el creador o un administrador
        if pnc.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para editar este registro', 'danger')
            return redirect(url_for('list_pnc_simple', category=category))
        
        form = PNCSimpleForm(obj=pnc)
        
        if form.validate_on_submit():
            # Obtener el folio actual para comparación
            folio_actual = pnc.folio
            
            # Actualizar la fecha y otros campos excepto el folio
            pnc.fecha = form.fecha.data
            pnc.turno = form.turno.data
            pnc.producto = form.producto.data
            pnc.horario = form.horario.data
            
            # Actualizar el folio si ha cambiado
            nuevo_folio = form.folio.data
            if nuevo_folio != folio_actual:
                # Verificar si el nuevo folio ya existe
                if PNCSimple.query.filter_by(folio=nuevo_folio).first():
                    flash(f'Error: Ya existe un registro con el folio {nuevo_folio}. No se puede actualizar.', 'danger')
                    return redirect(url_for('edit_pnc_simple', category=category, pnc_id=pnc_id))
                else:
                    pnc.folio = nuevo_folio
            
            try:
                db.session.commit()
                flash('Registro PNC Simple actualizado correctamente', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error al actualizar el registro: {str(e)}', 'danger')
                return redirect(url_for('edit_pnc_simple', category=category, pnc_id=pnc_id))
            
            return redirect(url_for('list_pnc_simple', category=category))
            
        # Para peticiones POST sin validación (desde el modal)
        if request.method == 'POST':
            try:
                # Obtener el folio actual para comparación
                folio_actual = pnc.folio
                
                # Actualizar fecha y otros campos
                pnc.fecha = datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date()
                pnc.turno = request.form.get('turno')
                pnc.producto = request.form.get('producto')
                pnc.horario = request.form.get('horario')
                # Convertir cantidad a numérico si existe
                try:
                    cantidad_val = float(request.form.get('cantidad', '0')) if request.form.get('cantidad') else None
                    pnc.cantidad = cantidad_val
                except ValueError:
                    pnc.cantidad = None
                pnc.unidad_cantidad = request.form.get('unidad_cantidad', '')
                pnc.origen = request.form.get('origen', '')
                pnc.no_conformidad = request.form.get('no_conformidad', '')
                pnc.status = request.form.get('status', '')
                pnc.detector = request.form.get('detector', '')
                pnc.rechazo = True if request.form.get('rechazo') == 'true' else False
                
                # Actualizar el folio si ha cambiado
                nuevo_folio = request.form.get('folio')
                if nuevo_folio != folio_actual:
                    # Verificar si el nuevo folio ya existe
                    if PNCSimple.query.filter_by(folio=nuevo_folio).first():
                        flash(f'Error: Ya existe un registro con el folio {nuevo_folio}. No se puede actualizar.', 'danger')
                        return redirect(url_for('list_pnc_simple', category=category))
                    else:
                        pnc.folio = nuevo_folio
                
                db.session.commit()
                flash('Registro PNC Simple actualizado correctamente', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error al actualizar el registro: {str(e)}', 'danger')
            
            return redirect(url_for('list_pnc_simple', category=category))
        
        return render_template('pnc/edit_pnc_simple.html',
                               title=f'Editar Registro PNC Simple - {pnc.folio}',
                               form=form,
                               pnc=pnc,
                               category=category)
    
    @app.route('/pnc_simple/<category>/delete/<int:pnc_id>', methods=['POST'])
    @login_required
    def delete_pnc_simple(category, pnc_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro PNC Simple
        pnc = PNCSimple.query.get_or_404(pnc_id)
        
        # Verificar que el PNC pertenece a la categoría correcta
        if pnc.categoria != category:
            flash('Registro PNC Simple no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_pnc_simple', category=category))
        
        # Verificar si es el creador o un administrador
        if pnc.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para eliminar este registro', 'danger')
            return redirect(url_for('list_pnc_simple', category=category))
        
        db.session.delete(pnc)
        db.session.commit()
        
        flash('Registro PNC Simple eliminado correctamente', 'success')
        return redirect(url_for('list_pnc_simple', category=category))
    
    # Rutas para el formulario de Calidad de Tortillas integrado en PNC
    @app.route('/calidad_tortilla/create', methods=['GET', 'POST'])
    @login_required
    def create_calidad_tortilla():
        form = CalidadTortillaForm()
        
        if form.validate_on_submit():
            # Crear nuevo registro
            new_registro = CalidadTortilla(
                folio=form.folio.data,
                fecha=form.fecha.data,
                turno=form.turno.data,
                producto=form.producto.data,
                horario=form.horario.data,
                created_by=current_user.id
            )
            
            db.session.add(new_registro)
            db.session.commit()
            
            flash(f'Registro con folio {form.folio.data} creado exitosamente.', 'success')
            return redirect(url_for('list_pnc', category='TORTILLA'))
        
        return render_template('calidad_tortilla/create.html',
                               title='Nuevo Registro de Calidad - TORTILLA',
                               form=form,
                               category='TORTILLA')
    
    @app.route('/calidad_tortilla/edit/<int:registro_id>', methods=['GET', 'POST'])
    @login_required
    def edit_calidad_tortilla(registro_id):
        # Obtener el registro
        registro = CalidadTortilla.query.get_or_404(registro_id)
        
        # Verificar si es el creador o un administrador
        if registro.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para editar este registro', 'danger')
            return redirect(url_for('list_pnc', category='TORTILLA'))
        
        form = CalidadTortillaForm(obj=registro)
        
        if form.validate_on_submit():
            # Actualizar el registro
            form.populate_obj(registro)
            db.session.commit()
            
            flash('Registro actualizado correctamente', 'success')
            return redirect(url_for('list_pnc', category='TORTILLA'))
        
        return render_template('calidad_tortilla/edit.html',
                               title=f'Editar Registro - {registro.folio}',
                               form=form,
                               registro=registro,
                               category='TORTILLA')
    
    @app.route('/calidad_tortilla/delete/<int:registro_id>', methods=['POST'])
    @login_required
    def delete_calidad_tortilla(registro_id):
        # Obtener el registro
        registro = CalidadTortilla.query.get_or_404(registro_id)
        
        # Verificar si es el creador o un administrador
        if registro.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para eliminar este registro', 'danger')
            return redirect(url_for('list_pnc', category='TORTILLA'))
        
        db.session.delete(registro)
        db.session.commit()
        
        flash('Registro eliminado correctamente', 'success')
        return redirect(url_for('list_pnc', category='TORTILLA'))

    @app.route('/laboratorio/<category>')
    @login_required
    def laboratorio(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        return render_template('dashboard/laboratorio.html', category=category)

    # Rutas para Análisis Fisicoquímicos
    @app.route('/analisis_fisicoquimicos/<category>', methods=['GET', 'POST'])
    @login_required
    def list_analisis_fisicoquimicos(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Procesar el formulario de creación de análisis fisicoquímico
        if request.method == 'POST' and 'submit_analisis' in request.form:
            # Verificar que el campo producto esté presente y no esté vacío
            if not request.form.get('producto'):
                flash('Error: El campo Producto es obligatorio y no puede estar vacío.', 'danger')
                return redirect(url_for('list_analisis_fisicoquimicos', category=category))
            
            # Generar el folio automáticamente con el formato: ANL_DDMM_XX_001
            fecha_actual = datetime.now()
            fecha_str = fecha_actual.strftime('%d%m')
            
            # Definir el sufijo según la categoría
            categoria_sufijo = ''
            if category == 'PAPA':
                categoria_sufijo = 'PA'
            elif category == 'EXTRUIDOS':
                categoria_sufijo = 'EX'
            elif category == 'TORTILLA':
                categoria_sufijo = 'TO'
            else:
                categoria_sufijo = 'XX'
            
            # Base del folio sin el número secuencial
            folio_base = f"ANL_{fecha_str}_{categoria_sufijo}"
            
            # Buscar el último número de folio para este día y categoría
            existing_folios = AnalisisCalidad.query.filter(
                AnalisisCalidad.folio.like(f"{folio_base}_%")
            ).all()
            
            # Encontrar el número más alto
            ultimo_numero = 0
            patron = re.compile(f"{folio_base}_(\\d+)")
            
            for analisis in existing_folios:
                match = patron.match(analisis.folio)
                if match:
                    try:
                        num = int(match.group(1))
                        if num > ultimo_numero:
                            ultimo_numero = num
                    except ValueError:
                        pass
            
            # Incrementar para el nuevo folio y formatear con ceros a la izquierda
            numero_folio = f"{ultimo_numero + 1:03d}"  # Formato 001, 002, etc.
            folio = f"{folio_base}_{numero_folio}"
            
            # Crear nuevo registro de análisis - sin campos lote, analista y peso para TORTILLA
            new_analisis = AnalisisCalidad(
                folio=folio,
                fecha=datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date(),
                turno=request.form.get('turno'),
                producto=request.form.get('producto'),
                horario=request.form.get('horario'),
                lote='' if category == 'TORTILLA' else request.form.get('lote', ''),
                analista='' if category == 'TORTILLA' else request.form.get('analista', ''),
                peso=None if category == 'TORTILLA' else request.form.get('peso'),
                humedad_base_frita=request.form.get('humedad_base_frita', ''),
                aceite_base_frita=request.form.get('aceite_base_frita', ''),
                # Campos PAPA - Cloruros y Sal Titulador
                cloruros_base=request.form.get('cloruros_base', ''),
                tanque1_aceite_pt=request.form.get('tanque1_aceite_pt', ''),
                tanque1_humedad_pt=request.form.get('tanque1_humedad_pt', ''),
                tanque1_sal_titulador=request.form.get('tanque1_sal_titulador', ''),
                tanque1_sal_pt=request.form.get('tanque1_sal_pt', ''),
                tanque2_aceite_pt=request.form.get('tanque2_aceite_pt', ''),
                tanque2_humedad_pt=request.form.get('tanque2_humedad_pt', ''),
                tanque2_sal_titulador=request.form.get('tanque2_sal_titulador', ''),
                tanque2_sal_pt=request.form.get('tanque2_sal_pt', ''),
                tanque3_aceite_pt=request.form.get('tanque3_aceite_pt', ''),
                tanque3_humedad_pt=request.form.get('tanque3_humedad_pt', ''),
                tanque3_sal_titulador=request.form.get('tanque3_sal_titulador', ''),
                tanque3_sal_pt=request.form.get('tanque3_sal_pt', ''),
                # CAMPOS PT PRODUCTO TERMINADO
                aceite_pt_producto_terminado=request.form.get('aceite_pt_producto_terminado', ''),
                humedad_pt_producto_terminado=request.form.get('humedad_pt_producto_terminado', ''),
                sal_pt_producto_terminado=request.form.get('sal_pt_producto_terminado', ''),
                observaciones=request.form.get('observaciones', ''),
                categoria=category,
                created_by=current_user.id
            )
            
            try:
                db.session.add(new_analisis)
                db.session.commit()
                flash(f'Registro de Análisis Fisicoquímico con folio {folio} creado exitosamente.', 'success')
                
                # Cerrar el modal después de guardar (usando JavaScript)
                # Este código se ejecutará automáticamente por el navegador
                response = redirect(url_for('list_analisis_fisicoquimicos', category=category))
                response.headers['X-Modal-Close'] = 'true'
                return response
            except Exception as e:
                db.session.rollback()
                flash(f'Error al guardar el registro: {str(e)}', 'danger')
            
            # Redireccionar a la misma página para mostrar el listado actualizado
            return redirect(url_for('list_analisis_fisicoquimicos', category=category))
        
        # ...existing code...
        if request.method == 'POST' and 'submit_edit_analisis' in request.form:
            analisis_id = request.form.get('edit_id')
            analisis = AnalisisCalidad.query.get_or_404(analisis_id)
            # Permitir solo al creador o admin editar
            if analisis.created_by != current_user.id and not current_user.is_admin:
                flash('No tienes permiso para editar este registro', 'danger')
                return redirect(url_for('list_analisis_fisicoquimicos', category=category))
            # Actualizar campos
            analisis.fecha = datetime.strptime(request.form.get('fecha'), '%Y-%m-%d').date()
            analisis.turno = request.form.get('turno')
            analisis.producto = request.form.get('producto')
            analisis.horario = request.form.get('horario')
           
            analisis.humedad_base_frita = request.form.get('humedad_base_frita')
            analisis.aceite_base_frita = request.form.get('aceite_base_frita')
            # Campos PAPA - Cloruros y Sal Titulador
            analisis.cloruros_base = request.form.get('cloruros_base')
            analisis.tanque1_aceite_pt = request.form.get('tanque1_aceite_pt')
            analisis.tanque1_humedad_pt = request.form.get('tanque1_humedad_pt')
            analisis.tanque1_sal_titulador = request.form.get('tanque1_sal_titulador')
            analisis.tanque1_sal_pt = request.form.get('tanque1_sal_pt')
            analisis.tanque2_aceite_pt = request.form.get('tanque2_aceite_pt')
            analisis.tanque2_humedad_pt = request.form.get('tanque2_humedad_pt')
            analisis.tanque2_sal_titulador = request.form.get('tanque2_sal_titulador')
            analisis.tanque2_sal_pt = request.form.get('tanque2_sal_pt')
            analisis.tanque3_aceite_pt = request.form.get('tanque3_aceite_pt')
            analisis.tanque3_humedad_pt = request.form.get('tanque3_humedad_pt')
            analisis.tanque3_sal_titulador = request.form.get('tanque3_sal_titulador')
            analisis.tanque3_sal_pt = request.form.get('tanque3_sal_pt')
            # CAMPOS PT PRODUCTO TERMINADO
            analisis.aceite_pt_producto_terminado = request.form.get('aceite_pt_producto_terminado')
            analisis.humedad_pt_producto_terminado = request.form.get('humedad_pt_producto_terminado')
            analisis.sal_pt_producto_terminado = request.form.get('sal_pt_producto_terminado')
            analisis.observaciones = request.form.get('observaciones')
            try:
                db.session.commit()
                flash('Registro de análisis actualizado correctamente.', 'success')
            except Exception as e:
                db.session.rollback()
                flash(f'Error al actualizar el registro: {str(e)}', 'danger')
            return redirect(url_for('list_analisis_fisicoquimicos', category=category))

        # Paginación server-side para evitar cargar todos los registros
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        # Limitar per_page a valores válidos
        if per_page not in [20, 50, 100, 200]:
            per_page = 50

        pagination = AnalisisCalidad.query.filter_by(categoria=category)\
            .order_by(AnalisisCalidad.created_at.desc())\
            .paginate(page=page, per_page=per_page, error_out=False)

        analisis_records = pagination.items

        # Obtener lista de documentos de la carpeta Doc
        documentos_lista = obtener_documentos_fisicoquimicos()

        return render_template('pnc/list_analisis_fisicoquimicos.html',
                               title=f'Análisis Fisicoquímicos - {category}',
                               category=category,
                               analisis_records=analisis_records,
                               pagination=pagination,
                               current_page=page,
                               per_page=per_page,
                               documentos_lista=documentos_lista)

    # Función auxiliar para generar thumbnail de un PDF
    def generar_thumbnail_pdf(pdf_path, thumbnail_path, width=300):
        """Genera una imagen thumbnail de la primera página de un PDF"""
        import os
        try:
            import fitz  # PyMuPDF
            doc = fitz.open(pdf_path)
            page = doc[0]  # Primera página

            # Calcular el zoom para obtener el ancho deseado
            zoom = width / page.rect.width
            matrix = fitz.Matrix(zoom, zoom)

            # Renderizar página como imagen
            pix = page.get_pixmap(matrix=matrix)
            pix.save(thumbnail_path)
            doc.close()
            return True
        except Exception as e:
            print(f"Error generando thumbnail: {e}")
            return False

    # Función auxiliar para obtener documentos de la carpeta Doc
    def obtener_documentos_fisicoquimicos():
        """Obtiene la lista de documentos PDF de la carpeta Doc con thumbnails"""
        import os
        documentos = []
        doc_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Doc')
        thumbnail_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'thumbnails')

        # Crear carpeta de thumbnails si no existe
        if not os.path.exists(thumbnail_folder):
            os.makedirs(thumbnail_folder)

        if os.path.exists(doc_folder):
            for filename in sorted(os.listdir(doc_folder)):
                if filename.endswith('.pdf'):
                    # Extraer código y nombre del documento
                    # Formato: IT-CS-LFQ-XXX-NOMBRE.pdf
                    parts = filename.replace('.pdf', '').split('-', 4)
                    if len(parts) >= 5:
                        codigo = '-'.join(parts[:4])  # IT-CS-LFQ-XXX
                        nombre = parts[4].replace('_', ' ')  # Nombre del documento
                    else:
                        codigo = filename.replace('.pdf', '')
                        nombre = filename.replace('.pdf', '')

                    # Generar thumbnail si no existe
                    thumbnail_name = filename.replace('.pdf', '.png')
                    thumbnail_path = os.path.join(thumbnail_folder, thumbnail_name)
                    pdf_path = os.path.join(doc_folder, filename)

                    if not os.path.exists(thumbnail_path):
                        generar_thumbnail_pdf(pdf_path, thumbnail_path)

                    # Verificar si el thumbnail existe
                    has_thumbnail = os.path.exists(thumbnail_path)

                    documentos.append({
                        'filename': filename,
                        'codigo': codigo,
                        'nombre': nombre,
                        'thumbnail': thumbnail_name if has_thumbnail else None
                    })

        return documentos

    # Ruta para ver documento PDF
    @app.route('/documentos/fisicoquimicos/ver/<path:filename>')
    @login_required
    def ver_documento_fisicoquimico(filename):
        """Sirve el documento PDF para visualización en el navegador"""
        import os
        from flask import send_from_directory
        doc_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Doc')
        return send_from_directory(doc_folder, filename, mimetype='application/pdf')

    # Ruta para descargar documento PDF
    @app.route('/documentos/fisicoquimicos/descargar/<path:filename>')
    @login_required
    def descargar_documento_fisicoquimico(filename):
        """Sirve el documento PDF para descarga"""
        import os
        from flask import send_from_directory
        doc_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Doc')
        return send_from_directory(doc_folder, filename, as_attachment=True)

    # API para obtener los datos completos de un registro de análisis fisicoquímico
    @app.route('/api/analisis_fisicoquimicos/<int:registro_id>', methods=['GET'])
    @login_required
    def api_analisis_fisicoquimico_get(registro_id):
        # Obtener el registro
        registro = AnalisisCalidad.query.get_or_404(registro_id)
        
        # Crear diccionario con los datos del registro
        data = {
            'id': registro.id,
            'folio': registro.folio,
            'fecha': registro.fecha.strftime('%d/%m/%Y'),
            'turno': registro.turno,
            'producto': registro.producto,
            'horario': registro.horario,
            'lote': registro.lote,
            'analista': registro.analista,
            'peso': registro.peso,
            'humedad_base_frita': registro.humedad_base_frita,
            'aceite_base_frita': registro.aceite_base_frita,
            'tanque1_aceite_pt': registro.tanque1_aceite_pt,
            'tanque1_humedad_pt': registro.tanque1_humedad_pt,
            'tanque1_sal_pt': registro.tanque1_sal_pt,
            'tanque2_aceite_pt': registro.tanque2_aceite_pt,
            'tanque2_humedad_pt': registro.tanque2_humedad_pt,
            'tanque2_sal_pt': registro.tanque2_sal_pt,
            'tanque3_aceite_pt': registro.tanque3_aceite_pt,
            'tanque3_humedad_pt': registro.tanque3_humedad_pt,
            'tanque3_sal_pt': registro.tanque3_sal_pt,
            'observaciones': registro.observaciones,
            'created_by': registro.created_by,
            'created_at': registro.created_at.strftime('%d/%m/%Y %H:%M')
        }
        
        return jsonify(data)

    # API para obtener el último valor de Cloruros Base
    @app.route('/api/ultimo_cloruros_base/<category>', methods=['GET'])
    @login_required
    def api_ultimo_cloruros_base(category):
        """Obtiene el último valor de Cloruros Base registrado para autocompletar"""

        # Verificar que sea PAPA
        if category != 'PAPA':
            return jsonify({'error': 'Esta funcionalidad solo aplica para PAPA'}), 400

        # Obtener producto del query string
        producto = request.args.get('producto')

        try:
            # Construir query base
            query = AnalisisCalidad.query.filter(
                AnalisisCalidad.categoria == category,
                AnalisisCalidad.cloruros_base.isnot(None),
                AnalisisCalidad.cloruros_base != '',
                AnalisisCalidad.cloruros_base != '0'
            )

            # Filtrar por producto si se especificó
            if producto and producto.strip():
                query = query.filter(AnalisisCalidad.producto == producto)
                print(f"🔍 Buscando último Cloruros Base para producto: {producto}")
            else:
                print(f"🔍 Buscando último Cloruros Base (cualquier producto)")

            # Obtener el último registro
            ultimo_registro = query.order_by(AnalisisCalidad.created_at.desc()).first()

            if ultimo_registro:
                print(f"✅ Encontrado: {ultimo_registro.cloruros_base} del {ultimo_registro.fecha}")
                return jsonify({
                    'success': True,
                    'valor': ultimo_registro.cloruros_base,
                    'fecha': ultimo_registro.fecha.strftime('%d/%m/%Y'),
                    'producto': ultimo_registro.producto
                })
            else:
                print(f"❌ No se encontró registro previo")
                return jsonify({
                    'success': False,
                    'mensaje': f'No se encontró ningún registro previo con Cloruros Base{" para " + producto if producto else ""}'
                })

        except Exception as e:
            print(f"Error obteniendo último Cloruros Base: {str(e)}")
            return jsonify({'error': str(e)}), 500

    # API para obtener datos de análisis fisicoquímicos para los gráficos (OPTIMIZADA)
    @app.route('/api/analisis_fisicoquimicos/<category>', methods=['GET'])
    @login_required
    def api_analisis_fisicoquimicos(category):
        """API optimizada para análisis fisicoquímicos con soporte completo para filtros de fecha"""
        
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            return jsonify({'error': 'Categoría no válida'}), 400
        
        try:
            # Obtener parámetros de filtrado
            periodo = request.args.get('periodo', 'mes')
            producto = request.args.get('producto', 'todos')
            
            # NUEVO: Obtener fechas personalizadas
            fecha_inicio_str = request.args.get('fecha_inicio')
            fecha_fin_str = request.args.get('fecha_fin')
            
            # Obtener fecha actual
            from datetime import timedelta
            today = datetime.now().date()
            
            # Lógica mejorada para determinar fechas según el periodo
            if periodo == 'personalizado' and fecha_inicio_str and fecha_fin_str:
                # Usar fechas personalizadas enviadas desde el frontend
                try:
                    fecha_limite = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                    fecha_hasta = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
                    
                    # Validar que fecha_inicio <= fecha_fin
                    if fecha_limite > fecha_hasta:
                        return jsonify({
                            'success': False,
                            'error': 'La fecha de inicio no puede ser mayor que la fecha de fin',
                            'datos': [],
                            'resumen': {'total_registros': 0, 'ultimo_registro': None, 'productos': []}
                        }), 400
                        
                except ValueError:
                    return jsonify({
                        'success': False,
                        'error': 'Formato de fecha inválido. Use YYYY-MM-DD',
                        'datos': [],
                        'resumen': {'total_registros': 0, 'ultimo_registro': None, 'productos': []}
                    }), 400
            else:
                # Usar lógica de periodos predefinidos
                if periodo == 'semana':
                    fecha_limite = today - timedelta(days=7)
                    fecha_hasta = today
                elif periodo == 'mes':
                    fecha_limite = today - timedelta(days=30)
                    fecha_hasta = today
                elif periodo == 'trimestre':
                    fecha_limite = today - timedelta(days=90)
                    fecha_hasta = today
                elif periodo == 'todo':
                    fecha_limite = None
                    fecha_hasta = None
                else:
                    # Fallback para período desconocido
                    fecha_limite = today - timedelta(days=30)
                    fecha_hasta = today
            
            # Construir la consulta base con optimización
            query = AnalisisCalidad.query.filter(AnalisisCalidad.categoria == category)
            
            # Aplicar filtros de fecha si están definidos
            if fecha_limite and fecha_hasta:
                query = query.filter(
                    AnalisisCalidad.fecha >= fecha_limite,
                    AnalisisCalidad.fecha <= fecha_hasta
                )
            elif fecha_limite:  # Solo fecha límite inferior (para 'todo')
                query = query.filter(AnalisisCalidad.fecha >= fecha_limite)
            
            # Filtrar por producto si es necesario
            if producto != 'todos' and producto:
                query = query.filter(AnalisisCalidad.producto == producto)
            
            # Ordenar por fecha y hora para mejor visualización
            query = query.order_by(AnalisisCalidad.fecha.asc(), AnalisisCalidad.created_at.asc())
            
            # Ejecutar la consulta
            registros = query.all()
            
            # Función auxiliar para convertir valores a float válido
            def to_float_or_none(value):
                if value is None or value == '' or value == 'None':
                    return None
                try:
                    return float(value)
                except (ValueError, TypeError):
                    return None
            
            # Construir respuesta con datos validados
            datos = []
            for registro in registros:
                # Procesar valores numéricos de forma segura
                dato = {
                    'id': registro.id,
                    'folio': registro.folio or '',
                    'fecha': registro.fecha.strftime('%d/%m/%Y'),
                    'fecha_iso': registro.fecha.strftime('%Y-%m-%d'),
                    'hora': registro.horario or '00:00',
                    'fecha_hora_display': f"{registro.fecha.strftime('%d/%m')} {registro.horario or '00:00'}",
                    'turno': registro.turno or '',
                    'producto': registro.producto or '',
                    'horario': registro.horario or '',
                    # Valores base frita validados
                    'humedad_base_frita': to_float_or_none(registro.humedad_base_frita),
                    'aceite_base_frita': to_float_or_none(registro.aceite_base_frita),
                    # Tambor 1 - validados
                    'tanque1_aceite_pt': to_float_or_none(registro.tanque1_aceite_pt),
                    'tanque1_humedad_pt': to_float_or_none(registro.tanque1_humedad_pt),
                    'tanque1_sal_pt': to_float_or_none(registro.tanque1_sal_pt),
                    # Tambor 2 - validados
                    'tanque2_aceite_pt': to_float_or_none(registro.tanque2_aceite_pt),
                    'tanque2_humedad_pt': to_float_or_none(registro.tanque2_humedad_pt),
                    'tanque2_sal_pt': to_float_or_none(registro.tanque2_sal_pt),
                    # Tambor 3 - validados (puede ser None para EXTRUIDOS)
                    'tanque3_aceite_pt': to_float_or_none(registro.tanque3_aceite_pt),
                    'tanque3_humedad_pt': to_float_or_none(registro.tanque3_humedad_pt),
                    'tanque3_sal_pt': to_float_or_none(registro.tanque3_sal_pt),
                    'observaciones': registro.observaciones or ''
                }
                datos.append(dato)
            
            # Construir resumen mejorado
            productos_unicos = list(set([d['producto'] for d in datos if d['producto'] and d['producto'].strip()]))
            
            # Estadísticas adicionales
            total_registros = len(datos)
            ultimo_registro = None
            if datos:
                ultimo_registro = datos[-1]['fecha']
            
            # Conteo por producto
            conteo_productos = {}
            for producto_actual in productos_unicos:
                conteo_productos[producto_actual] = len([d for d in datos if d['producto'] == producto_actual])
            
            resumen = {
                'total_registros': total_registros,
                'ultimo_registro': ultimo_registro,
                'productos': productos_unicos,
                'conteo_productos': conteo_productos,
                'rango_fechas': {
                    'desde': datos[0]['fecha'] if datos else None,
                    'hasta': datos[-1]['fecha'] if datos else None
                }
            }
            
            return jsonify({
                'success': True,
                'datos': datos,
                'resumen': resumen,
                'filtros': {
                    'periodo': periodo,
                    'producto': producto,
                    'categoria': category,
                    'fecha_inicio': fecha_inicio_str,
                    'fecha_fin': fecha_fin_str
                }
            })
            
        except Exception as e:
            # Log del error para debugging
            import traceback
            print(f"Error en API análisis fisicoquímicos: {str(e)}")
            print(f"Traceback: {traceback.format_exc()}")
            
            return jsonify({
                'success': False,
                'error': f'Error interno del servidor: {str(e)}',
                'datos': [],
                'resumen': {
                    'total_registros': 0,
                    'ultimo_registro': None,
                    'productos': []
                }
            }), 500
    
    # Rutas para el Análisis de Calidad
    @app.route('/analisis_calidad/<category>', methods=['GET'])
    @login_required
    def list_analisis_calidad(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener todos los registros de análisis para esta categoría
        analisis_records = AnalisisCalidad.query.filter_by(categoria=category).order_by(AnalisisCalidad.created_at.desc()).all()
        
        return render_template('pnc/list_analisis_calidad.html',
                               title=f'Análisis de Calidad - {category}',
                               category=category,
                               analisis_records=analisis_records)
    
    @app.route('/analisis_calidad/<category>/edit/<int:analisis_id>', methods=['GET', 'POST'])
    @login_required
    def edit_analisis_calidad(category, analisis_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro
        analisis = AnalisisCalidad.query.get_or_404(analisis_id)
        
        # Verificar que el análisis pertenece a la categoría correcta
        if analisis.categoria != category:
            flash('Registro de análisis no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_analisis_calidad', category=category))
        
        # Verificar si es el creador o un administrador
        if analisis.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para editar este registro', 'danger')
            return redirect(url_for('list_analisis_calidad', category=category))
        
        form = AnalisisCalidadForm(obj=analisis)
        
        if form.validate_on_submit():
            # Actualizar el registro
            form.populate_obj(analisis)
            db.session.commit()
            
            flash('Registro de análisis actualizado correctamente', 'success')
            return redirect(url_for('list_analisis_calidad', category=category))
        
        return render_template('pnc/edit_analisis_calidad.html',
                               title=f'Editar Análisis de Calidad - {analisis.folio}',
                               form=form,
                               analisis=analisis,
                               category=category)
    
    # Ruta para eliminar análisis fisicoquímicos
    @app.route('/analisis_fisicoquimicos/<category>/delete/<int:analisis_id>', methods=['POST'])
    @login_required
    def delete_analisis_fisicoquimico(category, analisis_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro
        analisis = AnalisisCalidad.query.get_or_404(analisis_id)
        
        # Verificar que el análisis pertenece a la categoría correcta
        if analisis.categoria != category:
            flash('Registro de análisis no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_analisis_fisicoquimicos', category=category))
        
        # Verificar si es el creador o un administrador
        if analisis.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para eliminar este registro', 'danger')
            return redirect(url_for('list_analisis_fisicoquimicos', category=category))
        
        # Guardar información del folio antes de eliminar
        folio = analisis.folio
        
        try:
            db.session.delete(analisis)
            db.session.commit()
            flash(f'Registro de análisis fisicoquímico {folio} eliminado correctamente', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al eliminar el registro: {str(e)}', 'danger')
        
        return redirect(url_for('list_analisis_fisicoquimicos', category=category))
    
    # Rutas para PAE (Evaluación de Producto por Atributos)
    @app.route('/pae/<category>', methods=['GET'])
    @login_required
    def pae_dashboard(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener fecha actual
        today = datetime.now().date()
        
        # Determinar el turno actual basado en la hora
        # Turno A: 7:00 AM a 6:00 PM (horas 7-18)
        # Turno B: 7:00 PM a 6:00 AM (horas 19-23, 0-6)
        now = datetime.now()
        today = now.date()
        current_hour = now.hour

        if 7 <= current_hour <= 18:
            current_turno = 'A'
            horas_turno = list(range(7, 19))  # [7,8,9,10,11,12,13,14,15,16,17,18]
            # Turno A: todos los registros son del día de hoy
            pae_registros = PAERegistro.query.filter(
                PAERegistro.categoria == category,
                PAERegistro.fecha == today,
                PAERegistro.turno == 'A'
            ).all()
        else:
            current_turno = 'B'
            horas_turno = list(range(19, 24)) + list(range(0, 7))  # [19,20,21,22,23,0,1,2,3,4,5,6]

            if current_hour >= 19:
                # Estamos en la primera parte del turno B (19:00 - 23:59)
                # Solo buscar registros de HOY con hora >= 19
                # (los registros de hora 0-6 de hoy pertenecen al turno B ANTERIOR)
                pae_registros = PAERegistro.query.filter(
                    PAERegistro.categoria == category,
                    PAERegistro.fecha == today,
                    PAERegistro.turno == 'B',
                    PAERegistro.hora_bloque >= 19
                ).all()
            else:  # 0 <= current_hour <= 6
                # Estamos en la segunda parte del turno B (00:00 - 06:59)
                # Buscar registros de AYER (19-23) y de HOY (0-6)
                from sqlalchemy import or_, and_
                ayer = today - timedelta(days=1)
                pae_registros = PAERegistro.query.filter(
                    PAERegistro.categoria == category,
                    PAERegistro.turno == 'B',
                    or_(
                        and_(PAERegistro.fecha == ayer, PAERegistro.hora_bloque >= 19),
                        and_(PAERegistro.fecha == today, PAERegistro.hora_bloque < 7)
                    )
                ).all()

        # Crear un diccionario para saber qué horas tienen registros
        horas_con_registros = {}
        for registro in pae_registros:
            horas_con_registros[registro.hora_bloque] = registro.id
        
        return render_template('pae/dashboard.html',
                              title=f'PAE - {category}',
                              category=category,
                              current_turno=current_turno,
                              horas_turno=horas_turno,
                              horas_con_registros=horas_con_registros,
                              current_hour=current_hour)

    @app.route('/api/pae/resumen_por_hora')
    def resumen_por_hora():
        categoria = request.args.get('categoria')
        periodo = request.args.get('periodo', 'turno')
        turno = request.args.get('turno', None)
        atributo = request.args.get('atributo', None)
        producto = request.args.get('producto', None)
        # Por ahora ignoramos 'linea' para compatibilidad

        now = datetime.now()
        current_date = now.date()

        # Determinar rango de fechas según el filtro de periodo
        if periodo == 'hoy':
            fecha_inicio = current_date
            fecha_fin = current_date
        elif periodo == 'ayer':
            fecha_inicio = current_date - timedelta(days=1)
            fecha_fin = current_date - timedelta(days=1)
        elif periodo == 'semana':
            fecha_inicio = current_date - timedelta(days=6)
            fecha_fin = current_date
        else:  # 'turno' (por defecto)
            fecha_inicio = current_date
            fecha_fin = current_date

        # Determinar turno y horas
        if turno is None or turno == 'all':
            horas = list(range(7, 19)) + list(range(19, 24)) + list(range(0, 7))
            turnos = ['A', 'B']
        elif turno == 'A':
            horas = list(range(7, 19))
            turnos = ['A']
        elif turno == 'B':
            horas = list(range(19, 24)) + list(range(0, 7))
            turnos = ['B']
        else:
            horas = list(range(7, 19)) + list(range(19, 24)) + list(range(0, 7))
            turnos = ['A', 'B']

        # Consulta por hora y periodo, ahora filtrando por producto si se especifica
        resumen = []
        for h in horas:
            query = PAERegistro.query.filter(
                PAERegistro.categoria == categoria,
                PAERegistro.fecha >= fecha_inicio,
                PAERegistro.fecha <= fecha_fin,
                PAERegistro.turno.in_(turnos),
                PAERegistro.hora_bloque == h
            )
            if producto and producto != 'all':
                query = query.filter(PAERegistro.producto == producto)
            count = query.count()
            resumen.append({'hora': h, 'cantidad': count})

        return jsonify({'turno': turno, 'periodo': periodo, 'resumen': resumen})

    @app.route('/pae/<category>/registro/<int:hora>', methods=['GET', 'POST'])
    @login_required
    def pae_registro(category, hora):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Verificar que la hora sea válida (0-23)
        if hora < 0 or hora > 23:
            flash('Hora no válida', 'danger')
            return redirect(url_for('pae_dashboard', category=category))
        
        # Determinar el turno basado en la hora
        # Turno A: 7:00 AM a 6:00 PM (horas 7-18)
        # Turno B: 7:00 PM a 6:00 AM (horas 19-23, 0-6)
        if 7 <= hora <= 18:
            turno = 'A'
        else:
            turno = 'B'

        # Fecha y hora actual
        now = datetime.now()
        current_hour = now.hour
        today = now.date()

        # Para BUSCAR registros existentes en turno B, necesitamos determinar la fecha correcta
        # según la hora del registro y la hora actual
        fecha_busqueda = today
        if turno == 'B':
            if hora >= 19:
                # El registro es de la parte 19:00-23:59 del turno B
                if current_hour < 7:
                    # Estamos después de medianoche, esa hora fue AYER
                    fecha_busqueda = today - timedelta(days=1)
                # Si current_hour >= 19, fecha_busqueda = today (correcto)
            else:
                # El registro es de la parte 00:00-06:59 del turno B
                # La fecha es siempre "hoy" cuando consultamos horas 0-6
                fecha_busqueda = today

        # Obtener línea de la URL si existe
        linea = request.args.get('linea', 'all')
        
        # MODIFICADO: Si viene un POST, usar la línea seleccionada en el formulario en lugar de la URL
        form_linea = None
        if request.method == 'POST':
            form_linea = request.form.get('linea')
        
        # Verificar si ya existe un registro para esta hora, fecha, categoría y línea
        # Usar fecha_busqueda para buscar (considera el cruce de medianoche en turno B)
        existing_record_query = PAERegistro.query.filter_by(
            categoria=category,
            fecha=fecha_busqueda,
            turno=turno,
            hora_bloque=hora
        )

        # Si estamos creando/editando y se especificó una línea en el form, buscar por esa línea
        if form_linea:
            existing_record_query = existing_record_query.filter_by(linea=form_linea)
        # Si no hay línea en el form pero hay una línea en la URL, usar esa
        elif linea != 'all':
            existing_record_query = existing_record_query.filter_by(linea=linea)

        existing_record = existing_record_query.first()

        # Si no se encontró un registro específico por línea, intentar buscar cualquier registro para esta hora
        if not existing_record:
            existing_record = PAERegistro.query.filter_by(
                categoria=category,
                fecha=fecha_busqueda,
                turno=turno,
                hora_bloque=hora
            ).first()
        
        if existing_record:
            # Si existe, mostrar formulario para editar
            form = PAEForm(obj=existing_record)
            is_new = False
            atributos_json_str = existing_record.data or "{}"
            hora_registro_val = existing_record.hora.strftime('%H:%M') if existing_record.hora else datetime.now().strftime('%H:%M')
        else:
            # Si no existe, mostrar formulario para crear
            form = PAEForm()
            is_new = True
            atributos_json_str = "{}"
            hora_registro_val = datetime.now().strftime('%H:%M')

        # Procesar el formulario cuando se envía
        if request.method == 'POST':
            try:
                # Obtener datos del formulario directamente desde request.form
                producto = request.form.get('producto', '').strip()
                hora_muestreo_str = request.form.get('hora_muestreo', '')
                data_json = request.form.get('data', '{}')
                
                # Campos sensoriales
                sensorial_apariencia = request.form.get('sensorial_apariencia', '')
                sensorial_apariencia_comentario = request.form.get('sensorial_apariencia_comentario', '')
                sensorial_textura = request.form.get('sensorial_textura', '')
                sensorial_textura_comentario = request.form.get('sensorial_textura_comentario', '')
                sensorial_sabor = request.form.get('sensorial_sabor', '')
                sensorial_sabor_comentario = request.form.get('sensorial_sabor_comentario', '')
                observaciones = request.form.get('observaciones', '')
                
                # Campos de Rotura (solo para PAPA)
                rotura_aplica = request.form.get('rotura_aplica') == 'y'
                hojuela_entera = request.form.get('hojuela_entera')
                hojuela_entera_fiesta = request.form.get('hojuela_entera_fiesta')
                peladeras_scrap = request.form.get('peladeras_scrap')
                rotura_observaciones = request.form.get('rotura_observaciones', '')

                # Campos de Registro cada 4 Horas (solo para EXTRUIDOS)
                # Extraer del JSON para guardar también en columnas
                registro_4horas_aplica = False
                extrusor_humedad_cereal = None
                freidor_tiempo_residencia = None
                freidor_temperatura = None
                sazonado_temp_slurry = None

                if category == 'EXTRUIDOS':
                    try:
                        data_dict = json.loads(data_json)
                        ext_humedad = data_dict.get('EXT-HUMEDAD')
                        ext_tiempo = data_dict.get('EXT-TIEMPO')
                        ext_temp = data_dict.get('EXT-TEMP')
                        ext_slurry = data_dict.get('EXT-SLURRY')

                        # Si hay algún dato de 4 horas, marcar como aplicable
                        if any([ext_humedad, ext_tiempo, ext_temp, ext_slurry]):
                            registro_4horas_aplica = True
                            extrusor_humedad_cereal = float(ext_humedad) if ext_humedad else None
                            freidor_tiempo_residencia = float(ext_tiempo) if ext_tiempo else None
                            freidor_temperatura = float(ext_temp) if ext_temp else None
                            sazonado_temp_slurry = float(ext_slurry) if ext_slurry else None
                    except (json.JSONDecodeError, ValueError, TypeError) as e:
                        # Si hay error, continuar sin campos de 4 horas
                        print(f"Advertencia: No se pudieron extraer campos 4 horas: {e}")
                        pass

                # Campos de Registro cada 4 Horas (solo para TORTILLA)
                registro_4horas_tortilla_aplica = False
                tortilla_tiempo_reposo = None
                tortilla_temp_masa = None
                tortilla_humedad_masa = None
                tortilla_peso_10_base = None
                tortilla_temp_freidor = None

                if category == 'TORTILLA':
                    try:
                        # Obtener campos del formulario directamente
                        t_tiempo_reposo = request.form.get('tortilla_tiempo_reposo')
                        t_temp_masa = request.form.get('tortilla_temp_masa')
                        t_humedad_masa = request.form.get('tortilla_humedad_masa')
                        t_peso_10_base = request.form.get('tortilla_peso_10_base')
                        t_temp_freidor = request.form.get('tortilla_temp_freidor')

                        # Si hay algún dato de 4 horas, marcar como aplicable
                        if any([t_tiempo_reposo, t_temp_masa, t_humedad_masa, t_peso_10_base, t_temp_freidor]):
                            registro_4horas_tortilla_aplica = True
                            tortilla_tiempo_reposo = float(t_tiempo_reposo) if t_tiempo_reposo else None
                            tortilla_temp_masa = float(t_temp_masa) if t_temp_masa else None
                            tortilla_humedad_masa = float(t_humedad_masa) if t_humedad_masa else None
                            tortilla_peso_10_base = float(t_peso_10_base) if t_peso_10_base else None
                            tortilla_temp_freidor = float(t_temp_freidor) if t_temp_freidor else None
                    except (ValueError, TypeError) as e:
                        print(f"Advertencia: No se pudieron extraer campos 4 horas TORTILLA: {e}")
                        pass

                # Validar campos obligatorios
                if not producto:
                    flash('Error: El campo Producto es obligatorio', 'danger')
                    return redirect(url_for('pae_registro', category=category, hora=hora))
                
                # Procesar hora de muestreo
                hora_muestreo = None
                if hora_muestreo_str:
                    try:
                        # Convertir string de hora a objeto time
                        hora_muestreo = datetime.strptime(hora_muestreo_str, '%H:%M').time()
                    except ValueError:
                        # Si falla, usar hora actual
                        hora_muestreo = datetime.now().time()
                else:
                    hora_muestreo = datetime.now().time()
                
                # Validar y procesar JSON de datos
                try:
                    # Verificar que sea JSON válido
                    json.loads(data_json)
                except json.JSONDecodeError:
                    flash('Error: Datos de atributos inválidos', 'danger')
                    return redirect(url_for('pae_registro', category=category, hora=hora))
                
                if is_new:
                    # Crear nuevo registro
                    # Usar fecha_busqueda para que el registro tenga la fecha correcta
                    # (considera el cruce de medianoche en turno B)
                    nuevo_registro = PAERegistro(
                        categoria=category,
                        fecha=fecha_busqueda,
                        turno=turno,
                        hora=datetime.now().time(),
                        hora_bloque=hora,
                        hora_muestreo=hora_muestreo,
                        producto=producto,
                        data=data_json,
                        # Campos sensoriales
                        sensorial_apariencia=sensorial_apariencia if sensorial_apariencia else None,
                        sensorial_apariencia_comentario=sensorial_apariencia_comentario,
                        sensorial_textura=sensorial_textura if sensorial_textura else None,
                        sensorial_textura_comentario=sensorial_textura_comentario,
                        sensorial_sabor=sensorial_sabor if sensorial_sabor else None,
                        sensorial_sabor_comentario=sensorial_sabor_comentario,
                        observaciones=observaciones,
                        # Campos de Rotura (solo para PAPA)
                        rotura_aplica=rotura_aplica if category == 'PAPA' else False,
                        hojuela_entera=float(hojuela_entera) if hojuela_entera and category == 'PAPA' else None,
                        hojuela_entera_fiesta=float(hojuela_entera_fiesta) if hojuela_entera_fiesta and category == 'PAPA' else None,
                        peladeras_scrap=float(peladeras_scrap) if peladeras_scrap and category == 'PAPA' else None,
                        rotura_observaciones=rotura_observaciones if category == 'PAPA' else None,
                        # Campos de Registro cada 4 Horas (solo para EXTRUIDOS)
                        registro_4horas_aplica=registro_4horas_aplica if category == 'EXTRUIDOS' else False,
                        extrusor_humedad_cereal=extrusor_humedad_cereal if category == 'EXTRUIDOS' else None,
                        freidor_tiempo_residencia=freidor_tiempo_residencia if category == 'EXTRUIDOS' else None,
                        freidor_temperatura=freidor_temperatura if category == 'EXTRUIDOS' else None,
                        sazonado_temp_slurry=sazonado_temp_slurry if category == 'EXTRUIDOS' else None,
                        # Campos de Registro cada 4 Horas (solo para TORTILLA)
                        registro_4horas_tortilla_aplica=registro_4horas_tortilla_aplica if category == 'TORTILLA' else False,
                        tortilla_tiempo_reposo=tortilla_tiempo_reposo if category == 'TORTILLA' else None,
                        tortilla_temp_masa=tortilla_temp_masa if category == 'TORTILLA' else None,
                        tortilla_humedad_masa=tortilla_humedad_masa if category == 'TORTILLA' else None,
                        tortilla_peso_10_base=tortilla_peso_10_base if category == 'TORTILLA' else None,
                        tortilla_temp_freidor=tortilla_temp_freidor if category == 'TORTILLA' else None,
                        created_by=current_user.id
                    )

                    db.session.add(nuevo_registro)
                    db.session.commit()

                    flash(f'Registro PAE creado correctamente para {category} a las {hora}:00', 'success')

                else:
                    # Actualizar registro existente
                    existing_record.hora_muestreo = hora_muestreo
                    existing_record.producto = producto
                    existing_record.data = data_json
                    # Campos sensoriales
                    existing_record.sensorial_apariencia = sensorial_apariencia if sensorial_apariencia else None
                    existing_record.sensorial_apariencia_comentario = sensorial_apariencia_comentario
                    existing_record.sensorial_textura = sensorial_textura if sensorial_textura else None
                    existing_record.sensorial_textura_comentario = sensorial_textura_comentario
                    existing_record.sensorial_sabor = sensorial_sabor if sensorial_sabor else None
                    existing_record.sensorial_sabor_comentario = sensorial_sabor_comentario
                    existing_record.observaciones = observaciones
                    # Actualizar campos de Rotura (solo para PAPA)
                    if category == 'PAPA':
                        existing_record.rotura_aplica = rotura_aplica
                        existing_record.hojuela_entera = float(hojuela_entera) if hojuela_entera else None
                        existing_record.hojuela_entera_fiesta = float(hojuela_entera_fiesta) if hojuela_entera_fiesta else None
                        existing_record.peladeras_scrap = float(peladeras_scrap) if peladeras_scrap else None
                        existing_record.rotura_observaciones = rotura_observaciones

                    # Actualizar campos de Registro cada 4 Horas (solo para EXTRUIDOS)
                    if category == 'EXTRUIDOS':
                        existing_record.registro_4horas_aplica = registro_4horas_aplica
                        existing_record.extrusor_humedad_cereal = extrusor_humedad_cereal
                        existing_record.freidor_tiempo_residencia = freidor_tiempo_residencia
                        existing_record.freidor_temperatura = freidor_temperatura
                        existing_record.sazonado_temp_slurry = sazonado_temp_slurry

                    # Actualizar campos de Registro cada 4 Horas (solo para TORTILLA)
                    if category == 'TORTILLA':
                        existing_record.registro_4horas_tortilla_aplica = registro_4horas_tortilla_aplica
                        existing_record.tortilla_tiempo_reposo = tortilla_tiempo_reposo
                        existing_record.tortilla_temp_masa = tortilla_temp_masa
                        existing_record.tortilla_humedad_masa = tortilla_humedad_masa
                        existing_record.tortilla_peso_10_base = tortilla_peso_10_base
                        existing_record.tortilla_temp_freidor = tortilla_temp_freidor

                    db.session.commit()
                    
                    flash(f'Registro PAE actualizado correctamente para {category} a las {hora}:00', 'success')
                
                return redirect(url_for('pae_dashboard', category=category))
                
            except Exception as e:
                db.session.rollback()
                flash(f'Error al guardar el registro: {str(e)}', 'danger')
                print(f"Error en pae_registro: {str(e)}")  # Para debugging
                # No redirigir, mostrar el formulario con el error
        
        # Hora formateada para mostrar (formato 12 horas)
        hora_display = f"{hora}:00"
        if hora == 0:
            hora_display = "12:00 AM"
        elif hora < 12:
            hora_display = f"{hora}:00 AM"
        elif hora == 12:
            hora_display = "12:00 PM"
        else:
            hora_display = f"{hora-12}:00 PM"
        
        atributos_json = json.loads(atributos_json_str)

        return render_template('pae/registro.html',
                              title=f'PAE - Registro {hora_display}',
                              category=category,
                              hora=hora,
                              hora_display=hora_display,
                              turno=turno,
                              linea=linea,
                              form=form,
                              is_new=is_new,
                              atributos_json=atributos_json,
                              hora_registro_val=hora_registro_val)
    
    # Ruta para descargar datos PAE en Excel
    @app.route('/pae/descargar-excel', methods=['GET'])
    @login_required
    def descargar_pae_excel():
        """Descarga datos PAE en formato Excel con filtros"""
        try:
            # Importar bibliotecas necesarias al inicio del app
            from io import BytesIO
            from flask import Response

            # Obtener parámetros de filtro
            categoria = request.args.get('categoria', 'EXTRUIDOS')
            fecha_inicio_str = request.args.get('fecha_inicio')
            fecha_fin_str = request.args.get('fecha_fin')
            turno = request.args.get('turno', 'all')
            producto = request.args.get('producto', 'all')
            tipo_descarga = request.args.get('tipo', 'detallado')

            # REDIRECCIÓN PARA PAPA: Usar ruta especializada con TODAS las columnas
            if categoria == 'PAPA':
                print(f"[DEBUG PAPA] Redirigiendo a ruta especializada papa_excel_routes")
                redirect_url = f'/excel-papa?fecha_inicio={fecha_inicio_str}&fecha_fin={fecha_fin_str}&turno={turno}&producto={producto}&incluir_rangos=true'
                print(f"[DEBUG PAPA] URL: {redirect_url}")
                return redirect(redirect_url)

            # Validar categoría
            if categoria not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
                flash('Categoría no válida', 'danger')
                return redirect(url_for('pae_dashboard', category=categoria))

            # Validar fechas
            if not fecha_inicio_str or not fecha_fin_str:
                flash('Fechas requeridas', 'danger')
                return redirect(url_for('pae_dashboard', category=categoria))
            
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Formato de fecha inválido', 'danger')
                return redirect(url_for('pae_dashboard', category=categoria))
            
            # Construir consulta base
            query = PAERegistro.query.filter(
                PAERegistro.categoria == categoria,
                PAERegistro.fecha >= fecha_inicio,
                PAERegistro.fecha <= fecha_fin
            )
            
            # Aplicar filtros adicionales
            if turno != 'all':
                query = query.filter(PAERegistro.turno == turno)
            
            if producto != 'all':
                query = query.filter(PAERegistro.producto == producto)
            
            # Ordenar por fecha y hora
            registros = query.order_by(PAERegistro.fecha, PAERegistro.hora_bloque).all()
            
            # Verificar si hay datos
            if not registros:
                flash('No hay datos para los filtros seleccionados', 'warning')
                return redirect(url_for('pae_dashboard', category=categoria))
            
            # Crear archivo Excel usando openpyxl directamente
            output = BytesIO()
            
            try:
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                if tipo_descarga == 'resumen':
                    # Crear resumen por atributos
                    datos_resumen = crear_resumen_pae_excel(registros, categoria)
                    df_resumen = pd.DataFrame(datos_resumen)
                    
                    # Escribir a Excel
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_resumen.to_excel(writer, sheet_name='Resumen PAE', index=False)
                        
                        # Formatear encabezados
                        workbook = writer.book
                        worksheet = writer.sheets['Resumen PAE']
                        
                        # Aplicar formato a encabezados
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                            
                else:
                    # Crear datos detallados
                    datos_detallados = crear_detallado_pae_excel(registros, categoria)
                    df_detallado = pd.DataFrame(datos_detallados)

                    # Escribir a Excel
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_detallado.to_excel(writer, sheet_name='Datos PAE Detallados', index=False)

                        # Formatear encabezados
                        workbook = writer.book
                        worksheet = writer.sheets['Datos PAE Detallados']

                        # Definir colores
                        VERDE_FILL = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
                        ROJO_FILL = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')

                        # Aplicar formato a encabezados
                        for cell in worksheet[1]:
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

                        # Aplicar colores a campos de 4 horas TORTILLA
                        if categoria == 'TORTILLA':
                            # Obtener índices de columnas
                            headers = [cell.value for cell in worksheet[1]]

                            col_tiempo_reposo = None
                            col_temp_masa = None
                            col_humedad_masa = None
                            col_peso_base = None
                            col_temp_freidor = None
                            col_producto = None

                            for idx, header in enumerate(headers, 1):
                                if header == 'Tiempo de Reposo':
                                    col_tiempo_reposo = idx
                                elif header == 'Temperatura de Masa (°C)':
                                    col_temp_masa = idx
                                elif header == 'Humedad de Masa (%)':
                                    col_humedad_masa = idx
                                elif header == 'Peso 10 Base Frita (g)':
                                    col_peso_base = idx
                                elif header == 'Temp. Freidor Recto (°C)':
                                    col_temp_freidor = idx
                                elif header == 'Producto':
                                    col_producto = idx

                            # Rangos de validación
                            RANGOS_4H_TORTILLA = {
                                'tiempo_reposo': (10, 18),
                                'temp_masa': (32, 38),
                                'humedad_masa': (49.5, 51.5),
                                'temp_freidor': (183.3, 186.7)
                            }

                            # Productos con rango de peso 24-27
                            PRODUCTOS_TOSTITOS = ['TOSTITOS SALSA VERDE', 'TOSTITOS FH']

                            # Aplicar colores a las filas de datos (desde fila 2)
                            for row_idx in range(2, worksheet.max_row + 1):
                                # Obtener producto de la fila
                                producto_cell = worksheet.cell(row=row_idx, column=col_producto) if col_producto else None
                                producto_valor = producto_cell.value if producto_cell else ''

                                # Tiempo de reposo (10-18)
                                if col_tiempo_reposo:
                                    cell = worksheet.cell(row=row_idx, column=col_tiempo_reposo)
                                    if cell.value and cell.value != '':
                                        try:
                                            valor = float(cell.value)
                                            if RANGOS_4H_TORTILLA['tiempo_reposo'][0] <= valor <= RANGOS_4H_TORTILLA['tiempo_reposo'][1]:
                                                cell.fill = VERDE_FILL
                                            else:
                                                cell.fill = ROJO_FILL
                                        except (ValueError, TypeError):
                                            pass

                                # Temperatura de masa (32-38)
                                if col_temp_masa:
                                    cell = worksheet.cell(row=row_idx, column=col_temp_masa)
                                    if cell.value and cell.value != '':
                                        try:
                                            valor = float(cell.value)
                                            if RANGOS_4H_TORTILLA['temp_masa'][0] <= valor <= RANGOS_4H_TORTILLA['temp_masa'][1]:
                                                cell.fill = VERDE_FILL
                                            else:
                                                cell.fill = ROJO_FILL
                                        except (ValueError, TypeError):
                                            pass

                                # Humedad de masa (49.5-51.5)
                                if col_humedad_masa:
                                    cell = worksheet.cell(row=row_idx, column=col_humedad_masa)
                                    if cell.value and cell.value != '':
                                        try:
                                            valor = float(cell.value)
                                            if RANGOS_4H_TORTILLA['humedad_masa'][0] <= valor <= RANGOS_4H_TORTILLA['humedad_masa'][1]:
                                                cell.fill = VERDE_FILL
                                            else:
                                                cell.fill = ROJO_FILL
                                        except (ValueError, TypeError):
                                            pass

                                # Peso 10 base frita (depende del producto)
                                if col_peso_base:
                                    cell = worksheet.cell(row=row_idx, column=col_peso_base)
                                    if cell.value and cell.value != '':
                                        try:
                                            valor = float(cell.value)
                                            # Determinar rango según producto
                                            if producto_valor in PRODUCTOS_TOSTITOS:
                                                min_val, max_val = 24, 27
                                            else:
                                                min_val, max_val = 23.5, 26.5

                                            if min_val <= valor <= max_val:
                                                cell.fill = VERDE_FILL
                                            else:
                                                cell.fill = ROJO_FILL
                                        except (ValueError, TypeError):
                                            pass

                                # Temperatura de freidor recto (183.3-186.7)
                                if col_temp_freidor:
                                    cell = worksheet.cell(row=row_idx, column=col_temp_freidor)
                                    if cell.value and cell.value != '':
                                        try:
                                            valor = float(cell.value)
                                            if RANGOS_4H_TORTILLA['temp_freidor'][0] <= valor <= RANGOS_4H_TORTILLA['temp_freidor'][1]:
                                                cell.fill = VERDE_FILL
                                            else:
                                                cell.fill = ROJO_FILL
                                        except (ValueError, TypeError):
                                            pass

            except ImportError:
                flash('Error: pandas u openpyxl no están instalados. Ejecute install_excel_dependencies.bat', 'danger')
                return redirect(url_for('pae_dashboard', category=categoria))
            
            output.seek(0)
            
            # Configurar respuesta con nombre de archivo seguro
            filename = f'pae_{categoria.lower()}_{fecha_inicio_str}_{fecha_fin_str}.xlsx'
            
            return Response(
                output.getvalue(),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={
                    'Content-Disposition': f'attachment; filename="{filename}"',
                }
            )
            
        except Exception as e:
            print(f"Error en descarga Excel PAE: {str(e)}")
            import traceback
            traceback.print_exc()
            flash(f'Error al generar el archivo Excel: {str(e)}', 'danger')
            return redirect(url_for('pae_dashboard', category=categoria))
    
    def crear_resumen_pae_excel(registros, categoria):
        """Crea lista de datos con resumen de datos PAE"""
        import json
        
        # Definir atributos según categoría
        atributos_map = {}
        if categoria == 'EXTRUIDOS':
            atributos_map = {
                'A': 'Quemado (<0%)',
                'B': 'Roto (<=3.8 cm) (<8%)',
                'C': 'Defectos Totales (<18%)',
                'D': 'Densidad (64-72 g/l)',
                'E': 'Densidad Base Frita (95-110 g/l)',
                'F': 'Diámetro 20 Collects (21-23 cm)',
                'G': 'Cobertura (90-100%)'
            }
        elif categoria == 'TORTILLA':
            atributos_map = {
                'A': 'Puntos Negros',
                'B': 'Quemado',
                'C': 'Manchas',
                'D': 'Doblados',
                'E': 'Pegados',
                'F': 'Aceitoso',
                'G': 'Puntos Tostados',
                'H': 'Forma',
                'I': 'Planos (Burbuja)',
                'J': 'Ampulas (Burbuja)',
                'K': 'Cumulos (Laminado)',
                'L': 'Hoyos (Laminado)',
                'M': 'Tiras de Masa (Laminado)',
                'N': 'Pliegue (Laminado)',
                'O': 'Dobles (Laminado)',
                'P': 'Tamaño (Tamaño)'
            }
        elif categoria == 'PAPA':
            atributos_map = {
                'A': 'Defectos de color',
                'B': 'Daño seco',
                'C': 'Color indeseable',
                'D': 'Defectos internos papa',
                'E': 'Defectos externos papa',
                'F': 'Defectos totales de papa',
                'G': 'Centros suaves + clusters',
                'H': 'Exceso de cáscara',
                'I': 'Hojuelas aceitosas',
                'J': 'Ampulas',
                'K': 'Puntos obscuros',
                'L': 'Defectos totales de proceso',
                'M': 'Hojuelas dobladas'
            }

        # Procesar datos
        resumen_data = []

        for codigo, nombre in atributos_map.items():
            valores = []

            for registro in registros:
                if registro.data:
                    try:
                        data_json = json.loads(registro.data)
                        valor = data_json.get(codigo, 0)
                        if valor:
                            valores.append(int(valor))
                    except (json.JSONDecodeError, ValueError, TypeError):
                        continue

            total = sum(valores)
            cantidad_registros = len(valores)
            promedio = total / cantidad_registros if cantidad_registros > 0 else 0
            maximo = max(valores) if valores else 0
            minimo = min(valores) if valores else 0

            resumen_data.append({
                'Atributo': nombre,
                'Codigo': codigo,
                'Total': total,
                'Cantidad Registros': cantidad_registros,
                'Promedio': round(promedio, 2),
                'Maximo': maximo,
                'Minimo': minimo
            })

        return resumen_data

    def crear_detallado_pae_excel(registros, categoria):
        """Crea lista de datos con datos detallados PAE"""
        import json

        # Definir atributos según categoría
        atributos_map = {}
        if categoria == 'EXTRUIDOS':
            atributos_map = {
                'A': 'Quemado',
                'B': 'Roto',
                'C': 'Defectos Totales',
                'D': 'Densidad',
                'E': 'Densidad Base Frita',
                'F': 'Diametro 20 Collects',
                'G': 'Cobertura',
                # Campos de Registro cada 4 Horas
                'EXT-HUMEDAD': 'Humedad Cereal Trompo (%)',
                'EXT-TIEMPO': 'Tiempo Residencia Freidor (seg)',
                'EXT-TEMP': 'Temperatura Freidor (°C)',
                'EXT-SLURRY': 'Temperatura Slurry Marmitas (°C)'
            }
        elif categoria == 'TORTILLA':
            atributos_map = {
                'A': 'Puntos Negros',
                'B': 'Quemado',
                'C': 'Manchas',
                'D': 'Doblados',
                'E': 'Pegados',
                'F': 'Aceitoso',
                'G': 'Puntos Tostados',
                'H': 'Forma',
                'I': 'Planos',
                'J': 'Ampulas',
                'K': 'Cumulos',
                'L': 'Hoyos',
                'M': 'Tiras de Masa',
                'N': 'Pliegue',
                'O': 'Dobles',
                'P': 'Tamaño',
                'Q': 'Defectos Totales'
            }
        elif categoria == 'PAPA':
            atributos_map = {
                'A': 'Defectos de color',
                'B': 'Daño seco',
                'C': 'Color indeseable',
                'D': 'Defectos internos papa',
                'E': 'Defectos externos papa',
                'F': 'Defectos totales de papa',
                'G': 'Centros suaves + clusters',
                'H': 'Exceso de cáscara',
                'I': 'Hojuelas aceitosas',
                'J': 'Ampulas',
                'K': 'Puntos obscuros',
                'L': 'Defectos totales de proceso',
                'M': 'Hojuelas dobladas'
            }

        # Procesar registros
        data_rows = []
        
        for registro in registros:
            row = {
                'ID': registro.id,
                'Fecha': registro.fecha.strftime('%Y-%m-%d'),
                'Turno': registro.turno,
                'Hora Bloque': f"{registro.hora_bloque}:00",
                'Hora Registro': registro.hora.strftime('%H:%M:%S') if registro.hora else '',
                'Hora Muestreo': registro.hora_muestreo.strftime('%H:%M:%S') if registro.hora_muestreo else '',
                'Producto': registro.producto or '',
                'Sensorial Apariencia': registro.sensorial_apariencia or '',
                'Sensorial Textura': registro.sensorial_textura or '',
                'Sensorial Sabor': registro.sensorial_sabor or '',
                'Observaciones': registro.observaciones or '',
                'Creado por': registro.creator.username if registro.creator else '',
                'Fecha Creacion': registro.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Agregar datos de atributos
            if registro.data:
                try:
                    data_json = json.loads(registro.data)
                    for codigo, nombre in atributos_map.items():
                        row[f'{nombre} ({codigo})'] = data_json.get(codigo, 0)
                except json.JSONDecodeError:
                    for codigo, nombre in atributos_map.items():
                        row[f'{nombre} ({codigo})'] = 0
            else:
                for codigo, nombre in atributos_map.items():
                    row[f'{nombre} ({codigo})'] = 0
            

            # Agregar columnas específicas de EXTRUIDOS (Registro cada 4 Horas)
            if categoria == 'EXTRUIDOS':
                row['Registro 4H Aplica'] = 'Sí' if (hasattr(registro, 'registro_4horas_aplica') and registro.registro_4horas_aplica) else 'No'
                row['Humedad Cereal Trompo (%)'] = registro.extrusor_humedad_cereal if (hasattr(registro, 'extrusor_humedad_cereal') and registro.extrusor_humedad_cereal) else ''
                row['Tiempo Residencia Freidor (seg)'] = registro.freidor_tiempo_residencia if (hasattr(registro, 'freidor_tiempo_residencia') and registro.freidor_tiempo_residencia) else ''
                row['Temperatura Freidor (°C)'] = registro.freidor_temperatura if (hasattr(registro, 'freidor_temperatura') and registro.freidor_temperatura) else ''
                row['Temperatura Slurry Marmitas (°C)'] = registro.sazonado_temp_slurry if (hasattr(registro, 'sazonado_temp_slurry') and registro.sazonado_temp_slurry) else ''

            # Agregar columnas específicas de TORTILLA (Registro cada 4 Horas)
            if categoria == 'TORTILLA':
                row['Registro 4H Aplica'] = 'Sí' if (hasattr(registro, 'registro_4horas_tortilla_aplica') and registro.registro_4horas_tortilla_aplica) else 'No'
                row['Tiempo de Reposo'] = registro.tortilla_tiempo_reposo if (hasattr(registro, 'tortilla_tiempo_reposo') and registro.tortilla_tiempo_reposo) else ''
                row['Temperatura de Masa (°C)'] = registro.tortilla_temp_masa if (hasattr(registro, 'tortilla_temp_masa') and registro.tortilla_temp_masa) else ''
                row['Humedad de Masa (%)'] = registro.tortilla_humedad_masa if (hasattr(registro, 'tortilla_humedad_masa') and registro.tortilla_humedad_masa) else ''
                row['Peso 10 Base Frita (g)'] = registro.tortilla_peso_10_base if (hasattr(registro, 'tortilla_peso_10_base') and registro.tortilla_peso_10_base) else ''
                row['Temp. Freidor Recto (°C)'] = registro.tortilla_temp_freidor if (hasattr(registro, 'tortilla_temp_freidor') and registro.tortilla_temp_freidor) else ''

            data_rows.append(row)

        return data_rows

    # Rutas para WEAK LINK
    @app.route('/weaklink/<category>', methods=['GET'])
    @login_required
    def weaklink_dashboard(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
            
        # Obtener todos los registros de WeakLink para esta categoría
        weaklink_records = WeakLink.query.filter_by(categoria=category).order_by(WeakLink.created_at.desc()).all()
        
        return render_template('weaklink/dashboard.html',
                              title=f'WEAK LINK - {category}',
                              category=category,
                              records=weaklink_records)
    
    @app.route('/weaklink/<category>/create', methods=['GET', 'POST'])
    @login_required
    def weaklink_create(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
            
        form = WeakLinkForm()
        
        # Configurar opciones dinámicas según la categoría (ANTES de validar)
        if category == 'TORTILLA':
            form.producto.choices = [
                ('DORITOS', 'DORITOS'),
                ('TOSTITOS SALSA VERDE', 'TOSTITOS SALSA VERDE'),
                ('TOSTITOS FH', 'TOSTITOS FH'),
                ('DORITOS PIZZEROLA', 'DORITOS PIZZEROLA'),
                ('DORITOS FH', 'DORITOS FH'),
                ('RANCHERITOS', 'RANCHERITOS'),
                ('DORITOS INCOGNITA', 'DORITOS INCOGNITA')
            ]
        elif category == 'EXTRUIDOS':
            form.producto.choices = [
                ('CHEETOS TORCIDITOS', 'CHEETOS TORCIDITOS'),
                ('CHEETOS XTRA FLAMIN HOT', 'CHEETOS XTRA FLAMIN HOT'),
                ('CHEETOS XTRA FH NUEVO', 'CHEETOS XTRA FH NUEVO'),
                ('OTROS', 'OTROS')
            ]
        elif category == 'PAPA':
            form.producto.choices = [
                ('PAPA SAL', 'PAPA SAL'),
                ('SABRITAS LIMON', 'SABRITAS LIMON'),
                ('RUFFLES SAL', 'RUFFLES SAL'),
                ('RUFFLES QUESO', 'RUFFLES QUESO'),
                ('SABRITAS XTRA FH', 'SABRITAS XTRA FH'),
                ('OTROS', 'OTROS')
            ]
        else:
            form.producto.choices = [('', 'Seleccionar...')]

        now = datetime.now()
        if request.method == 'GET':
            form.fecha.data = now.date()
            form.hora.data = now.time().replace(second=0, microsecond=0)
            # Configurar turno automático
            if 7 <= now.hour <= 18:
                form.turno.data = 'A'
            else:
                form.turno.data = 'B'

        if form.validate_on_submit():
            # Crear nuevo registro de WeakLink
            new_weaklink = WeakLink(
            categoria=category,
            fecha=form.fecha.data,
            hora=form.hora.data,
            turno=form.turno.data,
            operador=form.operador.data,
            orden=form.orden.data,
            maquina=form.maquina.data,
            producto=form.producto.data,
            observaciones=form.observaciones.data,

            # Sección LIL
            limpieza_pesadora=form.limpieza_pesadora.data,
            limpieza_cabezal=form.limpieza_cabezal.data,
            limpieza_mordazas=form.limpieza_mordazas.data,
            condicion_velcro=form.condicion_velcro.data,
            validacion_recetas=form.validacion_recetas.data,
            validacion_etiquetas=form.validacion_etiquetas.data,
            
            # Sección Centerlines
            temperatura_mordaza_frontal=form.temperatura_mordaza_frontal.data,
            temperatura_mordaza_trasera=form.temperatura_mordaza_trasera.data,
            temperatura_sellado_vertical=form.temperatura_sellado_vertical.data,
            bolsa_por_cajas=form.bolsa_por_cajas.data,
            
            # Sección Empaque
            codigo_empaque=form.codigo_empaque.data,
            eficiencia_promocion=form.eficiencia_promocion.data,
            volumen_llenado=form.volumen_llenado.data,
            fecha_frescura=form.fecha_frescura.data,
            acomodo_correcto=form.acomodo_correcto.data,
            apariencia_empaque=form.apariencia_empaque.data,
            hermeticidad=form.hermeticidad.data,
            
            # Sección Peso
            gramaje_impreso=form.gramaje_impreso.data,
            peso_ishida=form.peso_ishida.data,
            peso_muestra_1=form.peso_muestra_1.data,
            peso_muestra_2=form.peso_muestra_2.data,
            peso_muestra_3=form.peso_muestra_3.data,
            peso_muestra_4=form.peso_muestra_4.data,
            peso_muestra_5=form.peso_muestra_5.data,
            peso_muestra_6=form.peso_muestra_6.data,
            peso_muestra_7=form.peso_muestra_7.data,
            peso_muestra_8=form.peso_muestra_8.data,
            peso_muestra_9=form.peso_muestra_9.data,
            peso_muestra_10=form.peso_muestra_10.data,
            peso_promedio=form.peso_promedio.data,
            dif_vs_gramaje=form.dif_vs_gramaje.data,
            dif_vs_ishida=form.dif_vs_ishida.data,
            
            # Sección Cama de Aire
            cama_aire_tipo=form.cama_aire_tipo.data,
            cama_aire_muestra_1=form.cama_aire_muestra_1.data,
            cama_aire_muestra_2=form.cama_aire_muestra_2.data,
            cama_aire_muestra_3=form.cama_aire_muestra_3.data,
            cama_aire_promedio=form.cama_aire_promedio.data,
            
            # Sección Oxígeno Residual
            oxigeno_residual_muestra_1=form.oxigeno_residual_muestra_1.data,
            oxigeno_residual_muestra_2=form.oxigeno_residual_muestra_2.data,
            oxigeno_residual_muestra_3=None,  # Eliminamos este campo del formulario
            oxigeno_residual_promedio=form.oxigeno_residual_promedio.data,
            
                created_by=current_user.id
                )
            
            db.session.add(new_weaklink)
            db.session.commit()
            
            flash('Registro de WEAK LINK creado exitosamente', 'success')
            return redirect(url_for('weaklink_dashboard', category=category))
        
        # Debug: mostrar errores de validación
        if request.method == 'POST' and form.errors:
            print(f"Form data recibida: {request.form}")
            print(f"Producto enviado: '{request.form.get('producto')}'")
            print(f"Opciones válidas: {form.producto.choices}")
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f'Error en {field}: {error}', 'danger')
        
        return render_template('weaklink/create.html',
                              title=f'Nuevo WEAK LINK - {category}',
                              form=form,
                              category=category)
    
    @app.route('/api/weaklink/resultados', methods=['GET'])
    @login_required
    def api_weaklink_resultados():
        from models import WeakLink
        import datetime
        category = request.args.get('category')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        turno = request.args.get('turno')

        query = WeakLink.query
        if category:
            query = query.filter(WeakLink.categoria == category)
        if fecha_inicio:
            try:
                fecha_inicio_dt = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                query = query.filter(WeakLink.fecha >= fecha_inicio_dt)
            except Exception:
                pass
        if fecha_fin:
            try:
                fecha_fin_dt = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                query = query.filter(WeakLink.fecha <= fecha_fin_dt)
            except Exception:
                pass
        if turno:
            query = query.filter(WeakLink.turno == turno)

        resultados = query.order_by(WeakLink.fecha.desc(), WeakLink.hora.desc()).all()
        data = []
        for r in resultados:
            row = {}
            for c in r.__table__.columns:
                val = getattr(r, c.name)
                # Formatear fechas y horas para Excel
                if c.name == 'fecha' and val:
                    row[c.name] = val.strftime('%d/%m/%Y')
                elif c.name == 'hora' and val:
                    row[c.name] = val.strftime('%H:%M')
                elif c.name == 'created_at' and val:
                    row[c.name] = val.strftime('%d/%m/%Y %H:%M')
                elif c.name == 'updated_at' and val:
                    row[c.name] = val.strftime('%d/%m/%Y %H:%M')
                else:
                    row[c.name] = val
            data.append(row)
        return jsonify(data)

    @app.route('/api/weaklink/resumen-turno-actual', methods=['GET'])
    @login_required
    def api_weaklink_resumen_turno_actual():
        """API para obtener resumen por turno A o B"""
        from models import WeakLink
        from datetime import datetime, date, time, timedelta
        from sqlalchemy import func, and_, or_

        categoria = request.args.get('category')
        turno_solicitado = request.args.get('turno', None)  # 'A', 'B' o None para actual

        if not categoria or categoria not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            return jsonify({'error': 'Categoría no válida'}), 400

        now = datetime.now()
        current_date = now.date()
        current_time = now.time()

        # Horarios de turnos: A = 7:00-19:00, B = 19:00-7:00
        turno_a_inicio = time(7, 0)
        turno_a_fin = time(19, 0)

        # Determinar turno actual si no se especifica
        if turno_solicitado is None:
            if turno_a_inicio <= current_time < turno_a_fin:
                turno_solicitado = 'A'
            else:
                turno_solicitado = 'B'

        # Calcular fechas y horas según el turno solicitado
        if turno_solicitado == 'A':
            # Turno A: 7:00 AM - 7:00 PM del mismo día
            fecha_turno = current_date
            hora_inicio = turno_a_inicio
            hora_fin = turno_a_fin
            turno_label = 'A (7:00 - 19:00)'

            # Consultar registros del Turno A
            registros = WeakLink.query.filter(
                WeakLink.categoria == categoria,
                WeakLink.fecha == fecha_turno,
                WeakLink.turno == 'A'
            ).all()
        else:
            # Turno B: 7:00 PM - 7:00 AM del día siguiente
            # Si estamos antes de las 7 AM, el turno B empezó ayer
            if current_time < turno_a_inicio:
                fecha_inicio_b = current_date - timedelta(days=1)
                fecha_fin_b = current_date
            else:
                fecha_inicio_b = current_date
                fecha_fin_b = current_date + timedelta(days=1)

            turno_label = 'B (19:00 - 7:00)'

            # Consultar registros del Turno B
            registros = WeakLink.query.filter(
                WeakLink.categoria == categoria,
                WeakLink.turno == 'B',
                or_(
                    WeakLink.fecha == fecha_inicio_b,
                    WeakLink.fecha == fecha_fin_b
                )
            ).all()

            fecha_turno = fecha_inicio_b

        # Definir equipos según categoría
        if categoria == 'TORTILLA':
            equipos = [f'Tubo {i}' for i in range(1, 17)]  # Tubo 1-16
        elif categoria == 'EXTRUIDOS':
            equipos = [f'Tubo {i}' for i in range(17, 33)]  # Tubo 17-32
        elif categoria == 'PAPA':
            equipos = [f'Tubo {i}' for i in range(33, 51)]  # Tubo 33-50 (18 equipos)
        else:
            equipos = []

        # Contar registros por equipo
        conteo_por_equipo = {}
        for registro in registros:
            equipo = registro.maquina
            if equipo not in conteo_por_equipo:
                conteo_por_equipo[equipo] = 0
            conteo_por_equipo[equipo] += 1

        # Preparar resumen
        resumen = []
        total_registros = 0
        equipos_completados = 0

        for equipo in equipos:
            count = conteo_por_equipo.get(equipo, 0)
            total_registros += count

            # Determinar color según cantidad de registros
            if count == 0:
                color = 'danger'  # Rojo
                progreso = 0
                completado = False
            elif count < 3:
                color = 'warning'  # Amarillo
                progreso = (count / 3) * 100
                completado = False
            else:
                color = 'success'  # Verde
                progreso = 100
                completado = True
                equipos_completados += 1

            resumen.append({
                'equipo': equipo,
                'registros': count,
                'objetivo': 3,
                'progreso': round(progreso, 1),
                'color': color,
                'completado': completado
            })

        # Calcular progreso global
        total_equipos = len(equipos)
        porcentaje_completado = round((equipos_completados / total_equipos * 100), 1) if total_equipos > 0 else 0

        return jsonify({
            'categoria': categoria,
            'turno': turno_solicitado,
            'turno_label': turno_label,
            'fecha_turno': fecha_turno.strftime('%Y-%m-%d'),
            'hora_actual': now.strftime('%H:%M'),
            'total_equipos': total_equipos,
            'equipos_completados': equipos_completados,
            'total_registros': total_registros,
            'porcentaje_completado': porcentaje_completado,
            'resumen': resumen
        })

    @app.route('/weaklink/<category>/delete/<int:record_id>', methods=['POST'])
    @login_required
    def weaklink_delete(category, record_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro WeakLink
        record = WeakLink.query.get_or_404(record_id)
        
        # Verificar que el registro pertenece a la categoría correcta
        if record.categoria != category:
            flash('Registro no pertenece a esta categoría', 'danger')
            return redirect(url_for('weaklink_dashboard', category=category))
        
        # Verificar si es el creador o un administrador
        if record.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para eliminar este registro', 'danger')
            return redirect(url_for('weaklink_dashboard', category=category))
        
        db.session.delete(record)
        db.session.commit()
        
        flash('Registro WEAK LINK eliminado correctamente', 'success')
        return redirect(url_for('weaklink_dashboard', category=category))

    # API para obtener los datos completos de un registro WEAK LINK
    @app.route('/api/weaklink/<int:record_id>', methods=['GET'])
    @login_required
    def api_weaklink_get(record_id):
        # Obtener el registro WeakLink
        record = WeakLink.query.get_or_404(record_id)
        
        # Crear diccionario con los datos del registro
        data = {
            'id': record.id,
            'categoria': record.categoria,
            'fecha': record.fecha.strftime('%d/%m/%Y'),
            'hora': record.hora.strftime('%H:%M'),
            'turno': record.turno,
            'operador': record.operador,
            'linea': record.categoria,
            'orden': record.orden,
            'maquina': record.maquina,
            'producto': record.producto,
            'observaciones': record.observaciones,
            
            # Sección LIL
            'limpieza_pesadora': record.limpieza_pesadora,
            'limpieza_cabezal': record.limpieza_cabezal,
            'limpieza_mordazas': record.limpieza_mordazas,
            'condicion_velcro': record.condicion_velcro,
            'validacion_recetas': record.validacion_recetas,
            'validacion_etiquetas': record.validacion_etiquetas,
            
            # Sección Centerlines
            'temperatura_mordaza_frontal': record.temperatura_mordaza_frontal,
            'temperatura_mordaza_trasera': record.temperatura_mordaza_trasera,
            'temperatura_sellado_vertical': record.temperatura_sellado_vertical,
            'bolsa_por_cajas': record.bolsa_por_cajas,
            
            # Sección Empaque
            'codigo_empaque': record.codigo_empaque,
            'eficiencia_promocion': record.eficiencia_promocion,
            'porcentaje_eficiencia': record.porcentaje_eficiencia,
            'volumen_llenado': record.volumen_llenado,
            'fecha_frescura': record.fecha_frescura,
            'acomodo_correcto': record.acomodo_correcto,
            'apariencia_empaque': record.apariencia_empaque,
            'hermeticidad': record.hermeticidad,
            
            # Sección Peso
            'gramaje_impreso': record.gramaje_impreso,
            'peso_ishida': record.peso_ishida,
            'peso_muestra_1': record.peso_muestra_1,
            'peso_muestra_2': record.peso_muestra_2,
            'peso_muestra_3': record.peso_muestra_3,
            'peso_muestra_4': record.peso_muestra_4,
            'peso_muestra_5': record.peso_muestra_5,
            'peso_muestra_6': record.peso_muestra_6,
            'peso_muestra_7': record.peso_muestra_7,
            'peso_muestra_8': record.peso_muestra_8,
            'peso_muestra_9': record.peso_muestra_9,
            'peso_muestra_10': record.peso_muestra_10,
            'peso_promedio': record.peso_promedio,
            'dif_vs_gramaje': record.dif_vs_gramaje,
            'dif_vs_ishida': record.dif_vs_ishida,
            
            # Sección Cama de Aire
            'cama_aire_tipo': record.cama_aire_tipo,
            'cama_aire_muestra_1': record.cama_aire_muestra_1,
            'cama_aire_muestra_2': record.cama_aire_muestra_2,
            'cama_aire_muestra_3': record.cama_aire_muestra_3,
            'cama_aire_promedio': record.cama_aire_promedio,
            
            # Sección Oxígeno Residual
            'oxigeno_residual_muestra_1': record.oxigeno_residual_muestra_1,
            'oxigeno_residual_muestra_2': record.oxigeno_residual_muestra_2,
            'oxigeno_residual_promedio': record.oxigeno_residual_promedio,
            
            'created_by': record.created_by,
            'created_at': record.created_at.strftime('%d/%m/%Y %H:%M')
        }
        
        return jsonify(data)

    @app.route('/analisis_calidad/<category>/delete/<int:analisis_id>', methods=['POST'])
    @login_required
    def delete_analisis_calidad(category, analisis_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro
        analisis = AnalisisCalidad.query.get_or_404(analisis_id)
        
        # Verificar que el análisis pertenece a la categoría correcta
        if analisis.categoria != category:
            flash('Registro de análisis no pertenece a esta categoría', 'danger')
            return redirect(url_for('list_analisis_calidad', category=category))
        
        # Verificar si es el creador o un administrador
        if analisis.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para eliminar este registro', 'danger')
            return redirect(url_for('list_analisis_calidad', category=category))
        
        db.session.delete(analisis)
        db.session.commit()
        
        flash('Registro de análisis eliminado correctamente', 'success')
        return redirect(url_for('list_analisis_fisicoquimicos', category=category))

    @app.route('/weaklink/<category>/edit/<int:record_id>', methods=['GET', 'POST'])
    @login_required
    def weaklink_edit(category, record_id):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        # Obtener el registro WeakLink
        record = WeakLink.query.get_or_404(record_id)
        
        # Verificar que el registro pertenece a la categoría correcta
        if record.categoria != category:
            flash('Registro no pertenece a esta categoría', 'danger')
            return redirect(url_for('weaklink_dashboard', category=category))
        
        # Verificar si es el creador o un administrador
        if record.created_by != current_user.id and not current_user.is_admin:
            flash('No tienes permiso para editar este registro', 'danger')
            return redirect(url_for('weaklink_dashboard', category=category))
        
        # Crear formulario con los datos del registro
        form = WeakLinkForm(obj=record)
        
        # Configurar opciones dinA­micas segA§n la categorA-a
        if category == 'TORTILLA':
            form.producto.choices = [
                ('DORITOS', 'DORITOS'),
                ('TOSTITOS SALSA VERDE', 'TOSTITOS SALSA VERDE'),
                ('TOSTITOS FH', 'TOSTITOS FH'),
                ('DORITOS PIZZEROLA', 'DORITOS PIZZEROLA'),
                ('DORITOS FH', 'DORITOS FH'),
                ('RANCHERITOS', 'RANCHERITOS'),
                ('DORITOS INCOGNITA', 'DORITOS INCOGNITA')
            ]
        elif category == 'EXTRUIDOS':
            form.producto.choices = [
                ('CHEETOS TORCIDITOS', 'CHEETOS TORCIDITOS'),
                ('CHEETOS XTRA FLAMIN HOT', 'CHEETOS XTRA FLAMIN HOT'),
                ('CHEETOS XTRA FH NUEVO', 'CHEETOS XTRA FH NUEVO'),
                ('OTROS', 'OTROS')
            ]
        elif category == 'PAPA':
            form.producto.choices = [
                ('PAPA SAL', 'PAPA SAL'),
                ('SABRITAS LIMON', 'SABRITAS LIMON'),
                ('RUFFLES SAL', 'RUFFLES SAL'),
                ('RUFFLES QUESO', 'RUFFLES QUESO'),
                ('SABRITAS XTRA FH', 'SABRITAS XTRA FH'),
                ('OTROS', 'OTROS')
            ]
        else:
            form.producto.choices = [('', 'Seleccionar...')]
        
        if form.validate_on_submit():
            # Actualizar datos del registro
            form.populate_obj(record)
            record.oxigeno_residual_muestra_3 = None
            db.session.commit()
            
            flash('Registro WEAK LINK actualizado correctamente', 'success')
            return redirect(url_for('weaklink_dashboard', category=category))
        
        return render_template('weaklink/edit.html',
                              title=f'Editar WEAK LINK - {category}',
                              form=form,
                              record=record,
                              category=category)

    # === APIs PAE para datos reales ===
    
    @app.route('/pae/<category>/datos', methods=['GET'])
    @login_required
    def api_pae_datos(category):
        """API para obtener datos PAE reales para los gráficos"""
        try:
            # Verificar categoría válida
            if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
                return jsonify({'error': 'Categoría no válida'}), 400
            
            # Obtener parámetros de filtro
            periodo = request.args.get('periodo', 'turno')
            turno = request.args.get('turno', 'all')
            linea = request.args.get('linea', 'all')
            producto = request.args.get('producto', 'all')
            fecha_inicio_param = request.args.get('fecha_inicio')
            fecha_fin_param = request.args.get('fecha_fin')
            
            # Calcular rango de fechas
            now = datetime.now()
            today = now.date()
            
            if periodo == 'personalizado' and fecha_inicio_param and fecha_fin_param:
                # Usar fechas proporcionadas por el usuario
                try:
                    from datetime import datetime as dt
                    fecha_inicio = dt.strptime(fecha_inicio_param, '%Y-%m-%d').date()
                    fecha_fin = dt.strptime(fecha_fin_param, '%Y-%m-%d').date()
                except ValueError:
                    fecha_inicio = today
                    fecha_fin = today
            elif periodo == 'hoy':
                fecha_inicio = today
                fecha_fin = today
            elif periodo == 'ayer':
                fecha_inicio = today - timedelta(days=1)
                fecha_fin = today - timedelta(days=1)
            elif periodo == 'semana':
                fecha_inicio = today - timedelta(days=6)
                fecha_fin = today
            else:  # 'turno' por defecto
                fecha_inicio = today
                fecha_fin = today
            
            # Construir consulta base
            query = PAERegistro.query.filter(
                PAERegistro.categoria == category,
                PAERegistro.fecha >= fecha_inicio,
                PAERegistro.fecha <= fecha_fin
            )
            
            # Filtrar por turno si se especifica
            if turno != 'all':
                query = query.filter(PAERegistro.turno == turno)
            
            # Filtrar por producto si se especifica
            if producto != 'all':
                query = query.filter(PAERegistro.producto == producto)
            
            # Ordenar por fecha y hora
            registros = query.order_by(PAERegistro.fecha, PAERegistro.hora_bloque).all()
            
            # Procesar datos según la categoría
            if category == 'EXTRUIDOS':
                return procesar_datos_extruidos(registros, turno)
            elif category == 'TORTILLA':
                return procesar_datos_tortilla(registros, turno)
            elif category == 'PAPA':
                return procesar_datos_papa(registros, turno)
                
        except Exception as e:
            print(f"Error en API PAE datos: {str(e)}")
            return jsonify({'error': f'Error interno: {str(e)}'}), 500
    
    def procesar_datos_extruidos(registros, turno):
        """Procesa datos PAE específicos para EXTRUIDOS"""

        # Definir atributos para EXTRUIDOS
        atributos_extruidos = {
            'A': 'Quemado (<0%)',
            'B': 'Roto (<=3.8 cm) (<8%)',
            'C': 'Defectos Totales (<18%)',
            'D': 'Densidad (64-72 g/l)',
            'E': 'Densidad Base Frita (95-110 g/l)',
            'F': 'Diámetro 20 Collects (21-23 cm)',
            'G': 'Cobertura (90-100%)'
        }

        # Determinar horas según turno
        if turno == 'A':
            horas = list(range(7, 19))
        elif turno == 'B':
            horas = list(range(19, 24)) + list(range(0, 7))
        else:
            horas = list(range(7, 19)) + list(range(19, 24)) + list(range(0, 7))

        # Inicializar estructura de datos
        datos = {}
        for codigo in atributos_extruidos.keys():
            datos[codigo] = [0] * len(horas)

        # Contador de registros con valor por cada atributo (para calcular promedio correcto)
        conteo_por_atributo = {}
        for codigo in atributos_extruidos.keys():
            conteo_por_atributo[codigo] = 0

        # Procesar cada registro
        for registro in registros:
            if registro.data:
                try:
                    data_json = json.loads(registro.data)
                    hora_idx = None

                    # Encontrar índice de la hora
                    try:
                        hora_idx = horas.index(registro.hora_bloque)
                    except ValueError:
                        continue

                    # Extraer valores de atributos
                    for codigo in atributos_extruidos.keys():
                        valor_raw = data_json.get(codigo)
                        # Solo contar si el campo tiene un valor real (no vacío, no None)
                        if valor_raw is not None and valor_raw != '' and valor_raw != 0:
                            try:
                                valor_num = float(valor_raw)
                                datos[codigo][hora_idx] += valor_num
                                conteo_por_atributo[codigo] += 1
                            except (ValueError, TypeError):
                                pass
                        elif valor_raw == 0 or valor_raw == '0':
                            # El valor 0 es válido, debe contarse
                            datos[codigo][hora_idx] += 0
                            conteo_por_atributo[codigo] += 1

                except json.JSONDecodeError:
                    continue

        # Calcular resumen - promedio basado SOLO en formularios que tienen valor en ese campo
        num_registros = len(registros)
        resumen = []
        for codigo, nombre in atributos_extruidos.items():
            valores = datos[codigo]
            total = sum(valores)
            # Promedio = total / número de formularios que tienen valor en este campo específico
            conteo_campo = conteo_por_atributo[codigo]
            promedio = total / conteo_campo if conteo_campo > 0 else 0
            maximo = max(valores) if valores else 0
            hora_max = horas[valores.index(maximo)] if maximo > 0 else None

            resumen.append({
                'codigo': codigo,
                'atributo': nombre,
                'total': round(total, 2),
                'promedio': round(promedio, 2),
                'maximo': round(maximo, 2),
                'hora_max': hora_max,
                'registros_con_valor': conteo_campo
            })

        # Ordenar por total descendente
        resumen.sort(key=lambda x: x['total'], reverse=True)

        return jsonify({
            'horas': horas,
            'datos': datos,
            'resumen': resumen,
            'periodo': request.args.get('periodo', 'turno'),
            'turno': turno,
            'categoria': 'EXTRUIDOS',
            'num_registros': num_registros
        })
    
    def procesar_datos_tortilla(registros, turno):
        """Procesa datos PAE específicos para TORTILLA"""

        # Definir atributos para TORTILLA
        atributos_tortilla = {
            'A': 'Puntos Negros',
            'B': 'Quemado',
            'C': 'Manchas',
            'D': 'Doblados',
            'E': 'Pegados',
            'F': 'Aceitoso',
            'G': 'Puntos Tostados',
            'H': 'Forma',
            'I': 'Planos (Burbuja)',
            'J': 'Ámpulas (Burbuja)',
            'K': 'Cúmulos (Laminado)',
            'L': 'Hoyos (Laminado)',
            'M': 'Tiras de Masa (Laminado)',
            'N': 'Pliegue (Laminado)',
            'O': 'Dobles (Laminado)',
            'P': 'Tamaño (Tamaño)'
        }

        # Determinar horas según turno
        if turno == 'A':
            horas = list(range(7, 19))
        elif turno == 'B':
            horas = list(range(19, 24)) + list(range(0, 7))
        else:
            horas = list(range(7, 19)) + list(range(19, 24)) + list(range(0, 7))

        # Inicializar estructura de datos
        datos = {}
        for codigo in atributos_tortilla.keys():
            datos[codigo] = [0] * len(horas)

        # Contador de registros con valor por cada atributo
        conteo_por_atributo = {}
        for codigo in atributos_tortilla.keys():
            conteo_por_atributo[codigo] = 0

        # Procesar cada registro
        for registro in registros:
            if registro.data:
                try:
                    data_json = json.loads(registro.data)
                    hora_idx = None

                    # Encontrar índice de la hora
                    try:
                        hora_idx = horas.index(registro.hora_bloque)
                    except ValueError:
                        continue

                    # Extraer valores de atributos
                    for codigo in atributos_tortilla.keys():
                        valor_raw = data_json.get(codigo)
                        # Solo contar si el campo tiene un valor real
                        if valor_raw is not None and valor_raw != '' and valor_raw != 0:
                            try:
                                valor_num = float(valor_raw)
                                datos[codigo][hora_idx] += valor_num
                                conteo_por_atributo[codigo] += 1
                            except (ValueError, TypeError):
                                pass
                        elif valor_raw == 0 or valor_raw == '0':
                            datos[codigo][hora_idx] += 0
                            conteo_por_atributo[codigo] += 1

                except json.JSONDecodeError:
                    continue

        # Calcular resumen - promedio basado SOLO en formularios con valor en ese campo
        num_registros = len(registros)
        resumen = []
        for codigo, nombre in atributos_tortilla.items():
            valores = datos[codigo]
            total = sum(valores)
            conteo_campo = conteo_por_atributo[codigo]
            promedio = total / conteo_campo if conteo_campo > 0 else 0
            maximo = max(valores) if valores else 0
            hora_max = horas[valores.index(maximo)] if maximo > 0 else None

            resumen.append({
                'codigo': codigo,
                'atributo': nombre,
                'total': round(total, 2),
                'promedio': round(promedio, 2),
                'maximo': round(maximo, 2),
                'hora_max': hora_max,
                'registros_con_valor': conteo_campo
            })

        # Ordenar por total descendente
        resumen.sort(key=lambda x: x['total'], reverse=True)

        return jsonify({
            'horas': horas,
            'datos': datos,
            'resumen': resumen,
            'periodo': request.args.get('periodo', 'turno'),
            'turno': turno,
            'categoria': 'TORTILLA',
            'num_registros': num_registros
        })
    
    def procesar_datos_papa(registros, turno):
        """Procesa datos PAE específicos para PAPA"""

        # Definir atributos para PAPA (A-M según especificaciones)
        atributos_papa = {
            'A': 'Defectos de color',
            'B': 'Daño seco',
            'C': 'Color indeseable',
            'D': 'Defectos internos papa',
            'E': 'Defectos externos papa',
            'F': 'Defectos totales de papa',
            'G': 'Centros suaves + clusters',
            'H': 'Exceso de cáscara',
            'I': 'Hojuelas aceitosas',
            'J': 'Ampulas',
            'K': 'Puntos obscuros',
            'L': 'Defectos totales de proceso',
            'M': 'Hojuelas dobladas'
        }

        # Determinar horas según turno
        if turno == 'A':
            horas = list(range(7, 19))
        elif turno == 'B':
            horas = list(range(19, 24)) + list(range(0, 7))
        else:
            horas = list(range(7, 19)) + list(range(19, 24)) + list(range(0, 7))

        # Inicializar estructura de datos
        datos = {}
        for codigo in atributos_papa.keys():
            datos[codigo] = [0] * len(horas)

        # Contador de registros con valor por cada atributo
        conteo_por_atributo = {}
        for codigo in atributos_papa.keys():
            conteo_por_atributo[codigo] = 0

        # Procesar cada registro
        num_registros = len(registros)
        for registro in registros:
            if registro.data:
                try:
                    data_json = json.loads(registro.data)
                    hora_idx = None

                    # Encontrar índice de la hora
                    try:
                        hora_idx = horas.index(registro.hora_bloque)
                    except ValueError:
                        continue

                    # Extraer valores de atributos
                    for codigo in atributos_papa.keys():
                        valor_raw = data_json.get(codigo)
                        # Solo contar si el campo tiene un valor real (no vacío, no None)
                        if valor_raw is not None and valor_raw != '' and valor_raw != 0:
                            try:
                                valor_num = float(valor_raw)
                                datos[codigo][hora_idx] += valor_num
                                conteo_por_atributo[codigo] += 1
                            except (ValueError, TypeError):
                                pass
                        elif valor_raw == 0 or valor_raw == '0':
                            # Si el valor es explícitamente 0, también cuenta como llenado
                            datos[codigo][hora_idx] += 0
                            conteo_por_atributo[codigo] += 1

                except json.JSONDecodeError:
                    continue

        # Calcular resumen - promedio basado solo en formularios con valor para cada atributo
        resumen = []
        for codigo, nombre in atributos_papa.items():
            valores = datos[codigo]
            total = sum(valores)
            # Promedio = total / número de formularios que tienen valor en este campo específico
            conteo_campo = conteo_por_atributo[codigo]
            promedio = total / conteo_campo if conteo_campo > 0 else 0
            maximo = max(valores) if valores else 0
            hora_max = horas[valores.index(maximo)] if maximo > 0 else None

            resumen.append({
                'codigo': codigo,
                'atributo': nombre,
                'total': total,
                'promedio': round(promedio, 2),
                'maximo': maximo,
                'hora_max': hora_max,
                'formularios_con_valor': conteo_campo
            })

        # Ordenar por total descendente
        resumen.sort(key=lambda x: x['total'], reverse=True)

        return jsonify({
            'horas': horas,
            'datos': datos,
            'resumen': resumen,
            'periodo': request.args.get('periodo', 'turno'),
            'turno': turno,
            'categoria': 'PAPA',
            'num_registros': num_registros
        })
    
    @app.route('/api/pae/productos', methods=['GET'])
    @login_required
    def api_pae_productos():
        """API para obtener productos únicos de PAE"""
        try:
            categoria = request.args.get('categoria')
            if not categoria or categoria not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
                return jsonify({'error': 'Categoría no válida'}), 400
            
            # Obtener productos únicos de los registros PAE
            productos = db.session.query(PAERegistro.producto).filter(
                PAERegistro.categoria == categoria,
                PAERegistro.producto.isnot(None)
            ).distinct().all()
            
            productos_list = [p[0] for p in productos if p[0]]
            
            return jsonify({
                'productos': productos_list,
                'categoria': categoria
            })
            
        except Exception as e:
            return jsonify({'error': f'Error interno: {str(e)}'}), 500
    
    def crear_resumen_pae_excel(registros, categoria):
        """Crea lista de datos con resumen de datos PAE"""
        import json
        
        # Definir atributos según categoría
        atributos_map = {}
        if categoria == 'EXTRUIDOS':
            atributos_map = {
                'A': 'Quemado (<0%)',
                'B': 'Roto (<=3.8 cm) (<8%)',
                'C': 'Defectos Totales (<18%)',
                'D': 'Densidad (64-72 g/l)',
                'E': 'Densidad Base Frita (95-110 g/l)',
                'F': 'Diámetro 20 Collects (21-23 cm)',
                'G': 'Cobertura (100%)'
            }
        elif categoria == 'TORTILLA':
            atributos_map = {
                'A': 'Puntos Negros',
                'B': 'Quemado',
                'C': 'Manchas',
                'D': 'Doblados',
                'E': 'Pegados',
                'F': 'Aceitoso',
                'G': 'Puntos Tostados',
                'H': 'Forma',
                'I': 'Planos (Burbuja)',
                'J': 'Ampulas (Burbuja)',
                'K': 'Cumulos (Laminado)',
                'L': 'Hoyos (Laminado)',
                'M': 'Tiras de Masa (Laminado)',
                'N': 'Pliegue (Laminado)',
                'O': 'Dobles (Laminado)',
                'P': 'Tamaño (Tamaño)'
            }
        elif categoria == 'PAPA':
            atributos_map = {
                'A': 'Defectos de color',
                'B': 'Daño seco',
                'C': 'Color indeseable',
                'D': 'Defectos internos papa',
                'E': 'Defectos externos papa',
                'F': 'Defectos totales de papa',
                'G': 'Centros suaves + clusters',
                'H': 'Exceso de cáscara',
                'I': 'Hojuelas aceitosas',
                'J': 'Ampulas',
                'K': 'Puntos obscuros',
                'L': 'Defectos totales de proceso',
                'M': 'Hojuelas dobladas'
            }

        # Procesar datos
        resumen_data = []

        for codigo, nombre in atributos_map.items():
            valores = []

            for registro in registros:
                if registro.data:
                    try:
                        data_json = json.loads(registro.data)
                        valor = data_json.get(codigo, 0)
                        if valor:
                            valores.append(int(valor))
                    except (json.JSONDecodeError, ValueError, TypeError):
                        continue

            total = sum(valores)
            cantidad_registros = len(valores)
            promedio = total / cantidad_registros if cantidad_registros > 0 else 0
            maximo = max(valores) if valores else 0
            minimo = min(valores) if valores else 0

            resumen_data.append({
                'Atributo': nombre,
                'Codigo': codigo,
                'Total': total,
                'Cantidad Registros': cantidad_registros,
                'Promedio': round(promedio, 2),
                'Maximo': maximo,
                'Minimo': minimo
            })

        return resumen_data

    def crear_detallado_pae_excel(registros, categoria):
        """Crea lista de datos con datos detallados PAE"""
        import json

        # Definir atributos según categoría
        atributos_map = {}
        if categoria == 'EXTRUIDOS':
            atributos_map = {
                'A': 'Quemado',
                'B': 'Roto',
                'C': 'Defectos Totales',
                'D': 'Densidad',
                'E': 'Densidad Base Frita',
                'F': 'Diametro 20 Collects',
                'G': 'Cobertura'
            }
        elif categoria == 'TORTILLA':
            atributos_map = {
                'A': 'Puntos Negros',
                'B': 'Quemado',
                'C': 'Manchas',
                'D': 'Doblados',
                'E': 'Pegados',
                'F': 'Aceitoso',
                'G': 'Puntos Tostados',
                'H': 'Forma',
                'I': 'Planos',
                'J': 'Ampulas',
                'K': 'Cumulos',
                'L': 'Hoyos',
                'M': 'Tiras de Masa',
                'N': 'Pliegue',
                'O': 'Dobles',
                'P': 'Tamaño',
                'Q': 'Defectos Totales'
            }
        elif categoria == 'PAPA':
            atributos_map = {
                'A': 'Defectos de color',
                'B': 'Daño seco',
                'C': 'Color indeseable',
                'D': 'Defectos internos papa',
                'E': 'Defectos externos papa',
                'F': 'Defectos totales de papa',
                'G': 'Centros suaves + clusters',
                'H': 'Exceso de cáscara',
                'I': 'Hojuelas aceitosas',
                'J': 'Ampulas',
                'K': 'Puntos obscuros',
                'L': 'Defectos totales de proceso',
                'M': 'Hojuelas dobladas'
            }

        # Procesar registros
        data_rows = []
        
        for registro in registros:
            row = {
                'ID': registro.id,
                'Fecha': registro.fecha.strftime('%Y-%m-%d'),
                'Turno': registro.turno,
                'Hora Bloque': f"{registro.hora_bloque}:00",
                'Hora Registro': registro.hora.strftime('%H:%M:%S') if registro.hora else '',
                'Hora Muestreo': registro.hora_muestreo.strftime('%H:%M:%S') if registro.hora_muestreo else '',
                'Producto': registro.producto or '',
                'Sensorial Apariencia': registro.sensorial_apariencia or '',
                'Sensorial Textura': registro.sensorial_textura or '',
                'Sensorial Sabor': registro.sensorial_sabor or '',
                'Observaciones': registro.observaciones or '',
                'Creado por': registro.creator.username if registro.creator else '',
                'Fecha Creacion': registro.created_at.strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Agregar datos de atributos
            if registro.data:
                try:
                    data_json = json.loads(registro.data)
                    for codigo, nombre in atributos_map.items():
                        row[f'{nombre} ({codigo})'] = data_json.get(codigo, 0)
                except json.JSONDecodeError:
                    for codigo, nombre in atributos_map.items():
                        row[f'{nombre} ({codigo})'] = 0
            else:
                for codigo, nombre in atributos_map.items():
                    row[f'{nombre} ({codigo})'] = 0
            
            
            # Agregar columnas específicas de EXTRUIDOS (Registro cada 4 Horas)
            if categoria == 'EXTRUIDOS':
                row['Registro 4H Aplica'] = 'Sí' if (hasattr(registro, 'registro_4horas_aplica') and registro.registro_4horas_aplica) else 'No'
                row['Humedad Cereal Trompo (%)'] = registro.extrusor_humedad_cereal if (hasattr(registro, 'extrusor_humedad_cereal') and registro.extrusor_humedad_cereal) else ''
                row['Tiempo Residencia Freidor (seg)'] = registro.freidor_tiempo_residencia if (hasattr(registro, 'freidor_tiempo_residencia') and registro.freidor_tiempo_residencia) else ''
                row['Temperatura Freidor (°C)'] = registro.freidor_temperatura if (hasattr(registro, 'freidor_temperatura') and registro.freidor_temperatura) else ''
                row['Temperatura Slurry Marmitas (°C)'] = registro.sazonado_temp_slurry if (hasattr(registro, 'sazonado_temp_slurry') and registro.sazonado_temp_slurry) else ''
            
            data_rows.append(row)
        
        return data_rows
    
    # Ruta para descargar Excel PAE
    @app.route('/pae/<category>/export_excel', methods=['GET'])
    @login_required
    def export_pae_excel(category):
        """Exportar datos PAE a Excel"""
        try:
            # Importar openpyxl para generar Excel
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from flask import send_file
            import io
            
            # Verificar categoría válida
            if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
                flash('Categoría no válida', 'danger')
                return redirect(url_for('pae_dashboard', category=category))
            
            # Obtener parámetros de filtros
            fecha_inicio = request.args.get('fecha_inicio')
            fecha_fin = request.args.get('fecha_fin') 
            producto = request.args.get('producto')
            tipo_reporte = request.args.get('tipo', 'detallado')  # 'resumen' o 'detallado'
            
            # Construir consulta
            query = PAERegistro.query.filter_by(categoria=category)
            
            # Aplicar filtros de fecha si existen
            if fecha_inicio:
                query = query.filter(PAERegistro.fecha >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date())
            if fecha_fin:
                query = query.filter(PAERegistro.fecha <= datetime.strptime(fecha_fin, '%Y-%m-%d').date())
            if producto and producto != 'todos':
                query = query.filter(PAERegistro.producto == producto)
                
            # Obtener registros
            registros = query.order_by(PAERegistro.fecha.desc(), PAERegistro.hora_bloque.desc()).all()
            
            if not registros:
                flash('No hay datos para exportar con los filtros seleccionados', 'warning')
                return redirect(url_for('pae_dashboard', category=category))
            
            # Crear workbook
            wb = Workbook()
            
            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            if tipo_reporte == 'resumen':
                # Generar reporte de resumen
                datos_resumen = crear_resumen_pae_excel(registros, category)
                ws = wb.active
                ws.title = "Resumen PAE"
                
                # Headers para resumen
                headers = ['Atributo', 'Codigo', 'Total', 'Cantidad Registros', 'Promedio', 'Maximo', 'Minimo']
                
                # Escribir headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Escribir datos
                for row_idx, item in enumerate(datos_resumen, 2):
                    ws.cell(row=row_idx, column=1, value=item['Atributo'])
                    ws.cell(row=row_idx, column=2, value=item['Codigo'])
                    ws.cell(row=row_idx, column=3, value=item['Total'])
                    ws.cell(row=row_idx, column=4, value=item['Cantidad Registros'])
                    ws.cell(row=row_idx, column=5, value=item['Promedio'])
                    ws.cell(row=row_idx, column=6, value=item['Maximo'])
                    ws.cell(row=row_idx, column=7, value=item['Minimo'])
                    
            else:
                # Generar reporte detallado
                datos_detallados = crear_detallado_pae_excel(registros, category)
                ws = wb.active
                ws.title = "Datos Detallados PAE"
                
                if datos_detallados:
                    # Headers dinámicos basados en el primer registro
                    headers = list(datos_detallados[0].keys())
                    
                    # Escribir headers
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    # Escribir datos
                    for row_idx, item in enumerate(datos_detallados, 2):
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(row=row_idx, column=col_idx, value=item.get(header, ''))
            
            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generar nombre del archivo
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"pae_{category.lower()}_{timestamp}.xlsx"
            
            # Guardar en memoria
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError:
            flash('Error: La biblioteca openpyxl no está instalada. Contacte al administrador.', 'danger')
            return redirect(url_for('pae_dashboard', category=category))
        except Exception as e:
            flash(f'Error al generar el archivo Excel: {str(e)}', 'danger')
            return redirect(url_for('pae_dashboard', category=category))

    # Endpoint para descarga Excel de fisicoquímicos - SIN COLORES
        
    @login_required
    def obtener_rangos(categoria, producto):
                rangos_base = {
                    'TORTILLA': {
                        'default': {
                            'humedad_base': {'min': 0.9, 'max': 1.3, 'warning_low': 0.8, 'warning_high': 1.4},
                            'aceite_base': {'min': 22, 'max': 24, 'warning_low': 21, 'warning_high': 25},
                            'aceite_pt': {'min': 22.98, 'max': 25.98, 'warning_low': 21.98, 'warning_high': 26.98},
                            'humedad_pt': {'min': 0.94, 'max': 1.44, 'warning_low': 1.18, 'warning_high': 2.18},
                            'sal_pt': {'min': 1.38, 'max': 1.98, 'warning_low': 1.18, 'warning_high':2.18}
                        },
                        'TOSTITOS SALSA VERDE': {
                            'humedad_base': {'min': 0.9, 'max': 1.3, 'warning_low': 0.8, 'warning_high': 1.4},
                            'aceite_base': {'min': 22.00, 'max': 24.00, 'warning_low': 21, 'warning_high': 25},
                            'aceite_pt': {'min': 23.14, 'max': 26.14, 'warning_low': 22.14, 'warning_high': 27.14},
                            'humedad_pt': {'min': 1.03, 'max': 1.63, 'warning_low': 0.93, 'warning_high': 1.73},
                            'sal_pt': {'min': 0.97, 'max': 1.57, 'warning_low': 0.67, 'warning_high': 1.87}
                        },
                        'TOSTITOS FH': {
                            'humedad_base': {'min': 0.9, 'max': 1.3, 'warning_low': 0.8, 'warning_high': 1.4},
                            'aceite_base': {'min': 23.14, 'max': 26.14, 'warning_low': 22.14, 'warning_high': 27.15},
                            'aceite_pt': {'min': 23.14, 'max': 26.14, 'warning_low': 22.14, 'warning_high': 27.14},
                            'humedad_pt': {'min': 1.03, 'max': 1.63, 'warning_low': 0.93, 'warning_high': 1.73},
                            'sal_pt': {'min': 0.97, 'max': 1.57, 'warning_low': 0.67, 'warning_high': 1.87}
                        },
                        'DORITOS PIZZEROLA': {
                            'humedad_base': {'min': 1, 'max': 1.2, 'warning_low': 0.9, 'warning_high': 1.3},
                            'aceite_base': {'min': 20, 'max': 23, 'warning_low': 19.00, 'warning_high': 24},
                            'aceite_pt': {'min': 22.83, 'max': 25.83, 'warning_low': 21.83, 'warning_high': 26.83},
                            'humedad_pt': {'min': 0.99, 'max': 1.49, 'warning_low': 0.89, 'warning_high': 1.59},
                            'sal_pt': {'min': 1.10, 'max': 1.7, 'warning_low': 0.9, 'warning_high': 1.9}
                        },
                        'DORITOS FH': {
                            'humedad_base': {'min': 1, 'max': 1.2, 'warning_low': 0.9, 'warning_high': 1.3},
                            'aceite_base': {'min': 20, 'max': 23, 'warning_low': 19.0, 'warning_high': 24},
                            'aceite_pt': {'min': 22.71, 'max': 25.71, 'warning_low': 21.71, 'warning_high': 26.71},
                            'humedad_pt': {'min': 1.12, 'max': 1.72, 'warning_low': 1.07, 'warning_high': 1.77},
                            'sal_pt': {'min': 1.31, 'max': 1.91, 'warning_low': 1.11, 'warning_high': 2.11}
                        },
                        'RANCHERITOS': {
                            'humedad_base': {'min': 0.8, 'max': 1.40, 'warning_low': 0.6, 'warning_high': 1.6},
                            'aceite_base': {'min': 21.35, 'max': 22.75, 'warning_low': 20.25, 'warning_high': 23.75},
                            'aceite_pt': {'min': 22.01, 'max': 22.75, 'warning_low': 20.25, 'warning_high': 23.75},
                            'humedad_pt': {'min': 0.94, 'max': 1.44, 'warning_low': 0.84, 'warning_high': 1.54},
                            'sal_pt': {'min': 1.38, 'max': 1.98, 'warning_low': 1.18, 'warning_high': 2.18}
                        }

                    },
                    'EXTRUIDOS': {
                        'default': {
                            'humedad_base': {'min': 0.7, 'max': 1.7, 'warning_low': 0.6, 'warning_high': 1.8},
                            'aceite_base': {'min': 21.7, 'max': 27.7, 'warning_low': 20.7, 'warning_high': 28.7},
                            'aceite_pt': {'min': 32.46, 'max': 38.46, 'warning_low': 31.46, 'warning_high': 39.46},
                            'humedad_pt': {'min': 0.5, 'max':1.9, 'warning_low': 0.4999999, 'warning_high': 2.10},
                            'sal_pt': {'min': 0.95, 'max': 1.55, 'warning_low': 0.85, 'warning_high': 1.65}
                        },
                        'CHEETOS XTRA FLAMIN HOT': {
                            'humedad_base': {'min': 0.7, 'max': 1.7, 'warning_low': 0.6, 'warning_high': 1.8},
                            'aceite_base': {'min': 21.7, 'max': 27.7, 'warning_low': 20.7, 'warning_high': 28.7},
                            'aceite_pt': {'min': 29.52, 'max': 35.52, 'warning_low': 28.51, 'warning_high': 36.01},
                            'humedad_pt': {'min': 0.47, 'max': 1.67, 'warning_low': 0.47, 'warning_high': 2.07},
                            'sal_pt': {'min': 1.4, 'max': 1.8, 'warning_low': 1.19, 'warning_high': 2.01}
                        },
                        'CHEETOS JALAQUEÑO': {
                            'humedad_base': {'min': 0.7, 'max': 1.7, 'warning_low': 0.60, 'warning_high': 1.80},
                            'aceite_base': {'min': 21.7, 'max': 27.7, 'warning_low': 20.70, 'warning_high': 28.70},
                            'aceite_pt': {'min': 31.64, 'max': 37.64, 'warning_low': 29.64, 'warning_high': 39.64},
                            'humedad_pt': {'min': 0.5, 'max':1.9, 'warning_low': 0.499, 'warning_high': 2.10},
                            'sal_pt': {'min': 1.06, 'max': 1.66, 'warning_low': 0.95, 'warning_high': 1.77}
                        },
                        'CHEETOS XTRA FH NUEVO': {
                            'humedad_base': {'min': 0.7, 'max': 1.7, 'warning_low': 0.6, 'warning_high': 1.8},
                            'aceite_base': {'min': 21.7, 'max': 27.7, 'warning_low': 20.7, 'warning_high': 28.7},
                            'aceite_pt': {'min': 29.35, 'max': 35.35, 'warning_low': 27.35, 'warning_high': 37.35},
                            'humedad_pt': {'min': 0.5, 'max': 1.9,  },
                            'sal_pt': {'min': 1.16, 'max': 1.76, 'warning_low': 0.86, 'warning_high': 2.07}
                        }   
                    },
                    'PAPA': {
                        'default': {
                            'humedad_base': {'min': 1.35, 'max': 1.65, 'warning_low': 1.20, 'warning_high': 1.8},
                            'aceite_base': {'min': 31, 'max': 35, 'warning_low': 30, 'warning_high': 36},
                            'aceite_pt': {'min': 0, 'max': 0, 'warning_low': 0, 'warning_high': 0},
                            'humedad_pt': {'min': 1.35, 'max':1.8, 'warning_low': 1.20, 'warning_high': 2},
                            'sal_pt': {'min': 0.55, 'max': 0.85, 'warning_low': 0.45, 'warning_high': 0.95}
                        }
                    }
                }
                categoria_rangos = rangos_base.get(categoria, rangos_base['EXTRUIDOS'])
                return categoria_rangos.get(producto, categoria_rangos['default'])
            
    @app.route('/analisis_fisicoquimicos/descargar-excel')
    @login_required
    def descargar_excel_fisicoquimicos():
        """Descarga Excel con campos PT - con colores según rangos por producto"""
        import pandas as pd
        import tempfile
        import os
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        categoria = request.args.get('categoria', 'TORTILLA')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        turno = request.args.get('turno', 'all')
        producto = request.args.get('producto', 'all')

        # =====================================================
        # RANGOS POR CATEGORÍA Y PRODUCTO (igual que el JS)
        # =====================================================
        RANGOS_FISICOQUIMICOS = {
            'EXTRUIDOS': {
                'default': {
                    'humedad_base': {'verde': (0.7, 1.7), 'amarillo': [(0.60, 0.69), (1.71, 1.80)]},
                    'aceite_base': {'verde': (21.7, 27.7), 'amarillo': [(20.7, 21.69), (27.71, 28.7)]},
                    'humedad_pt': {'verde': (0.5, 1.9), 'amarillo': [(1.91, 2.10)]},
                    'aceite_pt': {'verde': (32.46, 38.46), 'amarillo': [(31.46, 32.45), (38.47, 39.46)]},
                    'sal_pt': {'verde': (0.95, 1.55), 'amarillo': [(0.85, 0.94), (1.56, 1.65)]}
                },
                'CHEETOS TORCIDITOS': {
                    'humedad_base': {'verde': (0.7, 1.7), 'amarillo': [(0.60, 0.69), (1.71, 1.80)]},
                    'aceite_base': {'verde': (21.7, 27.7), 'amarillo': [(20.7, 21.69), (27.71, 28.7)]},
                    'humedad_pt': {'verde': (0.5, 1.9), 'amarillo': [(1.91, 2.10)]},
                    'aceite_pt': {'verde': (32.46, 38.46), 'amarillo': [(31.46, 32.45), (38.47, 39.46)]},
                    'sal_pt': {'verde': (0.95, 1.55), 'amarillo': [(0.85, 0.94), (1.56, 1.65)]}
                },
                'CHEETOS XTRA FLAMIN HOT': {
                    'humedad_base': {'verde': (0.7, 1.7), 'amarillo': [(0.60, 0.69), (1.71, 1.80)]},
                    'aceite_base': {'verde': (21.7, 27.7), 'amarillo': [(20.7, 21.69), (27.71, 28.7)]},
                    'humedad_pt': {'verde': (0.47, 1.67), 'amarillo': [(1.68, 2.07)]},
                    'aceite_pt': {'verde': (29.52, 35.52), 'amarillo': [(28.51, 29.51), (35.53, 36.01)]},
                    'sal_pt': {'verde': (1.40, 1.80), 'amarillo': [(1.19, 1.39), (1.81, 2.01)]}
                },
                'CHEETOS JALAQUEÑO': {
                    'humedad_base': {'verde': (0.7, 1.7), 'amarillo': [(0.60, 0.69), (1.71, 1.80)]},
                    'aceite_base': {'verde': (21.7, 27.7), 'amarillo': [(20.70, 21.69), (27.71, 28.70)]},
                    'humedad_pt': {'verde': (0.5, 1.9), 'amarillo': [(1.91, 2.10)]},
                    'aceite_pt': {'verde': (31.64, 37.64), 'amarillo': [(29.64, 31.63), (37.65, 39.64)]},
                    'sal_pt': {'verde': (1.06, 1.66), 'amarillo': [(0.95, 1.05), (1.67, 1.77)]}
                },
                'CHEETOS XTRA FH NUEVO': {
                    'humedad_base': {'verde': (0.7, 1.7), 'amarillo': [(0.60, 0.69), (1.71, 1.80)]},
                    'aceite_base': {'verde': (21.7, 27.7), 'amarillo': [(20.7, 21.69), (27.71, 28.7)]},
                    'humedad_pt': {'verde': (0.5, 1.9), 'amarillo': [(1.91, 2.10)]},
                    'aceite_pt': {'verde': (29.35, 35.35), 'amarillo': [(27.35, 29.34), (35.36, 37.35)]},
                    'sal_pt': {'verde': (1.16, 1.76), 'amarillo': [(0.86, 1.15), (1.77, 2.07)]}
                }
            },
            'TORTILLA': {
                'default': {
                    'humedad_base': {'verde': (1.00, 1.20), 'amarillo': [(0.80, 0.99), (1.21, 1.30)]},
                    'aceite_base': {'verde': (20.00, 23.00), 'amarillo': [(21.00, 21.99), (23.01, 24.00)]},
                    'humedad_pt': {'verde': (0.78, 1.58), 'amarillo': [(0.68, 0.77), (1.59, 1.68)]},
                    'aceite_pt': {'verde': (23.45, 26.45), 'amarillo': [(22.45, 23.44), (26.46, 27.45)]},
                    'sal_pt': {'verde': (0.90, 1.50), 'amarillo': [(0.80, 0.89), (1.51, 1.60)]}
                },
                'DORITOS': {
                    'humedad_base': {'verde': (1.00, 1.20), 'amarillo': [(0.90, 0.99), (1.21, 1.30)]},
                    'aceite_base': {'verde': (20.00, 23.00), 'amarillo': [(19.00, 19.99), (23.01, 24.00)]},
                    'humedad_pt': {'verde': (0.78, 1.58), 'amarillo': [(0.63, 0.77), (1.59, 1.73)]},
                    'aceite_pt': {'verde': (23.45, 26.45), 'amarillo': [(22.45, 23.44), (26.46, 27.45)]},
                    'sal_pt': {'verde': (0.90, 1.50), 'amarillo': [(0.70, 0.89), (1.51, 1.70)]}
                },
                'DORITOS INCÓGNITA': {
                    'humedad_base': {'verde': (1.00, 1.20), 'amarillo': [(0.90, 0.99), (1.21, 1.30)]},
                    'aceite_base': {'verde': (20.00, 23.00), 'amarillo': [(19.00, 19.99), (23.01, 24.00)]},
                    'humedad_pt': {'verde': (1.02, 1.62), 'amarillo': [(0.97, 1.01), (1.63, 1.67)]},
                    'aceite_pt': {'verde': (22.35, 25.36), 'amarillo': [(21.35, 22.34), (25.36, 26.35)]},
                    'sal_pt': {'verde': (0.72, 1.32), 'amarillo': [(0.52, 0.71), (1.33, 1.52)]}
                },
                'DORITOS PIZZEROLA': {
                    'humedad_base': {'verde': (1.00, 1.20), 'amarillo': [(0.90, 0.99), (1.21, 1.30)]},
                    'aceite_base': {'verde': (20.00, 23.00), 'amarillo': [(19, 19.99), (23.01, 24.00)]},
                    'humedad_pt': {'verde': (0.99, 1.49), 'amarillo': [(0.89, 0.98), (1.50, 1.59)]},
                    'aceite_pt': {'verde': (22.83, 25.83), 'amarillo': [(21.83, 22.82), (25.84, 26.83)]},
                    'sal_pt': {'verde': (1.10, 1.70), 'amarillo': [(0.90, 1.09), (1.71, 1.90)]}
                },
                'DORITOS FH': {
                    'humedad_base': {'verde': (1.12, 1.72), 'amarillo': [(1.07, 1.11), (1.73, 1.77)]},
                    'aceite_base': {'verde': (20, 23), 'amarillo': [(19.0, 19.99), (23.01, 24)]},
                    'humedad_pt': {'verde': (1.12, 1.72), 'amarillo': [(1.07, 1.11), (1.73, 1.77)]},
                    'aceite_pt': {'verde': (22.71, 25.71), 'amarillo': [(21.83, 22.82), (25.84, 26.83)]},
                    'sal_pt': {'verde': (1.31, 1.91), 'amarillo': [(1.11, 1.3), (1.92, 2.11)]}
                },
                'RANCHERITOS': {
                    'humedad_base': {'verde': (0.8, 1.40), 'amarillo': [(0.60, 0.79), (1.41, 1.60)]},
                    'aceite_base': {'verde': (21.25, 22.75), 'amarillo': [(20.25, 21.24), (22.76, 23.75)]},
                    'humedad_pt': {'verde': (0.94, 1.44), 'amarillo': [(0.84, 0.93), (1.45, 1.54)]},
                    'aceite_pt': {'verde': (22.01, 25.01), 'amarillo': [(21.01, 22.00), (25.02, 26.01)]},
                    'sal_pt': {'verde': (1.38, 1.98), 'amarillo': [(1.18, 1.37), (1.99, 2.18)]}
                },
                'TOSTITOS SALSA VERDE': {
                    'humedad_base': {'verde': (0.90, 1.30), 'amarillo': [(0.80, 0.89), (1.31, 1.40)]},
                    'aceite_base': {'verde': (22.00, 24.00), 'amarillo': [(21.00, 21.99), (24.01, 25.00)]},
                    'humedad_pt': {'verde': (1.03, 1.63), 'amarillo': [(0.93, 1.02), (1.64, 1.73)]},
                    'aceite_pt': {'verde': (23.14, 26.14), 'amarillo': [(22.14, 23.13), (26.15, 27.14)]},
                    'sal_pt': {'verde': (0.97, 1.57), 'amarillo': [(0.67, 0.96), (1.58, 1.87)]}
                },
                'TOSTITOS FH': {
                    'humedad_base': {'verde': (0.90, 1.30), 'amarillo': [(0.80, 0.89), (1.31, 1.40)]},
                    'aceite_base': {'verde': (22.00, 24.00), 'amarillo': [(21.00, 21.99), (24.01, 25.00)]},
                    'humedad_pt': {'verde': (0.94, 1.44), 'amarillo': [(0.84, 0.93), (1.45, 1.54)]},
                    'aceite_pt': {'verde': (22.98, 25.98), 'amarillo': [(21.98, 22.97), (25.99, 26.98)]},
                    'sal_pt': {'verde': (1.38, 1.98), 'amarillo': [(1.18, 1.37), (1.99, 2.18)]}
                }
            },
            'PAPA': {
                'default': {
                    'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.8)]},
                    'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                    'humedad_pt': {'verde': (1.35, 1.8), 'amarillo': [(1.20, 2)]},
                    'aceite_pt': {'verde': (0, 0), 'amarillo': []},
                    'sal_pt': {'verde': (0.55, 0.85), 'amarillo': [(0.45, 0.54), (0.86, 0.95)]},
                    'cloruros_base': {'verde': (0, 1), 'amarillo': []}
                },
                'PAPA SAL': {
                    'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                    'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                    'humedad_pt': {'verde': (1.35, 1.8), 'amarillo': [(1.20, 2)]},
                    'aceite_pt': {'verde': (0, 0), 'amarillo': []},
                    'sal_pt': {'verde': (0.55, 0.85), 'amarillo': [(0.45, 0.54), (0.86, 0.95)]},
                    'cloruros_base': {'verde': (0, 1), 'amarillo': []}
                },
                'SABRITAS LIMON': {
                    'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                    'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                    'humedad_pt': {'verde': (0, 0), 'amarillo': []},
                    'aceite_pt': {'verde': (0, 0), 'amarillo': []},
                    'sal_pt': {'verde': (1.23, 1.50), 'amarillo': [(1.10, 1.22), (1.51, 1.63)]},
                    'cloruros_base': {'verde': (0, 100), 'amarillo': []}
                },
                'RUFFLES SAL': {
                    'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                    'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                    'humedad_pt': {'verde': (1.35, 1.8), 'amarillo': [(1.20, 2)]},
                    'aceite_pt': {'verde': (0, 0), 'amarillo': []},
                    'sal_pt': {'verde': (0.55, 0.85), 'amarillo': [(0.45, 0.54), (0.86, 0.95)]},
                    'cloruros_base': {'verde': (0, 1), 'amarillo': []}
                },
                'RUFFLES QUESO': {
                    'humedad_base': {'verde': (1.20, 1.5), 'amarillo': [(1.05, 1.19), (1.51, 1.65)]},
                    'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                    'humedad_pt': {'verde': (1.35, 1.8), 'amarillo': [(1.20, 2)]},
                    'aceite_pt': {'verde': (0, 0), 'amarillo': []},
                    'sal_pt': {'verde': (1.24, 1.54), 'amarillo': [(1.19, 1.23), (1.55, 1.59)]},
                    'cloruros_base': {'verde': (0, 1), 'amarillo': []}
                },
                'SABRITAS XTRA FH': {
                    'humedad_base': {'verde': (1.35, 1.65), 'amarillo': [(1.20, 1.34), (1.66, 1.80)]},
                    'aceite_base': {'verde': (31, 35), 'amarillo': [(30, 30.9), (35.1, 36)]},
                    'humedad_pt': {'verde': (1.41, 1.71), 'amarillo': [(1.21, 1.40), (1.70, 1.91)]},
                    'aceite_pt': {'verde': (32.21, 32.51), 'amarillo': [(32.1, 32.2), (32.61, 32.71)]},
                    'sal_pt': {'verde': (1.58, 1.88), 'amarillo': [(1.38, 1.57), (1.89, 2.08)]},
                    'cloruros_base': {'verde': (0, 1), 'amarillo': []}
                }
            }
        }

        # Colores para Excel
        VERDE_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        AMARILLO_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        ROJO_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        HEADER_FILL = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        HEADER_FONT = Font(color='FFFFFF', bold=True)

        def obtener_rangos_producto(cat, prod):
            """Obtiene los rangos para un producto específico"""
            cat_rangos = RANGOS_FISICOQUIMICOS.get(cat, RANGOS_FISICOQUIMICOS['EXTRUIDOS'])
            return cat_rangos.get(prod, cat_rangos.get('default', {}))

        def determinar_color(valor, tipo_campo, cat, prod):
            """Determina el color según el valor y los rangos"""
            if valor is None or valor == '':
                return None
            try:
                valor_num = float(valor)
            except (ValueError, TypeError):
                return None

            rangos = obtener_rangos_producto(cat, prod)
            rango_campo = rangos.get(tipo_campo)

            if not rango_campo:
                return None

            verde = rango_campo.get('verde', (0, 0))
            amarillo = rango_campo.get('amarillo', [])

            # Verificar verde
            if verde[0] <= valor_num <= verde[1]:
                return VERDE_FILL

            # Verificar amarillo
            for rango_am in amarillo:
                if len(rango_am) >= 2 and rango_am[0] <= valor_num <= rango_am[1]:
                    return AMARILLO_FILL

            # Si no está en verde ni amarillo, es rojo
            return ROJO_FILL

        # Query
        query = AnalisisCalidad.query.filter_by(categoria=categoria)

        if fecha_inicio:
            try:
                query = query.filter(AnalisisCalidad.fecha >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date())
            except:
                pass
        if fecha_fin:
            try:
                query = query.filter(AnalisisCalidad.fecha <= datetime.strptime(fecha_fin, '%Y-%m-%d').date())
            except:
                pass
        if turno != 'all':
            query = query.filter(AnalisisCalidad.turno == turno)
        if producto != 'all':
            query = query.filter(AnalisisCalidad.producto == producto)

        registros = query.order_by(AnalisisCalidad.fecha.desc()).all()

        # Crear datos con TODOS los campos PT
        data = []
        for r in registros:
            row = {
                'Fecha': r.fecha.strftime('%d/%m/%Y') if r.fecha else '',
                'Turno': r.turno or '',
                'Producto': r.producto or '',
                'Horario': r.horario or '',
                'Folio': r.folio or '',
                'Humedad_Base': r.humedad_base_frita or '',
                'Aceite_Base': r.aceite_base_frita or '',
                'Aceite_PT_General': r.aceite_pt_producto_terminado or '',
                'Humedad_PT_General': r.humedad_pt_producto_terminado or '',
                'Sal_PT_General': r.sal_pt_producto_terminado or '',
                'T1_Aceite': r.tanque1_aceite_pt or '',
                'T1_Humedad': r.tanque1_humedad_pt or '',
                'T1_Sal_Titulador': r.tanque1_sal_titulador or '',
                'T1_Sal_PT': r.tanque1_sal_pt or '',
                'T2_Aceite': r.tanque2_aceite_pt or '',
                'T2_Humedad': r.tanque2_humedad_pt or '',
                'T2_Sal_Titulador': r.tanque2_sal_titulador or '',
                'T2_Sal_PT': r.tanque2_sal_pt or '',
                'T3_Aceite': r.tanque3_aceite_pt or '',
                'T3_Humedad': r.tanque3_humedad_pt or '',
                'T3_Sal_Titulador': r.tanque3_sal_titulador or '',
                'T3_Sal_PT': r.tanque3_sal_pt or '',
            }

            # Agregar Cloruros Base solo para PAPA (antes de Observaciones)
            if categoria == 'PAPA':
                row['Cloruros_Base'] = r.cloruros_base or ''

            row['Observaciones'] = r.observaciones or ''
            data.append(row)

        df = pd.DataFrame(data)

        # Mapeo de columnas a tipos de campo para colorear
        # Columnas: Fecha(1), Turno(2), Producto(3), Horario(4), Folio(5),
        #           Humedad_Base(6), Aceite_Base(7),
        #           Aceite_PT_General(8), Humedad_PT_General(9), Sal_PT_General(10),
        #           T1_Aceite(11), T1_Humedad(12), T1_Sal_Titulador(13), T1_Sal_PT(14),
        #           T2_Aceite(15), T2_Humedad(16), T2_Sal_Titulador(17), T2_Sal_PT(18),
        #           T3_Aceite(19), T3_Humedad(20), T3_Sal_Titulador(21), T3_Sal_PT(22),
        #           [Cloruros_Base(23) solo PAPA], Observaciones
        COLUMNAS_COLOR = {
            6: 'humedad_base',      # Humedad_Base
            7: 'aceite_base',       # Aceite_Base
            8: 'aceite_pt',         # Aceite_PT_General
            9: 'humedad_pt',        # Humedad_PT_General
            10: 'sal_pt',           # Sal_PT_General
            11: 'aceite_pt',        # T1_Aceite
            12: 'humedad_pt',       # T1_Humedad
            # 13: T1_Sal_Titulador - sin color (es entrada)
            14: 'sal_pt',           # T1_Sal_PT
            15: 'aceite_pt',        # T2_Aceite
            16: 'humedad_pt',       # T2_Humedad
            # 17: T2_Sal_Titulador - sin color (es entrada)
            18: 'sal_pt',           # T2_Sal_PT
            19: 'aceite_pt',        # T3_Aceite
            20: 'humedad_pt',       # T3_Humedad
            # 21: T3_Sal_Titulador - sin color (es entrada)
            22: 'sal_pt',           # T3_Sal_PT
        }

        # Para PAPA, agregar Cloruros_Base (columna 23)
        if categoria == 'PAPA':
            COLUMNAS_COLOR[23] = 'cloruros_base'

        # USAR ARCHIVO TEMPORAL (evita corrupción)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            try:
                df.to_excel(tmp.name, index=False, engine='openpyxl')
                tmp.flush()

                # Abrir el archivo para aplicar colores
                wb = load_workbook(tmp.name)
                ws = wb.active

                # Aplicar formato al header
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Aplicar colores a cada fila de datos
                for row_idx in range(2, ws.max_row + 1):
                    # Obtener el producto de esta fila (columna 3 = Producto)
                    producto_fila = ws.cell(row=row_idx, column=3).value or 'default'

                    # Aplicar colores a las columnas configuradas
                    for col_idx, tipo_campo in COLUMNAS_COLOR.items():
                        if col_idx <= ws.max_column:
                            celda = ws.cell(row=row_idx, column=col_idx)
                            valor = celda.value

                            color_fill = determinar_color(valor, tipo_campo, categoria, producto_fila)
                            if color_fill:
                                celda.fill = color_fill

                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Guardar cambios
                wb.save(tmp.name)
                wb.close()

                filename = f"fisicoquimicos_{categoria.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

                return send_file(
                    tmp.name,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )

            except Exception as e:
                flash(f'Error generando Excel: {str(e)}', 'danger')
                return redirect(url_for('list_analisis_fisicoquimicos', category=categoria))
            finally:
                # Limpiar archivo temporal después
                try:
                    os.unlink(tmp.name)
                except:
                    pass

    return app

# Crear y ejecutar la aplicación

if __name__ == "__main__":
    app = create_app()
    app.run(host='0.0.0.0', port=5000, debug=True)