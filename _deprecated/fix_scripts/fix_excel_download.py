import re

# Código corregido para la función de descarga Excel SIN COLORES
excel_function = '''
    # Endpoint para descarga Excel de fisicoquímicos
    @app.route('/analisis_fisicoquimicos/descargar-excel')
    @login_required
    def descargar_excel_fisicoquimicos():
        try:
            # Importar al inicio
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            import io
            
            # Obtener parámetros
            categoria = request.args.get('categoria', 'TORTILLA')
            fecha_inicio = request.args.get('fecha_inicio')
            fecha_fin = request.args.get('fecha_fin')
            turno = request.args.get('turno', 'all')
            producto = request.args.get('producto', 'all')
            incluir_rangos = request.args.get('incluir_rangos', 'false').lower() == 'true'
            
            # Validaciones básicas
            if categoria not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
                return "Categoría no válida", 400
            
            if not fecha_inicio or not fecha_fin:
                # Usar últimos 30 días por defecto
                fecha_fin_dt = datetime.now().date()
                fecha_inicio_dt = fecha_fin_dt - timedelta(days=30)
            else:
                try:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                except ValueError:
                    return "Formato de fecha inválido", 400
            
            # Construir consulta
            query = AnalisisCalidad.query.filter(
                AnalisisCalidad.categoria == categoria,
                AnalisisCalidad.fecha >= fecha_inicio_dt,
                AnalisisCalidad.fecha <= fecha_fin_dt
            )
            
            if turno != 'all':
                query = query.filter(AnalisisCalidad.turno == turno)
            
            if producto != 'all':
                query = query.filter(AnalisisCalidad.producto == producto)
            
            registros = query.order_by(AnalisisCalidad.fecha.desc(), AnalisisCalidad.created_at.desc()).all()
            
            # Crear Excel con openpyxl
            wb = Workbook()
            ws = wb.active
            ws.title = "Analisis_Fisicoquimicos"
            
            # Headers
            headers = ['Fecha', 'Turno', 'Producto', 'Horario',
                      'Humedad Base', 'Aceite Base',
                      'T1 Aceite', 'T1 Humedad', 'T1 Sal',
                      'T2 Aceite', 'T2 Humedad', 'T2 Sal']
            
            if categoria != 'EXTRUIDOS':
                headers.extend(['T3 Aceite', 'T3 Humedad', 'T3 Sal'])
            headers.append('Observaciones')
            
            # Escribir headers con formato básico
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Escribir datos SIN COLORES
            for row_idx, registro in enumerate(registros, 2):
                col_idx = 1
                
                # Fecha
                ws.cell(row=row_idx, column=col_idx, value=registro.fecha.strftime('%d/%m/%Y'))
                col_idx += 1
                
                # Turno
                ws.cell(row=row_idx, column=col_idx, value=registro.turno)
                col_idx += 1
                
                # Producto
                ws.cell(row=row_idx, column=col_idx, value=registro.producto)
                col_idx += 1
                
                # Horario
                ws.cell(row=row_idx, column=col_idx, value=registro.horario)
                col_idx += 1
                
                # Humedad Base SIN color
                ws.cell(row=row_idx, column=col_idx, value=registro.humedad_base_frita)
                col_idx += 1
                
                # Aceite Base SIN color
                ws.cell(row=row_idx, column=col_idx, value=registro.aceite_base_frita)
                col_idx += 1
                
                # Tambor 1 SIN colores
                for campo, valor in [('aceite', registro.tanque1_aceite_pt), 
                                     ('humedad', registro.tanque1_humedad_pt),
                                     ('sal', registro.tanque1_sal_pt)]:
                    ws.cell(row=row_idx, column=col_idx, value=valor)
                    col_idx += 1
                
                # Tambor 2 SIN colores
                for campo, valor in [('aceite', registro.tanque2_aceite_pt),
                                     ('humedad', registro.tanque2_humedad_pt),
                                     ('sal', registro.tanque2_sal_pt)]:
                    ws.cell(row=row_idx, column=col_idx, value=valor)
                    col_idx += 1
                
                # Tambor 3 SIN colores (si aplica)
                if categoria != 'EXTRUIDOS':
                    for campo, valor in [('aceite', registro.tanque3_aceite_pt),
                                         ('humedad', registro.tanque3_humedad_pt),
                                         ('sal', registro.tanque3_sal_pt)]:
                        ws.cell(row=row_idx, column=col_idx, value=valor)
                        col_idx += 1
                
                # Observaciones
                ws.cell(row=row_idx, column=col_idx, value=registro.observaciones or '')
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value or '')) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 40)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generar archivo
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"analisis_fisicoquimicos_{categoria.lower()}_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except ImportError as e:
            flash('Error: Las librerías necesarias no están instaladas. Por favor ejecute: pip install openpyxl', 'danger')
            return redirect(url_for('list_analisis_fisicoquimicos', category=categoria))
        except Exception as e:
            flash(f'Error al generar el archivo Excel: {str(e)}', 'danger')
            return redirect(url_for('list_analisis_fisicoquimicos', category=categoria))
'''

# Leer archivo actual
with open('app.py', 'r', encoding='utf-8') as f:
    content = f.read()

# Buscar y reemplazar la función existente
import re

# Patrón para encontrar la función actual
pattern = r'@app\.route\(\'/analisis_fisicoquimicos/descargar-excel\'\).*?(?=@app\.route|$)'

# Si encuentra la función, reemplazarla
if re.search(pattern, content, re.DOTALL):
    content = re.sub(pattern, excel_function.strip() + '\n\n    ', content, flags=re.DOTALL)
    print("Función encontrada y reemplazada")
else:
    # Si no la encuentra, buscar dónde insertarla (después de list_analisis_fisicoquimicos)
    pattern2 = r'(def list_analisis_fisicoquimicos.*?return render_template.*?\))\s*\n'
    match = re.search(pattern2, content, re.DOTALL)
    if match:
        insert_pos = match.end()
        content = content[:insert_pos] + '\n' + excel_function + '\n' + content[insert_pos:]
        print("Función insertada después de list_analisis_fisicoquimicos")

# Guardar archivo
with open('app.py', 'w', encoding='utf-8') as f:
    f.write(content)

print("Archivo app.py actualizado con la función de descarga Excel SIN COLORES")
