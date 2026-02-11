# Función de Excel completamente nueva y funcional
def create_excel_route_function():
    return '''
    @app.route('/analisis_fisicoquimicos/descargar-excel')
    @login_required
    def descargar_excel_fisico_simple():
        try:
            import tempfile
            import os
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            # Parámetros
            categoria = request.args.get('categoria', 'TORTILLA')
            fecha_inicio = request.args.get('fecha_inicio')
            fecha_fin = request.args.get('fecha_fin')
            turno = request.args.get('turno', 'all')
            producto = request.args.get('producto', 'all')
            
            # Query
            query = AnalisisCalidad.query.filter_by(categoria=categoria)
            
            if fecha_inicio:
                try:
                    query = query.filter(AnalisisCalidad.fecha >= datetime.strptime(fecha_inicio, '%Y-%m-%d').date())
                except:
                    pass
            if fecha_fin:
                try:
                    query = query.filter(AnalisisCalidad.fecha <= datetime.strptime(fecha_fin, '%Y-%m-%d').date())
                except:
                    pass
            if turno != 'all':
                query = query.filter(AnalisisCalidad.turno == turno)
            if producto != 'all':
                query = query.filter(AnalisisCalidad.producto == producto)
            
            registros = query.order_by(AnalisisCalidad.fecha.desc()).all()
            
            # Crear Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos"
            
            # Headers
            headers = ['Fecha', 'Turno', 'Producto', 'Horario', 'Folio', 'Humedad_Base_Frita', 'Aceite_Base_Frita',
                      'T1_Aceite_PT', 'T1_Humedad_PT', 'T1_Sal_PT', 'T2_Aceite_PT', 'T2_Humedad_PT', 'T2_Sal_PT']
            
            if categoria != 'EXTRUIDOS':
                headers.extend(['T3_Aceite_PT', 'T3_Humedad_PT', 'T3_Sal_PT'])
            headers.append('Observaciones')
            
            # Escribir headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            
            # Escribir datos
            for row, registro in enumerate(registros, 2):
                ws.cell(row=row, column=1).value = registro.fecha.strftime('%d/%m/%Y') if registro.fecha else ''
                ws.cell(row=row, column=2).value = registro.turno or ''
                ws.cell(row=row, column=3).value = registro.producto or ''
                ws.cell(row=row, column=4).value = registro.horario or ''
                ws.cell(row=row, column=5).value = registro.folio or ''
                ws.cell(row=row, column=6).value = registro.humedad_base_frita or ''
                ws.cell(row=row, column=7).value = registro.aceite_base_frita or ''
                ws.cell(row=row, column=8).value = registro.tanque1_aceite_pt or ''
                ws.cell(row=row, column=9).value = registro.tanque1_humedad_pt or ''
                ws.cell(row=row, column=10).value = registro.tanque1_sal_pt or ''
                ws.cell(row=row, column=11).value = registro.tanque2_aceite_pt or ''
                ws.cell(row=row, column=12).value = registro.tanque2_humedad_pt or ''
                ws.cell(row=row, column=13).value = registro.tanque2_sal_pt or ''
                
                if categoria != 'EXTRUIDOS':
                    ws.cell(row=row, column=14).value = registro.tanque3_aceite_pt or ''
                    ws.cell(row=row, column=15).value = registro.tanque3_humedad_pt or ''
                    ws.cell(row=row, column=16).value = registro.tanque3_sal_pt or ''
                    obs_col = 17
                else:
                    obs_col = 14
                
                ws.cell(row=row, column=obs_col).value = registro.observaciones or ''
            
            # Ajustar columnas
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15
            
            # Guardar en archivo temporal
            filename = f"analisis_fisicoquimicos_{categoria.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                wb.save(tmp.name)
                
                with open(tmp.name, 'rb') as f:
                    excel_data = f.read()
                
                os.unlink(tmp.name)
            
            # Respuesta
            response = app.response_class(
                excel_data,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                headers={'Content-Disposition': f'attachment; filename="{filename}"'}
            )
            return response
            
        except Exception as e:
            print(f"Error Excel: {e}")
            flash(f'Error: {str(e)}', 'danger')
            return redirect(request.referrer or url_for('index'))
    '''

if __name__ == "__main__":
    print(create_excel_route_function())
