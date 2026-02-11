import re

# Leer app.py actual
with open('app.py', 'r', encoding='utf-8') as f:
    content = f.read()

# Nueva función CON campos PT
nueva_funcion = '''    @app.route('/analisis_fisicoquimicos/descargar-excel')
    @login_required
    def descargar_excel_fisicoquimicos():
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            import io
            
            categoria = request.args.get('categoria', 'TORTILLA')
            fecha_inicio = request.args.get('fecha_inicio')
            fecha_fin = request.args.get('fecha_fin')
            turno = request.args.get('turno', 'all')
            producto = request.args.get('producto', 'all')
            
            if not fecha_inicio or not fecha_fin:
                fecha_fin_dt = datetime.now().date()
                fecha_inicio_dt = fecha_fin_dt - timedelta(days=30)
            else:
                try:
                    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                    fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                except ValueError:
                    return "Formato de fecha inválido", 400
            
            query = AnalisisCalidad.query.filter(
                AnalisisCalidad.categoria == categoria,
                AnalisisCalidad.fecha >= fecha_inicio_dt,
                AnalisisCalidad.fecha <= fecha_fin_dt
            )
            
            if turno != 'all':
                query = query.filter(AnalisisCalidad.turno == turno)
            if producto != 'all':
                query = query.filter(AnalisisCalidad.producto == producto)
            
            registros = query.order_by(AnalisisCalidad.fecha.desc()).all()
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Analisis_Fisicoquimicos"
            
            # Headers CON campos PT
            headers = ['Fecha', 'Turno', 'Producto', 'Horario',
                      'Humedad Base', 'Aceite Base',
                      'Aceite PT General', 'Humedad PT General', 'Sal PT General',
                      'T1 Aceite', 'T1 Humedad', 'T1 Sal',
                      'T2 Aceite', 'T2 Humedad', 'T2 Sal']
            
            if categoria != 'EXTRUIDOS':
                headers.extend(['T3 Aceite', 'T3 Humedad', 'T3 Sal'])
            headers.append('Observaciones')
            
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Datos CON campos PT
            for row_idx, registro in enumerate(registros, 2):
                ws.cell(row=row_idx, column=1, value=registro.fecha.strftime('%d/%m/%Y'))
                ws.cell(row=row_idx, column=2, value=registro.turno)
                ws.cell(row=row_idx, column=3, value=registro.producto)
                ws.cell(row=row_idx, column=4, value=registro.horario)
                ws.cell(row=row_idx, column=5, value=registro.humedad_base_frita)
                ws.cell(row=row_idx, column=6, value=registro.aceite_base_frita)
                
                # *** CAMPOS PT (LOS QUE FALTABAN) ***
                ws.cell(row=row_idx, column=7, value=registro.aceite_pt_producto_terminado)
                ws.cell(row=row_idx, column=8, value=registro.humedad_pt_producto_terminado)
                ws.cell(row=row_idx, column=9, value=registro.sal_pt_producto_terminado)
                
                # Tambores
                ws.cell(row=row_idx, column=10, value=registro.tanque1_aceite_pt)
                ws.cell(row=row_idx, column=11, value=registro.tanque1_humedad_pt)
                ws.cell(row=row_idx, column=12, value=registro.tanque1_sal_pt)
                ws.cell(row=row_idx, column=13, value=registro.tanque2_aceite_pt)
                ws.cell(row=row_idx, column=14, value=registro.tanque2_humedad_pt)
                ws.cell(row=row_idx, column=15, value=registro.tanque2_sal_pt)
                
                if categoria != 'EXTRUIDOS':
                    ws.cell(row=row_idx, column=16, value=registro.tanque3_aceite_pt)
                    ws.cell(row=row_idx, column=17, value=registro.tanque3_humedad_pt)
                    ws.cell(row=row_idx, column=18, value=registro.tanque3_sal_pt)
                    obs_col = 19
                else:
                    obs_col = 16
                
                ws.cell(row=row_idx, column=obs_col, value=registro.observaciones or '')
            
            # Ajustar columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value or '')) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"analisis_fisicoquimicos_{categoria.lower()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('list_analisis_fisicoquimicos', category=categoria))'''

# Buscar y reemplazar función existente
pattern = r'@app\.route\(\'/analisis_fisicoquimicos/descargar-excel\'\).*?(?=\n[ ]{4}@app\.route|\n[ ]{4}def |\nif __name__|$)'

if re.search(pattern, content, re.DOTALL):
    content = re.sub(pattern, nueva_funcion, content, flags=re.DOTALL)
    print("✅ Función reemplazada exitosamente")
else:
    print("❌ No se encontró la función existente")
    # Buscar punto de inserción
    insert_point = content.find("if __name__ == '__main__':")
    if insert_point != -1:
        content = content[:insert_point] + nueva_funcion + "\n\n    " + content[insert_point:]
        print("✅ Función insertada al final")

# Guardar
with open('app.py', 'w', encoding='utf-8') as f:
    f.write(content)

print("🎯 Excel ahora incluye:")
print("   - aceite_pt_producto_terminado")
print("   - humedad_pt_producto_terminado")
print("   - sal_pt_producto_terminado")
