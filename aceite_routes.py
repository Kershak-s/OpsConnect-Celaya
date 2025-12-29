# Rutas para el módulo de Control de Aceite
# Este archivo debe ser importado en app.py

from flask import jsonify
from models import db, AnalisisAceite
from datetime import datetime

def setup_aceite_routes(app):
    from flask import render_template, redirect, url_for, flash, request
    from flask_login import login_required, current_user
    
    @app.route('/aceite/<category>')
    @login_required
    def aceite_dashboard(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))
        
        registros = AnalisisAceite.query.filter_by(categoria=category).order_by(AnalisisAceite.fecha.desc()).all()
        registros_json = [
            {
                "id": r.id,
                "fecha": r.fecha.strftime('%Y-%m-%d'),
                "turno": r.turno,
                "horario": r.horario.strftime('%H:%M'),
                "producto": r.producto,
                "ov": r.ov,
                "agl": r.agl,
                "observaciones": r.observaciones or ""
            }
            for r in registros
        ]
        # Renderizar la plantilla
        return render_template(
            'aceite.html',
            title=f'Análisis de Aceite - {category}',
            category=category,
            registros=registros_json
        )
    
    @app.route('/aceite/<category>/analisis')
    @login_required
    def analisis_aceite(category):
        # Verificar categoría válida
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('index'))

        # Obtener registros (aunque sea vacío)
        registros = AnalisisAceite.query.filter_by(categoria=category).order_by(AnalisisAceite.fecha.desc()).all()
        registros_json = [
            {
                "id": r.id,
                "fecha": r.fecha.strftime('%Y-%m-%d'),
                "turno": r.turno,
                "horario": r.horario.strftime('%H:%M'),
                "producto": r.producto,
                "ov": r.ov,
                "agl": r.agl,
                "observaciones": r.observaciones or ""
            }
            for r in registros
        ]
        # Renderizar la plantilla
        return render_template('aceite.html',
                              title=f'Análisis de Aceite - {category}',
                              category=category,
                              registros=registros_json,
                              active_tab='analisis')
    
    @app.route('/aceite/<category>/create', methods=['POST'])
    @login_required
    def aceite_create(category):
        # Validar categoría
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('aceite_dashboard', category=category))

        # Obtener datos del formulario
        fecha = request.form.get('fecha')
        turno = request.form.get('turno')
        horario = request.form.get('horario')
        producto = request.form.get('producto')
        ov = request.form.get('ov')
        agl = request.form.get('agl')
        observaciones = request.form.get('observaciones', '')

        # Validar campos requeridos
        if not (fecha and turno and horario and producto and ov and agl):
            flash('Todos los campos obligatorios deben ser completados.', 'danger')
            return redirect(url_for('aceite_dashboard', category=category))

        try:
            registro = AnalisisAceite(
                fecha=datetime.strptime(fecha, '%Y-%m-%d').date(),
                turno=turno,
                horario=datetime.strptime(horario, '%H:%M').time(),
                producto=producto,
                ov=float(ov),
                agl=float(agl),
                observaciones=observaciones,
                categoria=category,
                created_by=current_user.id
            )
            db.session.add(registro)
            db.session.commit()
            flash('Registro de análisis de aceite guardado correctamente.', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Error al guardar el registro: ' + str(e), 'danger')

        return redirect(url_for('aceite_dashboard', category=category))
    
    @app.route('/aceite/<category>/edit/<int:id>', methods=['POST'])
    @login_required
    def aceite_edit(category, id):
        # Validar categoría
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            flash('Línea de producción no válida', 'danger')
            return redirect(url_for('aceite_dashboard', category=category))

        registro = AnalisisAceite.query.get_or_404(id)

        # Obtener datos del formulario
        fecha = request.form.get('fecha')
        turno = request.form.get('turno')
        horario = request.form.get('horario')
        producto = request.form.get('producto')
        ov = request.form.get('ov')
        agl = request.form.get('agl')
        observaciones = request.form.get('observaciones', '')

        # Validar campos requeridos
        if not (fecha and turno and horario and producto and ov and agl):
            flash('Todos los campos obligatorios deben ser completados.', 'danger')
            return redirect(url_for('aceite_dashboard', category=category))

        try:
            registro.fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
            registro.turno = turno
            registro.horario = datetime.strptime(horario, '%H:%M').time()
            registro.producto = producto
            registro.ov = float(ov)
            registro.agl = float(agl)
            registro.observaciones = observaciones
            db.session.commit()
            flash('Registro actualizado correctamente.', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Error al actualizar el registro: ' + str(e), 'danger')

        return redirect(url_for('aceite_dashboard', category=category))
    
    @app.route('/aceite/<category>/delete/<int:id>', methods=['POST'])
    @login_required
    def aceite_delete(category, id):
        if category not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
            return jsonify({'success': False, 'error': 'Categoría inválida'}), 400
        registro = AnalisisAceite.query.get_or_404(id)
        try:
            db.session.delete(registro)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    @app.route('/aceite/descargar-excel')
    @login_required
    def descargar_excel_aceite():
        try:
            # Obtener parámetros
            categoria = request.args.get('categoria', 'TORTILLA')
            fecha_inicio = request.args.get('fecha_inicio')
            fecha_fin = request.args.get('fecha_fin')
            turno = request.args.get('turno', 'all')
            producto = request.args.get('producto', 'all')
            incluir_rangos = request.args.get('incluir_rangos', 'false').lower() == 'true'
            
            # Validaciones
            if categoria not in ['EXTRUIDOS', 'TORTILLA', 'PAPA']:
                return "Categoría no válida", 400
            
            if not fecha_inicio or not fecha_fin:
                return "Fechas requeridas", 400
            
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                return "Formato de fecha inválido", 400
            
            # Construir consulta
            query = AnalisisAceite.query.filter(
                AnalisisAceite.categoria == categoria,
                AnalisisAceite.fecha >= fecha_inicio_dt,
                AnalisisAceite.fecha <= fecha_fin_dt
            )
            
            if turno != 'all':
                query = query.filter(AnalisisAceite.turno == turno)
            
            if producto != 'all':
                query = query.filter(AnalisisAceite.producto == producto)
            
            registros = query.order_by(AnalisisAceite.fecha.asc(), AnalisisAceite.created_at.asc()).all()
            
            if not registros:
                return "No hay datos para exportar", 404
            
            # Función para obtener color Excel con rangos por categoría
            def obtener_color_excel(valor, tipo, categoria):
                if valor is None:
                    return None
                try:
                    val = float(valor)
                    if categoria == 'PAPA':
                        if tipo == 'ov':
                            # VO: Verde 0-15, Amarillo 15.1-25, Rojo >25
                            if 0 <= val <= 15:
                                return 'D4EDDA'  # Verde
                            elif 15.1 <= val <= 25:
                                return 'FFF3CD'  # Amarillo
                            else:
                                return 'F8D7DA'  # Rojo
                        elif tipo == 'agl':
                            # AGL: Verde 0-0.25, Amarillo 0.26-0.35, Rojo >0.35
                            if 0 <= val <= 0.25:
                                return 'D4EDDA'  # Verde
                            elif 0.26 <= val <= 0.35:
                                return 'FFF3CD'  # Amarillo
                            else:
                                return 'F8D7DA'  # Rojo
                    else:
                        # Otras categorías (EXTRUIDOS, TORTILLA)
                        if tipo == 'ov':
                            return 'D4EDDA' if 0 <= val <= 50 else 'F8D7DA'
                        elif tipo == 'agl':
                            return 'D4EDDA' if 0 <= val <= 5 else 'F8D7DA'
                except (ValueError, TypeError):
                    return None
                return None
            
            # Crear Excel
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font
                from flask import send_file
                from io import BytesIO
                
                wb = Workbook()
                ws_datos = wb.active
                ws_datos.title = "Datos_Aceite"
                
                # Headers
                headers = ['Fecha', 'Turno', 'Horario', 'Producto', 'OV (%)', 'AGL (%)', 'Observaciones']
                
                # Escribir headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws_datos.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
                
                # Escribir datos con colores
                for row_idx, registro in enumerate(registros, 2):
                    # Datos básicos
                    ws_datos.cell(row=row_idx, column=1, value=registro.fecha.strftime('%d/%m/%Y'))
                    ws_datos.cell(row=row_idx, column=2, value=registro.turno or '')
                    ws_datos.cell(row=row_idx, column=3, value=registro.horario.strftime('%H:%M') if registro.horario else '')
                    ws_datos.cell(row=row_idx, column=4, value=registro.producto or '')
                    
                    # OV con color
                    ov_cell = ws_datos.cell(row=row_idx, column=5, value=registro.ov)
                    ov_color = obtener_color_excel(registro.ov, 'ov', categoria)
                    if ov_color:
                        ov_cell.fill = PatternFill(start_color=ov_color, end_color=ov_color, fill_type="solid")
                    
                    # AGL con color
                    agl_cell = ws_datos.cell(row=row_idx, column=6, value=registro.agl)
                    agl_color = obtener_color_excel(registro.agl, 'agl', categoria)
                    if agl_color:
                        agl_cell.fill = PatternFill(start_color=agl_color, end_color=agl_color, fill_type="solid")
                    
                    # Observaciones
                    ws_datos.cell(row=row_idx, column=7, value=registro.observaciones or '')
                
                # Ajustar anchos
                for column in ws_datos.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 25)
                    ws_datos.column_dimensions[column_letter].width = adjusted_width
                
                # Hoja de rangos si se solicita
                if incluir_rangos:
                    ws_rangos = wb.create_sheet(title="Rangos_Referencia")
                    rangos_headers = ['Parámetro', 'Rango Mínimo', 'Rango Máximo', 'Descripción']
                    
                    for col_idx, header in enumerate(rangos_headers, 1):
                        cell = ws_rangos.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
                    
                    # Datos de rangos según categoría
                    if categoria == 'PAPA':
                        rangos_data = [
                            ['VO Verde (Valor de Oxidación)', '0 ppm', '15 ppm', 'Rango óptimo para valor de oxidación'],
                            ['VO Amarillo (Valor de Oxidación)', '15.1 ppm', '25 ppm', 'Rango aceptable para valor de oxidación'],
                            ['VO Rojo (Valor de Oxidación)', '>25 ppm', '-', 'Fuera del rango aceptable'],
                            ['AGL Verde (Ácidos Grasos Libres)', '0%', '0.25%', 'Rango óptimo para ácidos grasos libres'],
                            ['AGL Amarillo (Ácidos Grasos Libres)', '0.26%', '0.35%', 'Rango aceptable para ácidos grasos libres'],
                            ['AGL Rojo (Ácidos Grasos Libres)', '>0.35%', '-', 'Fuera del rango aceptable']
                        ]
                    else:
                        rangos_data = [
                            ['OV (Valor de Oxidación)', '0%', '50%', 'Rango aceptable para valor de oxidación'],
                            ['AGL (Ácidos Grasos Libres)', '0%', '5%', 'Rango aceptable para ácidos grasos libres']
                        ]
                    
                    for row_idx, (param, min_val, max_val, desc) in enumerate(rangos_data, 2):
                        ws_rangos.cell(row=row_idx, column=1, value=param)
                        ws_rangos.cell(row=row_idx, column=2, value=min_val)
                        ws_rangos.cell(row=row_idx, column=3, value=max_val)
                        ws_rangos.cell(row=row_idx, column=4, value=desc)
                
                # Hoja de información
                ws_info = wb.create_sheet(title="Info_Reporte")
                info_data = [
                    ['Parámetros de Exportación', ''],
                    ['Categoría', categoria],
                    ['Fecha Inicio', fecha_inicio],
                    ['Fecha Fin', fecha_fin],
                    ['Turno', turno if turno != 'all' else 'Todos'],
                    ['Producto', producto if producto != 'all' else 'Todos'],
                    ['Total Registros', len(registros)],
                    ['Fecha Generación', datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
                    ['Usuario', current_user.username],
                    ['', ''],
                    ['Leyenda de Colores', ''],
                    ['Verde', 'Valores dentro del rango óptimo'],
                    ['Amarillo', 'Valores en rango aceptable (solo PAPA)'],
                    ['Rojo', 'Valores fuera del rango aceptable']
                ]
                
                for row_idx, (campo, valor) in enumerate(info_data, 1):
                    ws_info.cell(row=row_idx, column=1, value=campo)
                    ws_info.cell(row=row_idx, column=2, value=valor)
                    if row_idx == 1 or row_idx == 11:
                        ws_info.cell(row=row_idx, column=1).font = Font(bold=True)
                
                # Generar archivo
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"analisis_aceite_{categoria.lower()}_{timestamp}.xlsx"
                
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
                
            except ImportError:
                return "Error: openpyxl no está instalada", 500
            except Exception as e:
                return f"Error al generar Excel: {str(e)}", 500
                
        except Exception as e:
            return f"Error interno: {str(e)}", 500
